from collections import OrderedDict
from datetime import datetime
from pathlib import Path

from django.db.models import Case, CharField, F, Value, When
from django.db.models.functions import ExtractYear
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import (
    extend_schema,
    extend_schema_view,
)
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework import mixins, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny

from apps.common.utils import XlsxRenderer
from apps.country.models import HouseholdSize
from apps.country.rest_filters import RestHouseholdSizeFilterSet
from apps.country.serializers import HouseholdSizeSerializer
from utils.common import client_id


@client_id
class ListOnlyViewSetMixin(mixins.ListModelMixin, viewsets.GenericViewSet):
    def get(self, request, *args, **kwargs):
        return self.list(request, *args, **kwargs)


@extend_schema_view(
    list=extend_schema(
        description=Path("docs/disaster/main-description.md").read_text(),
        responses=HouseholdSizeSerializer(many=True),
        tags=["AHHS"],
    ),
)
class HouseholdSizeViewSet(ListOnlyViewSetMixin):
    filter_class = RestHouseholdSizeFilterSet
    serializer_class = HouseholdSizeSerializer
    filter_backends = (DjangoFilterBackend,)
    renderer_classes = [XlsxRenderer]

    def get_queryset(self):
        return (
            HouseholdSize.objects.filter(is_active=True)
            .annotate(
                reference_year=ExtractYear("created_at"),
                gap_filling_method=Case(
                    When(reference_year__lt=F("year"), then=Value("BACKWARD_FILLING")),
                    When(reference_year__gt=F("year"), then=Value("FORWARD_FILLING")),
                    default=Value("EXACT"),
                    output_field=CharField(),
                ),
            )
            .order_by("country__region")
            .all()
        )

    @extend_schema(
        description=Path("docs/ahhs/xlsx-export-description.md").read_text(),
        responses={
            (200, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"): OpenApiTypes.BINARY,
        },
        filters=True,
        tags=["AHHS"],
    )
    @action(
        detail=False,
        methods=["get"],
        url_path="householdsize-export",
        permission_classes=[AllowAny],
        pagination_class=None,
    )

    # FIXME: consider async export
    def export_householdsize(self, request):
        headers = OrderedDict(
            country__region__name="Region Name",
            country__idmc_short_name="Country",
            year="Year",
            size="AHHS",
            reference_year="Reference Year",
            data_source_category="Data Source Category",
            source="Source",
            source_link="Source Link",
            gap_filling_method="Gap Filling Method",
            notes="Notes",
        )
        values = self.filter_queryset(self.get_queryset().values(*headers.keys()))

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Household_Size_Data")

        # Headers
        ws.append(list(headers.values()))

        # Data rows
        for item in values:
            ws.append(
                [
                    item["country__region__name"],
                    item["country__idmc_short_name"],
                    item["year"],
                    item["size"],
                    item["reference_year"],
                    item["data_source_category"],
                    item["source"],
                    item["source_link"],
                    item["gap_filling_method"],
                    item["notes"],
                ]
            )

        ws2 = wb.create_sheet("README")
        readme_text = [
            ["TITLE: Average Household Size"],
            [],
            ["FILENAME: IDMC_Average_Household_Size_Data"],
            [],
            ["SOURCE: Internal Displacement Monitoring Centre (IDMC)"],
            [],
            [f"DATE EXTRACTED: {datetime.now().strftime('%B %d, %Y')}"],
            [],
            ["DESCRIPTION:"],
            [
                # change this description: consult with Keyur
                "The Internal Displacement Monitoring Centre (IDMC) monitors internal displacement events globally, "
                "triggered by disasters, conflict, and other forms of violence. It gathers and analyses both "
                "structured and unstructured secondary data from diverse sources—including government agencies, "
                "UN agencies, the International Federation of the Red Cross and Red Crescent, and the media."
            ],
            [],
            ["KEY DEFINITIONS:"],
            [],
            [
                # add more definitions: consult with Keyur
                "Internal Displacements (flows): This metric represents the number of internal displacements, or "
                "internal displacement population flows, reported from January 1st to December 31st of a reporting year. "
                "This figure may include individuals who are displaced multiple times during the year by different events."
            ],
            [],
            [
                "USE LICENSE: This content is licensed under CC BY-NC. Detailed licensing information is available at "
                "Creative Commons License (See: https://creativecommons.org/licenses/by-nc/4.0/)."
            ],
            [],
            [
                "COVERAGE: Global. The GIDD provides data on internal displacements triggered by disasters dates back "
                "to 2008, and the metrics on the total number of IDPs from disaster-related events are available from "
                "2019 onwards."
            ],
            [],
            ["CITATION:"],
            [
                "All derived work from IDMC data could cite IDMC following this example: Internal Displacement "
                "Monitoring Centre. Global Internal Displacement Database - Disasters. IDMC (2023). Available at: "
                "https://www.internal-displacement.org/database/displacement-data/ (Accessed: [date of access])."
            ],
            [],
            ["CONTACT: ch.datainfo@idmc.ch"],
            [],
        ]
        readme = [
            ("Region Name", "IDMC regions"),
            ("Country", "Country's or territory short name"),
            ("Year", "Year of displacement"),
            ("AHHS", "Average household size. This values are comapiled by IDMC from UN and national sources."),
            ("Reference Year", "Year of data reference"),
            ("Data Source Category", "Data Source Category"),
            ("Source", "Source of the household size"),
            ("Source Link", "Link of the source."),
            ("Gap Filling Method", "Gap Filling Method"),
            ("Notes", "Notes"),
        ]
        for item in readme:
            ws2.append([" : ".join(item)])

        for additional_readme in readme_text:
            ws2.append(additional_readme)

        ws2.append([])

        response = HttpResponse(content=save_virtual_workbook(wb))

        filename = "IDMC_Average_Household_Size_Data.xlsx"
        response["Content-Disposition"] = f"attachment; filename={filename}"
        response["Content-Type"] = "application/octet-stream"

        return response
