import tempfile
from io import StringIO

from django.core.management import call_command
from django.core.management.base import CommandError
from openpyxl import Workbook, load_workbook

from apps.organization.models import Organization, OrganizationKind
from apps.users.enums import USER_ROLE
from apps.users.utils import HelixInternalBot
from utils.factories import CountryFactory, OrganizationFactory, OrganizationKindFactory
from utils.tests import HelixTestCase, create_user_with_role


def write_sheet(headers, rows, sheet_name="Data"):
    """Write a temporary .xlsx with the given headers + rows and return its path."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    workbook.save(tmp.name)
    return tmp.name


class TestImportOrganizationsCommand(HelixTestCase):
    def setUp(self):
        self.kind = OrganizationKindFactory.create(name="Government")
        self.bot_user = HelixInternalBot().user

    def test_create_rows(self):
        country = CountryFactory.create(iso3="NPL")
        path = write_sheet(
            ["name", "short_name", "category", "countries", "organization_kind"],
            [
                {"name": "Alpha Org", "short_name": "ALPHA", "category": "INTERNATIONAL"},
                {"name": "Beta Org", "category": "NATIONAL", "countries": "NPL", "organization_kind": "Government"},
            ],
        )
        call_command("import_organizations", path)

        self.assertEqual(Organization.objects.count(), 2)
        beta = Organization.objects.get(name="Beta Org")
        self.assertEqual(beta.category, Organization.ORGANIZATION_CATEGORY.NATIONAL.value)
        self.assertEqual(list(beta.countries.all()), [country])
        self.assertEqual(beta.organization_kind, self.kind)
        self.assertEqual(beta.created_by, self.bot_user)

    def test_update_blank_leaves_field_unchanged(self):
        org = OrganizationFactory.create(name="Old Name", short_name="KEEP")
        path = write_sheet(
            ["id", "name", "short_name"],
            [{"id": org.id, "name": "New Name", "short_name": "   "}],  # blank/whitespace
        )
        user = create_user_with_role(USER_ROLE.ADMIN.name)
        call_command("import_organizations", path, "--user-email", user.email)

        org.refresh_from_db()
        self.assertEqual(org.name, "New Name")
        self.assertEqual(org.short_name, "KEEP")  # blank did NOT clear it
        self.assertEqual(org.last_modified_by, user)
        self.assertEqual(Organization.objects.count(), 1)  # updated, not created

    def test_update_clear_token_clears_field(self):
        org = OrganizationFactory.create(name="Old Name", short_name="OLD")
        path = write_sheet(
            ["id", "name", "short_name"],
            [{"id": org.id, "name": "New Name", "short_name": "<clear>"}],
        )
        call_command("import_organizations", path)

        org.refresh_from_db()
        self.assertEqual(org.name, "New Name")
        self.assertIsNone(org.short_name)  # clear token cleared it

    def test_all_or_nothing_rollback(self):
        path = write_sheet(
            ["name", "countries"],
            [
                {"name": "Good Org", "countries": ""},
                {"name": "Bad Org", "countries": "ZZZ"},  # unknown ISO3
            ],
        )
        with self.assertRaises(CommandError):
            call_command("import_organizations", path)
        self.assertEqual(Organization.objects.count(), 0)  # nothing committed

    def test_missing_update_target_errors(self):
        path = write_sheet(["id", "name"], [{"id": 999999, "name": "Ghost"}])
        with self.assertRaises(CommandError):
            call_command("import_organizations", path)
        self.assertEqual(Organization.objects.count(), 0)

    def test_lookup_error_not_overwritten_by_required(self):
        # Regression: when a lookup fails on a field that the serializer also requires,
        # the failed lookup leaves the field absent from the payload, so the serializer
        # would report a generic "this field is required" for the same column. The
        # specific ResolutionError must win, not be clobbered by the generic message.
        from rest_framework import serializers

        from apps.contrib.bulk_operations.tasks import generate_dummy_request
        from apps.contrib.management.base import BaseImportCommand, FKByName
        from apps.organization.serializers import OrganizationSerializer, OrganizationUpdateSerializer

        class RequiredKindSerializer(OrganizationSerializer):
            # organization_kind is nullable on the model; force it required so a failed
            # lookup produces the competing "required" error this test guards against.
            organization_kind = serializers.PrimaryKeyRelatedField(queryset=OrganizationKind.objects.all(), required=True)

        class Command(BaseImportCommand):
            model = Organization
            create_serializer = RequiredKindSerializer
            update_serializer = OrganizationUpdateSerializer
            lookups = [FKByName("organization_kind", OrganizationKind, "name")]

        request = generate_dummy_request(self.bot_user)
        raw_row = {"name": "Test Org", "organization_kind": "Nonexistent Kind"}
        _serializer, _is_update, row_errors = Command().prepare_row(raw_row, request)

        # The specific lookup error is preserved; the generic "required" did not overwrite it.
        self.assertIn("organization_kind", row_errors)
        self.assertIn("Nonexistent Kind", row_errors["organization_kind"])
        self.assertNotIn("required", row_errors["organization_kind"].lower())

    def test_parent_not_exposed(self):
        # `parent` is in the command's local denylist (EXTRA_EXCLUDED_FIELDS): not a template
        # column, and rejected on import.
        out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        call_command("import_organizations", "--make-template", out.name)
        self.assertNotIn("parent", [cell.value for cell in load_workbook(out.name)["Data"][1]])

        path = write_sheet(["name", "parent"], [{"name": "Child", "parent": "IDMC - Nepal"}])
        with self.assertRaises(CommandError):
            call_command("import_organizations", path)
        self.assertFalse(Organization.objects.filter(name="Child").exists())

    def test_qualified_fk_lookup_resolves_by_name_and_country(self):
        # The dormant parent lookup's "<name> - <country>" logic stays covered even though
        # parent is not exposed for import.
        from apps.contrib.management.base import QualifiedFKByName

        country = CountryFactory.create(idmc_short_name="Nepal", iso3="NPL")
        parent = OrganizationFactory.create(name="IDMC")
        parent.countries.add(country)
        lookup = QualifiedFKByName(
            "parent", Organization, parent_lookup="name", child_lookup="countries__idmc_short_name", separator=" - "
        )
        self.assertEqual(lookup.resolve("IDMC - Nepal"), parent.id)

        OrganizationFactory.create(name="IDMC").countries.add(country)  # now two "IDMC - Nepal"
        lookup.reset()
        from apps.contrib.management.base import ResolutionError

        with self.assertRaises(ResolutionError):
            lookup.resolve("IDMC - Nepal")

    def test_qualified_fk_lookup_without_country_uses_name_only(self):
        # An organization with no country keys on its name alone, not "<name> - None".
        from apps.contrib.management.base import QualifiedFKByName

        solo = OrganizationFactory.create(name="Solo Org")  # no countries
        lookup = QualifiedFKByName(
            "parent", Organization, parent_lookup="name", child_lookup="countries__idmc_short_name", separator=" - "
        )
        self.assertEqual(lookup.resolve("Solo Org"), solo.id)
        self.assertIn("Solo Org", lookup.enumerate_values())
        self.assertNotIn("Solo Org - None", lookup.enumerate_values())

    def test_dry_run_commits_nothing(self):
        path = write_sheet(["name"], [{"name": "Ephemeral"}])
        call_command("import_organizations", path, "--dry-run")
        self.assertEqual(Organization.objects.count(), 0)

    def test_unknown_header_rejected(self):
        path = write_sheet(["name", "bogus_column"], [{"name": "X", "bogus_column": "y"}])
        with self.assertRaises(CommandError):
            call_command("import_organizations", path)

    def test_make_template(self):
        out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        call_command("import_organizations", "--make-template", out.name)

        workbook = load_workbook(out.name)
        self.assertIn("Data", workbook.sheetnames)
        data_headers = [cell.value for cell in workbook["Data"][1]]
        for expected in ["id", "name", "short_name", "category", "countries", "organization_kind"]:
            self.assertIn(expected, data_headers)
        self.assertNotIn("parent", data_headers)  # local denylist (EXTRA_EXCLUDED_FIELDS)

        # Only README + Data sheets; the separate Reference sheet was removed.
        self.assertEqual(workbook.sheetnames, ["README", "Data"])

        readme_text = " ".join(
            str(cell.value) for row in workbook["README"].iter_rows() for cell in row if cell.value is not None
        )
        # H1 title + metadata caption.
        self.assertIn("Organization Import Template", readme_text)
        self.assertIn("Generated:", readme_text)
        self.assertIn("Source version:", readme_text)
        self.assertIn("Environment:", readme_text)
        # H2 sections.
        self.assertIn("How to use this template", readme_text)
        self.assertIn("Template Shape", readme_text)
        self.assertIn("Allowed Choices", readme_text)
        # Usage + types + notes.
        self.assertIn("<clear>", readme_text)
        self.assertIn("single choice, case-sensitive", readme_text)  # category type + case
        self.assertIn("multiple reference, case-sensitive", readme_text)  # countries type + case
        self.assertIn("by iso3", readme_text)  # note for the non-name reference
        # Allowed Choices grid values.
        self.assertIn("INTERNATIONAL", readme_text)
        self.assertIn("Government", readme_text)

        # Written cells use the portable Arial 10 font.
        readme = workbook["README"]
        self.assertEqual((readme["A2"].font.name, readme["A2"].font.size), ("Arial", 10))  # metadata line
        self.assertEqual(workbook["Data"]["A1"].font.name, "Arial")  # data header

    def test_make_template_warns_on_duplicate_allowed_values(self):
        OrganizationKindFactory.create(name="Dup Kind")
        OrganizationKindFactory.create(name="Dup Kind")  # duplicate organization_kind name
        err = StringIO()
        out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        call_command("import_organizations", "--make-template", out.name, stderr=err)
        message = err.getvalue()
        self.assertIn("duplicate allowed value", message)
        self.assertIn("Dup Kind", message)

    def test_blacklisted_system_fields_absent_from_template(self):
        # System/audit fields must never appear as columns, even though
        # OrganizationSerializer uses fields="__all__" and exposes them as writable.
        out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        call_command("import_organizations", "--make-template", out.name)
        data_headers = [cell.value for cell in load_workbook(out.name)["Data"][1]]
        for hidden in ["old_id", "version_id", "deleted_on", "created_by", "created_at", "modified_at", "last_modified_by"]:
            self.assertNotIn(hidden, data_headers)

    def test_blacklisted_column_rejected_on_import(self):
        path = write_sheet(["name", "version_id"], [{"name": "X", "version_id": "v1"}])
        with self.assertRaises(CommandError):
            call_command("import_organizations", path)
        self.assertFalse(Organization.objects.filter(name="X").exists())
