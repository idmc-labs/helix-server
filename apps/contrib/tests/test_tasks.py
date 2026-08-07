"""Unit tests for `apps.contrib.tasks.get_excel_sheet_content`.

Focused on the Phase 2a iterator-dispatch change in `append_to_worksheet`:
the consumer must accept both Django QuerySets (which expose `.iterator(chunk_size=...)`)
and plain Python iterables (the new explode generator).
"""

import io
from collections import OrderedDict
from unittest.mock import MagicMock

from openpyxl import load_workbook

from apps.contrib.tasks import get_excel_sheet_content
from utils.tests import HelixTestCase


class TestAppendToWorksheetDispatch(HelixTestCase):
    """Asserts the `hasattr(_data, "iterator")` dispatch in `append_to_worksheet`."""

    @staticmethod
    def _read_main_sheet_rows(workbook):
        """Round-trip the write-only workbook through bytes so it's readable, return Main rows."""
        buf = io.BytesIO()
        workbook.save(buf)
        buf.seek(0)
        ws = load_workbook(buf)["Main"]
        return [tuple(cell.value for cell in row) for row in ws.iter_rows()]

    def test_plain_generator_is_iterated_directly(self):
        """A plain generator (no `.iterator` attr) is iterated directly."""

        def gen():
            yield {"id": 1, "name": "alpha"}
            yield {"id": 2, "name": "beta"}

        headers = OrderedDict(id="ID", name="Name")
        workbook = get_excel_sheet_content(headers=headers, data=gen())
        rows = self._read_main_sheet_rows(workbook)
        # First row is the header row; then two data rows.
        self.assertEqual(rows[0], ("ID", "Name"))
        self.assertEqual(rows[1], (1, "alpha"))
        self.assertEqual(rows[2], (2, "beta"))

    def test_queryset_like_object_uses_iterator_with_chunk_size(self):
        """A QuerySet-like mock with `.iterator(chunk_size=...)` is called with chunk_size=2000."""
        mock_qs = MagicMock()
        mock_qs.iterator.return_value = iter(
            [
                {"id": 10, "name": "gamma"},
                {"id": 11, "name": "delta"},
            ]
        )

        headers = OrderedDict(id="ID", name="Name")
        workbook = get_excel_sheet_content(headers=headers, data=mock_qs)

        mock_qs.iterator.assert_called_once_with(chunk_size=2000)
        rows = self._read_main_sheet_rows(workbook)
        self.assertEqual(rows[0], ("ID", "Name"))
        self.assertEqual(rows[1], (10, "gamma"))
        self.assertEqual(rows[2], (11, "delta"))
