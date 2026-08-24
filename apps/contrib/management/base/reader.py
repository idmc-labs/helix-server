"""Reading and validating rows from an import .xlsx sheet."""

import typing

from django.core.management.base import CommandError

from .utils import DISPLAY_SEP, is_empty


def read_rows(
    file_path: str,
    *,
    data_sheet: str,
    allowed_columns: typing.Iterable[str],
) -> typing.List[typing.Dict]:
    """
    Read the data sheet into a list of {header: value} dicts.

    The header row must contain only columns in `allowed_columns` (any other column is an error).
    Fully blank rows are skipped.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook[data_sheet] if data_sheet in workbook.sheetnames else workbook.active

    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise CommandError("The sheet is empty.")

    from openpyxl.utils import get_column_letter

    # Trailing empty cells are phantom columns left by formatting, and carry no data of their own.
    header_cells = list(header_row)
    while header_cells and is_empty(header_cells[-1]):
        header_cells.pop()

    # Headers keep their position: a value is read against the header above it, so a blank header
    # among the columns cannot be dropped. Dropping it would pair every header to its right with
    # the previous column's value and write the wrong field without saying so.
    headers = ["" if cell is None else str(cell).strip() for cell in header_cells]
    unheaded = [get_column_letter(index + 1) for index, header in enumerate(headers) if not header]
    if unheaded:
        raise CommandError(
            f"Column(s) {DISPLAY_SEP.join(unheaded)} have no header. Delete the column or restore its "
            "header: a blank header would read the values beside it against the wrong columns."
        )

    # zip() pairs a header with the cell below it, so a header appearing twice keeps only the
    # rightmost column and discards the other in silence.
    seen: typing.Set[str] = set()
    repeated = sorted({header for header in headers if header in seen or seen.add(header)})
    if repeated:
        raise CommandError(
            f"Duplicate column(s): {DISPLAY_SEP.join(repeated)}. Each column may appear once: "
            "a repeated header would keep one of its cells and drop the other without saying which."
        )

    allowed = set(allowed_columns)
    unknown = [header for header in headers if header not in allowed]
    if unknown:
        raise CommandError(f"Unknown column(s): {DISPLAY_SEP.join(unknown)}. Allowed columns: {', '.join(sorted(allowed))}")

    rows = []
    for row_number, values in enumerate(rows_iter, start=2):
        if all(is_empty(value) for value in values):
            continue  # skip fully blank rows
        # A value past the last header belongs to a column the header row does not name, so it
        # would be dropped silently.
        beyond = [
            get_column_letter(index + 1)
            for index, value in enumerate(values)
            if index >= len(headers) and not is_empty(value)
        ]
        if beyond:
            raise CommandError(
                f"Row {row_number} has value(s) in column(s) {DISPLAY_SEP.join(beyond)}, which the "
                "header row does not name. Add a header or clear the column."
            )
        rows.append({header: value for header, value in zip(headers, values)})
    return rows
