"""Generating a blank import template: a README sheet (usage + allowed values) and the Data sheet."""

import typing

from .lookups import BaseLookup

FONT_NAME = "Arial"
FONT_SIZE = 10


def _font(size: int = FONT_SIZE, bold: bool = False):
    from openpyxl.styles import Font

    return Font(name=FONT_NAME, size=size, bold=bold)


def _cell(sheet, row: int, col: int, value, *, bold: bool = False, size: int = FONT_SIZE, wrap: bool = False):
    """Write a value with the portable Arial font (openpyxl 3.0.6 cannot set a true workbook default)."""
    from openpyxl.styles import Alignment

    cell = sheet.cell(row=row, column=col, value=value)
    cell.font = _font(size=size, bold=bold)
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell


def _style_header_row(sheet, ncols: int):
    """Bold + shaded header row, frozen so it stays visible while scrolling."""
    from openpyxl.styles import Alignment, PatternFill

    fill = PatternFill("solid", fgColor="D9E1F2")
    for col in range(1, ncols + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = _font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
    if ncols:
        sheet.freeze_panes = "A2"


def _autosize(sheet, widths: typing.List[int]):
    from openpyxl.utils import get_column_letter

    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 12), 60)


def _write_readme(
    workbook,
    *,
    title,
    metadata,
    data_sheet,
    columns,
    lookups,
    clear_token,
    required_columns,
    column_types,
    column_notes,
    update_only,
):
    span = 4  # widest section (the Template Shape table): Column | Type | Required | Note

    sheet = workbook.active
    sheet.title = "README"
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 36
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 44

    row = 1

    def heading(text, size):
        nonlocal row
        _cell(sheet, row, 1, text, size=size, bold=True)
        row += 1

    def line(text):
        nonlocal row
        _cell(sheet, row, 1, text, wrap=True)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        row += 1

    def blank():
        nonlocal row
        row += 1

    # H1 + metadata caption.
    heading(title, 16)
    for key, value in metadata.items():
        line(f"{key}: {value}")
    blank()

    # H2 How to use this template.
    heading("How to use this template", 12)
    blank()
    id_usage = (
        "Enter an existing record's id in every row. This importer only UPDATES; it never creates records."
        if update_only
        else "Leave 'id' blank to CREATE a new record; enter an existing record's id to UPDATE it."
    )
    for text in [
        f"Fill in the '{data_sheet}' sheet, one row per record.",
        id_usage,
        "On update, only the columns you fill are changed; a blank or whitespace-only cell leaves that field unchanged.",
        f"To clear a field on update, put {clear_token} in the cell.",
        "List columns accept multiple values separated by ';'.",
        "Choice and name-based reference values are case-sensitive and must match the Allowed Choices exactly.",
        "Reference-by-id columns take numeric ids (see the Note column).",
        "Any column not listed under Template Shape is rejected.",
    ]:
        line(text)
    blank()

    # H2 Template Shape.
    heading("Template Shape", 12)
    blank()
    required_header = "Required" if update_only else "Required (create)"
    id_note = "id of the row to update" if update_only else "leave blank to create; set an existing id to update"
    for col, text in enumerate(["Column", "Type", required_header, "Note"], start=1):
        _cell(sheet, row, col, text, bold=True)
    row += 1
    for column in columns:
        if column == "id":
            required = "yes" if update_only else "no"
        else:
            required = "yes" if column in required_columns else "no"
        note = id_note if column == "id" else column_notes.get(column, "")
        _cell(sheet, row, 1, column)
        _cell(sheet, row, 2, column_types.get(column, "text"))
        _cell(sheet, row, 3, required)
        _cell(sheet, row, 4, note, wrap=True)
        row += 1
    blank()

    # H2 Allowed Choices: one column per coded field, values down the rows.
    heading("Allowed Choices", 12)
    blank()
    coded = [lookup for lookup in lookups if lookup.list_values and lookup.enumerate_values()]
    if coded:
        from openpyxl.utils import get_column_letter

        header_row = row
        for index, lookup in enumerate(coded, start=1):
            _cell(sheet, header_row, index, lookup.field, bold=True)
            if index > span:  # keep the shape-table widths on A-D; size any extra grid columns
                sheet.column_dimensions[get_column_letter(index)].width = 24
        max_len = max(len(lookup.enumerate_values()) for lookup in coded)
        for offset in range(max_len):
            for index, lookup in enumerate(coded, start=1):
                values = lookup.enumerate_values()
                if offset < len(values):
                    _cell(sheet, header_row + 1 + offset, index, values[offset])
    else:
        line("(none)")


def write_template(
    out_path: str,
    *,
    title: str,
    metadata: typing.Dict[str, str],
    data_sheet: str,
    columns: typing.List[str],
    lookups: typing.List[BaseLookup],
    clear_token: str,
    required_columns: typing.Iterable[str],
    column_types: typing.Dict[str, str],
    column_notes: typing.Dict[str, str],
    update_only: bool = False,
) -> None:
    """
    Write a blank template workbook. All written cells use Arial 10 for portability (openpyxl
    3.0.6 cannot set a true workbook-default font, so rows the user adds later use their app default).
    - a `README` sheet: an H1 title + metadata, then "How to use this template", "Template Shape"
      (Column | Type | Required | Note), and "Allowed Choices" (one column per coded field)
    - a `data_sheet` with one styled header row (the importable columns) and no data rows
    """
    from openpyxl import Workbook

    workbook = Workbook()

    _write_readme(
        workbook,
        title=title,
        metadata=metadata,
        data_sheet=data_sheet,
        columns=columns,
        lookups=lookups,
        clear_token=clear_token,
        required_columns=set(required_columns),
        column_types=column_types,
        column_notes=column_notes,
        update_only=update_only,
    )

    data = workbook.create_sheet(data_sheet)
    data.append(columns)
    _style_header_row(data, len(columns))
    _autosize(data, [len(column) for column in columns])

    workbook.save(out_path)
