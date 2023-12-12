from __future__ import annotations

import logging
import typing
import re
import time
import json
from datetime import timedelta

from django.core.files import File
from django.conf import settings
from django.utils import timezone
from django.db.models import F, Value, CharField
from django.db.models.functions import Concat
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from utils.common import get_temp_file, redis_lock
# from helix.settings import QueuePriority
from helix.celery import app as celery_app

from apps.entry.tasks import PDF_TASK_TIMEOUT
from apps.report.tasks import REPORT_TIMEOUT
from apps.contrib.redis_client_track import (
    get_client_tracked_cache_keys,
    pull_track_data_from_redis,
    delete_external_redis_record_by_key,
)

if typing.TYPE_CHECKING:
    from apps.contrib.models import BulkApiOperation


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def get_excel_sheet_content(headers, data, **kwargs):
    wb = Workbook(write_only=True)

    ws = wb.create_sheet('Main')

    def clean_data_item(item):
        # NOTE: we are using isinstance(item, int) because 0 is falsy value
        if isinstance(item, int):
            return item
        elif isinstance(item, str):
            return re.sub(ILLEGAL_CHARACTERS_RE, '', item)
        elif item:
            return str(item)
        return ''

    def append_to_worksheet(_ws, _headers, _data, _transformer):
        keys = _headers.keys()
        _ws.append([_headers[key] for key in keys])
        for _datum in _data.iterator(chunk_size=2000):
            transformed_datum = _datum
            if _transformer:
                transformed_datum = _transformer(_datum)
            _ws.append([
                clean_data_item(transformed_datum.get(key))
                for key in keys
            ])

    # append the primary sheet
    append_to_worksheet(ws, headers, data, kwargs.get('transformer'))
    # append the secondary sheets if provided
    for other in kwargs.get('other', []):
        ws = wb.create_sheet(other['title'])
        headers = other['results']['headers']
        data = other['results']['data']
        transformer = other['results'].get('transformer')
        append_to_worksheet(ws, headers, data, transformer)

    for data in kwargs.get('readme_data', []):
        ws = wb.create_sheet(data['title'])
        headers = data['results']['headers']
        data = data['results']['data']
        ws.append(list(headers.values()))
        row = []
        for item in data:
            for key in headers:
                row.append(item[key])
            ws.append(row)
            row = []
    return wb


def save_download_file(download, workbook, path):
    with get_temp_file() as tmp:
        workbook.save(tmp.name)
        workbook.close()
        file = File(tmp)
        download.file_size = file.size
        download.file.save(path, file)
        del workbook


@celery_app.task(time_limit=settings.EXCEL_EXPORT_PROGRESS_STATE_TIMEOUT)
def generate_excel_file(download_id, user_id, model_instance_id=None):
    '''
    Fetch the filter data from excel download
    Fetch the request from the task argument
    Call appropriate model excel data getter
    '''
    from apps.contrib.models import ExcelDownload
    download = ExcelDownload.objects.get(id=download_id)
    download.started_at = timezone.now()
    download.status = ExcelDownload.EXCEL_GENERATION_STATUS.IN_PROGRESS
    download.save()
    try:
        then = time.time()
        path = f'{download.download_type.name}-{download.started_at.isoformat()}.xlsx'

        logger.warn(f'Starting sheet generation for ExcelDownload={download_id}...')
        if (
            ExcelDownload.DOWNLOAD_TYPES.INDIVIDUAL_REPORT == download.download_type and
            model_instance_id is not None
        ):
            from apps.report.models import Report
            from apps.report.utils import report_get_excel_sheets_data
            from apps.report.tasks import generate_excel_file as report_generate_excel_file
            report = Report.objects.get(id=model_instance_id)
            excel_sheet_data = report_get_excel_sheets_data(report).items()
            workbook = report_generate_excel_file(excel_sheet_data)
            save_download_file(download, workbook, path)
        else:
            sheet_data_getter = download.get_model_sheet_data_getter()
            sheet_data = sheet_data_getter(user_id=user_id, filters=download.filters)
            workbook = get_excel_sheet_content(**sheet_data)
            save_download_file(download, workbook, path)
        download.status = ExcelDownload.EXCEL_GENERATION_STATUS.COMPLETED
        download.completed_at = timezone.now()
        download.save()

        logger.warn(f'Completed sheet generation for ExcelDownload={download_id} in {time.time() - then}')
    except Exception as e:  # NOQA E722
        logger.error(f'Error: Sheet generation for ExcelDownload={download_id}', exc_info=True)
        download.status = ExcelDownload.EXCEL_GENERATION_STATUS.FAILED
        download.completed_at = timezone.now()
        download.save(update_fields=['status'])


@celery_app.task
def kill_all_old_excel_exports():
    from apps.contrib.models import ExcelDownload
    # if a task has been pending for too long, move it to killed
    pending = ExcelDownload.objects.filter(
        status=ExcelDownload.EXCEL_GENERATION_STATUS.PENDING,
        created_at__lte=timezone.now() - timedelta(seconds=settings.EXCEL_EXPORT_PENDING_STATE_TIMEOUT),
    ).update(status=ExcelDownload.EXCEL_GENERATION_STATUS.KILLED)

    # if a task has been in progress beyond timeout, move it to killed
    progress = ExcelDownload.objects.filter(
        status=ExcelDownload.EXCEL_GENERATION_STATUS.IN_PROGRESS,
        started_at__lte=timezone.now() - timedelta(seconds=settings.EXCEL_EXPORT_PROGRESS_STATE_TIMEOUT),
    ).update(status=ExcelDownload.EXCEL_GENERATION_STATUS.KILLED)

    logger.info(f'Updated EXCEL EXPORTS to killed:\n{pending=}\n{progress=}')


@celery_app.task
def kill_all_long_running_previews():
    from apps.contrib.models import SourcePreview

    progress = SourcePreview.objects.filter(
        status=SourcePreview.PREVIEW_STATUS.IN_PROGRESS,
        created_at__lte=timezone.now() - timedelta(seconds=PDF_TASK_TIMEOUT * 5),
    ).update(status=SourcePreview.PREVIEW_STATUS.KILLED)

    logger.info(f'Updated SOURCE PREVIEWS to killed:\n{progress=}')


@celery_app.task
def kill_all_long_running_report_generations():
    from apps.report.models import ReportGeneration

    progress = ReportGeneration.objects.filter(
        status=ReportGeneration.REPORT_GENERATION_STATUS.IN_PROGRESS,
        created_at__lte=timezone.now() - timedelta(seconds=REPORT_TIMEOUT * 2),
    ).update(status=ReportGeneration.REPORT_GENERATION_STATUS.KILLED)

    logger.info(f'Updated REPORT GENERATION to killed:\n{progress=}')


def generate_external_endpoint_dump_file(
    endpoint_type,
    serializer,
    get_data,
    filename,
):
    from apps.entry.models import ExternalApiDump
    external_api_dump, _ = ExternalApiDump.objects.get_or_create(api_type=endpoint_type)
    try:
        data = get_data()
        serializer = serializer(data, many=True)
        with get_temp_file(mode="w+") as tmp:
            json.dump(serializer.data, tmp)
            external_api_dump.dump_file.save(
                filename,
                File(tmp),
            )
        external_api_dump.status = ExternalApiDump.Status.COMPLETED
        logger.info(f'{endpoint_type}: file dump created')
    except Exception:
        external_api_dump.status = ExternalApiDump.Status.FAILED
        logger.error(f'{endpoint_type}: file dump generation failed', exc_info=True)
        return False
    external_api_dump.save()
    return True


def _generate_idus_dump_file(api_type):
    from apps.entry.serializers import FigureReadOnlySerializer
    from apps.entry.views import get_idu_data
    from apps.entry.models import ExternalApiDump
    from apps.crisis.models import Crisis

    if api_type == ExternalApiDump.ExternalApiType.IDUS_ALL:
        return generate_external_endpoint_dump_file(
            ExternalApiDump.ExternalApiType.IDUS_ALL,
            FigureReadOnlySerializer,
            get_idu_data,
            'idus_all.json',
        )
    if api_type == ExternalApiDump.ExternalApiType.IDUS_ALL_DISASTER:
        return generate_external_endpoint_dump_file(
            ExternalApiDump.ExternalApiType.IDUS_ALL_DISASTER,
            FigureReadOnlySerializer,
            lambda: get_idu_data().filter(figure_cause=Crisis.CRISIS_TYPE.DISASTER),
            'idus_all_disaster.json',
        )
    idu_date_from = timezone.now() - timedelta(days=180)
    return generate_external_endpoint_dump_file(
        ExternalApiDump.ExternalApiType.IDUS,
        FigureReadOnlySerializer,
        lambda: get_idu_data().filter(displacement_date__gte=idu_date_from),
        'idus.json',
    )


def _run_bulk_api_operation(operation: BulkApiOperation):
    print(operation)


@celery_app.task
def generate_idus_dump_file():
    from apps.entry.models import ExternalApiDump
    return _generate_idus_dump_file(ExternalApiDump.ExternalApiType.IDUS)


@celery_app.task
def generate_idus_all_dump_file():
    from apps.entry.models import ExternalApiDump
    return _generate_idus_dump_file(ExternalApiDump.ExternalApiType.IDUS_ALL)


@celery_app.task
def generate_idus_all_disaster_dump_file():
    from apps.entry.models import ExternalApiDump
    return _generate_idus_dump_file(ExternalApiDump.ExternalApiType.IDUS_ALL_DISASTER)


@celery_app.task
@redis_lock('remaining_lead_extract', 60 * 5)
def save_and_delete_tracked_data_from_redis_to_db():
    from apps.contrib.models import ClientTrackInfo

    tracking_keys = get_client_tracked_cache_keys()
    tracked_data_from_redis = pull_track_data_from_redis(tracking_keys)

    # Update track records count with max value if they already exist in database
    existing_track_info_qs = ClientTrackInfo.objects.annotate(
        redis_tracking_key=Concat(
            Value('trackinfo:'),
            F('tracked_date'),
            Value(':'),
            F('api_type'),
            Value(':'),
            F('client__code'),
            output_field=CharField()
        )
    ).filter(
        redis_tracking_key__in=tracked_data_from_redis.keys()
    )

    existing_track_info_map = {
        track_info.redis_tracking_key: track_info
        for track_info in existing_track_info_qs
    }

    new_objects = []
    update_objects = []
    for key, tracked_item in tracked_data_from_redis.items():
        existing_track_info = existing_track_info_map.get(key)
        if existing_track_info:
            # If there are more than one tracking info objects per
            # day take maximum value and update.
            existing_track_info.requests_per_day = max(
                tracked_item['requests_per_day'],
                existing_track_info.requests_per_day
            )
            update_objects.append(existing_track_info)
        else:
            new_objects.append(
                ClientTrackInfo(
                    api_type=tracked_item['api_type'],
                    client_id=tracked_item['client_id'],
                    tracked_date=tracked_item['tracked_date'],
                    requests_per_day=tracked_item['requests_per_day'],
                ),
            )

    # Save to database from redis
    ClientTrackInfo.objects.bulk_create(new_objects)
    ClientTrackInfo.objects.bulk_update(
        update_objects,
        fields=['requests_per_day'],
    )

    # Finally delete redis keys after save
    delete_external_redis_record_by_key(tracking_keys)


@celery_app.task
def run_bulk_api_operation(operation_id):
    from apps.contrib.models import BulkApiOperation
    operation = BulkApiOperation.objects.get(pk=operation_id)
    return _run_bulk_api_operation(operation)
