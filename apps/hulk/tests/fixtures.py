"""
Loaders for the hulk-bulk test fixture under ``artifacts/fixtures/hulk-bulk/``.

The fixture is authored as five xlsx files (one per resource type) plus the
"expected" output files. ``_build_fixtures.py`` regenerates everything from
the row definitions in that script.

This module:

* exposes the deterministic UUIDs used in the fixture so tests can assert
  by UUID rather than re-deriving them;
* reads the raw xlsx files and substitutes the test-time placeholder IDs
  (country / publisher / source / sub-types) into the rows;
* serialises the substituted rows to JSONL bytes ready to be uploaded to
  ``HulkBulkImport.import_<resource>``.

The expected output files (``expected/jsonl``, ``expected/success``,
``expected/failure``) are exposed via ``read_expected()`` for assertions.
"""

from __future__ import annotations

import json
import typing
from pathlib import Path

from openpyxl import load_workbook

# Repo root → artifacts/fixtures/hulk-bulk
FIXTURE_DIR = Path(__file__).resolve().parents[3] / "artifacts" / "fixtures" / "hulk-bulk"
RAW_DIR = FIXTURE_DIR / "raw"
EXPECTED_JSONL_DIR = FIXTURE_DIR / "expected" / "jsonl"
EXPECTED_SUCCESS_DIR = FIXTURE_DIR / "expected" / "success"
EXPECTED_FAILURE_DIR = FIXTURE_DIR / "expected" / "failure"

# Re-export from the builder so callers have one import path.
from apps.hulk.tests._build_fixtures import (  # noqa: E402, F401
    ATTACHMENT_UUIDS,
    DUMMY_FILE_URL,
    EMPTY_STRING_SENTINEL,
    ENTRY_UUIDS,
    EVENT_UUIDS,
    EXPECTED_OUTCOMES,
    FIGURE_UUIDS,
    LOCATION_UUIDS,
    PH_COUNTRY,
    PH_DISASTER,
    PH_OTHER,
    PH_PUBLISHER,
    PH_SOURCE,
    PH_VIOLENCE,
    SOURCE_PREVIEW_UUIDS,
)

__all__ = [
    "ATTACHMENT_UUIDS",
    "DUMMY_FILE_URL",
    "ENTRY_UUIDS",
    "EVENT_UUIDS",
    "EXPECTED_OUTCOMES",
    "FIGURE_UUIDS",
    "FixtureContext",
    "LOCATION_UUIDS",
    "RESOURCES",
    "SOURCE_PREVIEW_UUIDS",
    "build_jsonl_bundle",
    "read_expected_failure",
    "read_expected_input_rows",
    "read_expected_success",
]

RESOURCES = ("attachments", "source_previews", "entries", "events", "figures")


class FixtureContext(typing.TypedDict):
    country_id: int
    """A real Country.id whose iso2 matches the location.country_code below ('NP')."""
    publisher_id: int
    """A real Organization.id."""
    source_id: int
    """A real Organization.id used as figure source."""
    violence_sub_type_id: int
    disaster_sub_type_id: int
    other_sub_type_id: int


def _load_raw(resource: str) -> list[dict]:
    """Read raw/<resource>.xlsx and decode JSON-encoded cells back to dict/list."""
    path = RAW_DIR / f"{resource}.xlsx"
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = list(next(it))
    rows: list[dict] = []
    for raw in it:
        row: dict = {}
        for header, value in zip(headers, raw):
            if value == EMPTY_STRING_SENTINEL:
                value = ""
            elif isinstance(value, str) and value and value[0] in "[{":
                try:
                    value = json.loads(value)
                except json.JSONDecodeError:
                    pass
            row[header] = value
        rows.append(row)
    wb.close()
    return rows


def _substitute(value, mapping: dict[str, typing.Any]):
    """Recursively replace placeholder strings (``{{COUNTRY_ID}}`` etc) with real values."""
    if isinstance(value, str) and value in mapping:
        return mapping[value]
    if isinstance(value, list):
        return [_substitute(v, mapping) for v in value]
    if isinstance(value, dict):
        return {k: _substitute(v, mapping) for k, v in value.items()}
    return value


def _placeholder_mapping(ctx: FixtureContext) -> dict[str, typing.Any]:
    # IDs are stored as strings in the JSONL (matching the export shape used
    # in case-01/02) — pyhelix coerces them as needed.
    return {
        PH_COUNTRY: ctx["country_id"],
        PH_PUBLISHER: str(ctx["publisher_id"]),
        PH_SOURCE: str(ctx["source_id"]),
        PH_VIOLENCE: ctx["violence_sub_type_id"],
        PH_DISASTER: ctx["disaster_sub_type_id"],
        PH_OTHER: ctx["other_sub_type_id"],
    }


def _rows_for_ctx(resource: str, ctx: FixtureContext) -> list[dict]:
    mapping = _placeholder_mapping(ctx)
    return [_substitute(row, mapping) for row in _load_raw(resource)]


def _to_jsonl_bytes(rows: list[dict]) -> bytes:
    return ("\n".join(json.dumps(r) for r in rows) + "\n").encode("utf-8") if rows else b""


def build_jsonl_bundle(ctx: FixtureContext) -> dict[str, bytes]:
    """Substitute placeholders and return ``{resource: jsonl_bytes}``."""
    return {resource: _to_jsonl_bytes(_rows_for_ctx(resource, ctx)) for resource in RESOURCES}


# ---------------------------------------------------------------------------
# Expected-output loaders. These read the committed expected files so tests
# can bit-compare or set-compare against them.
# ---------------------------------------------------------------------------
def _read_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        return []
    rows = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def read_expected_input_rows(resource: str) -> list[dict]:
    """The expected unsubstituted JSONL the loader produces from raw xlsx."""
    return _read_jsonl(EXPECTED_JSONL_DIR / f"{resource}.jsonl")


def read_expected_success(resource: str) -> list[dict]:
    """Expected ``success_<resource>`` rows ({uuid, message}, no factory IDs)."""
    return _read_jsonl(EXPECTED_SUCCESS_DIR / f"{resource}.jsonl")


def read_expected_failure(resource: str) -> list[dict]:
    """Expected ``failure_<resource>`` rows ({uuid, error_key})."""
    return _read_jsonl(EXPECTED_FAILURE_DIR / f"{resource}.jsonl")
