import datetime
import tempfile
from io import StringIO

from django.core.management import call_command
from django.core.management.base import CommandError
from openpyxl import Workbook, load_workbook

from apps.crisis.models import Crisis
from apps.report.models import Report
from apps.users.enums import USER_ROLE
from utils.factories import (
    CountryRegionFactory,
    CrisisFactory,
    ViolenceFactory,
    ViolenceSubTypeFactory,
)
from utils.tests import HelixTestCase, create_user_with_role


def write_sheet(headers, rows, sheet_name="Data"):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    workbook.save(tmp.name)
    return tmp.name


class TestImportReportsCommand(HelixTestCase):
    def setUp(self):
        self.admin = create_user_with_role(USER_ROLE.ADMIN.name)
        self.crisis = CrisisFactory.create(name="Sahel Crisis")
        violence = ViolenceFactory.create(name="Armed Conflict")
        self.subtype = ViolenceSubTypeFactory.create(name="State based", violence=violence)

    def test_create_report_with_relations_and_enums(self):
        path = write_sheet(
            [
                "name",
                "filter_figure_start_after",
                "filter_figure_end_before",
                "filter_figure_crises",
                "filter_figure_crisis_types",
                "filter_figure_violence_sub_types",
            ],
            [
                {
                    "name": "Report A",
                    "filter_figure_start_after": "2020-01-01",  # non-GIDD reports require a date range
                    "filter_figure_end_before": "2020-12-31",
                    "filter_figure_crises": str(self.crisis.id),  # crises referenced by id
                    "filter_figure_crisis_types": "CONFLICT;DISASTER",
                    "filter_figure_violence_sub_types": "State based",
                }
            ],
        )
        call_command("import_reports", path)

        report = Report.objects.get(name="Report A")
        self.assertEqual(list(report.filter_figure_crises.all()), [self.crisis])
        self.assertEqual(
            report.filter_figure_crisis_types,
            [Crisis.CRISIS_TYPE.CONFLICT.value, Crisis.CRISIS_TYPE.DISASTER.value],
        )
        self.assertEqual(list(report.filter_figure_violence_sub_types.all()), [self.subtype])
        self.assertIsNotNone(report.created_by)  # attributed to the default internal bot

    def test_update_requires_permitted_user(self):
        report = Report.objects.create(
            name="Old",
            created_by=self.admin,
            # partial update re-runs validate(); the instance must already have the required date range
            filter_figure_start_after=datetime.date(2020, 1, 1),
            filter_figure_end_before=datetime.date(2020, 12, 31),
        )
        path = write_sheet(["id", "name"], [{"id": report.id, "name": "Renamed"}])
        call_command("import_reports", path, "--user-email", self.admin.email)

        report.refresh_from_db()
        self.assertEqual(report.name, "Renamed")
        self.assertEqual(report.last_modified_by, self.admin)
        self.assertEqual(Report.objects.count(), 1)

    def test_all_or_nothing_rollback_on_bad_enum(self):
        path = write_sheet(
            ["name", "filter_figure_crisis_types"],
            [
                {"name": "Good", "filter_figure_crisis_types": "CONFLICT"},
                {"name": "Bad", "filter_figure_crisis_types": "NOT_A_TYPE"},
            ],
        )
        with self.assertRaises(CommandError):
            call_command("import_reports", path)
        self.assertEqual(Report.objects.count(), 0)

    def test_ambiguous_name_reference_errors(self):
        # Name-based references (regions) still error on ambiguity.
        CountryRegionFactory.create(name="Dup Region")
        CountryRegionFactory.create(name="Dup Region")
        path = write_sheet(["name", "filter_figure_regions"], [{"name": "R", "filter_figure_regions": "Dup Region"}])
        with self.assertRaises(CommandError):
            call_command("import_reports", path)
        self.assertFalse(Report.objects.filter(name="R").exists())

    def test_unknown_crisis_id_errors(self):
        # Id-based references error when the id does not exist.
        path = write_sheet(["name", "filter_figure_crises"], [{"name": "R", "filter_figure_crises": "999999"}])
        with self.assertRaises(CommandError):
            call_command("import_reports", path)
        self.assertFalse(Report.objects.filter(name="R").exists())

    def test_dry_run_commits_nothing(self):
        path = write_sheet(
            ["name", "filter_figure_start_after", "filter_figure_end_before"],
            [{"name": "Ephemeral", "filter_figure_start_after": "2020-01-01", "filter_figure_end_before": "2020-12-31"}],
        )
        call_command("import_reports", path, "--dry-run")
        self.assertEqual(Report.objects.count(), 0)

    def test_make_template(self):
        out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        call_command("import_reports", "--make-template", out.name)

        workbook = load_workbook(out.name)
        data_headers = [cell.value for cell in workbook["Data"][1]]
        for expected in [
            "id",
            "name",
            "filter_figure_countries",
            "filter_figure_crisis_types",
            "filter_figure_violence_sub_types",
        ]:
            self.assertIn(expected, data_headers)

        # Only README + Data sheets; allowed values live in the README, not a separate sheet.
        self.assertEqual(workbook.sheetnames, ["README", "Data"])
        readme_text = " ".join(
            str(cell.value) for row in workbook["README"].iter_rows() for cell in row if cell.value is not None
        )
        self.assertIn("Report Import Template", readme_text)  # H1 from verbose_name
        self.assertIn("Template Shape", readme_text)
        self.assertIn("Allowed Choices", readme_text)
        self.assertIn("multiple choice, case-sensitive", readme_text)  # crisis_types / roles type + case
        self.assertIn("CONFLICT", readme_text)  # enum-array value
        self.assertIn("RECOMMENDED", readme_text)  # Figure.ROLE value
        self.assertIn("State based", readme_text)  # violence sub type by name (small table, kept)
        self.assertNotIn("Sahel Crisis", readme_text)  # large table excluded from allowed choices

    def test_make_template_warns_on_separator_in_value(self):
        CountryRegionFactory.create(name="Flood; Fluvial")  # regions matched by name, split on ';'
        err = StringIO()
        out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        call_command("import_reports", "--make-template", out.name, stderr=err)
        message = err.getvalue()
        self.assertIn("filter_figure_regions", message)
        self.assertIn("';' separator", message)

    def test_clear_empties_an_id_based_m2m(self):
        # M2MById inherited the base clear value of None, which a many-to-many serializer field
        # refuses, so an id-based relation could be set but never emptied.
        report = Report.objects.create(
            name="Has crises",
            created_by=self.admin,
            filter_figure_start_after=datetime.date(2020, 1, 1),
            filter_figure_end_before=datetime.date(2020, 12, 31),
        )
        report.filter_figure_crises.set([self.crisis])

        path = write_sheet(["id", "filter_figure_crises"], [{"id": report.id, "filter_figure_crises": "<clear>"}])
        call_command("import_reports", path, "--user-email", self.admin.email)

        report.refresh_from_db()
        self.assertEqual(report.filter_figure_crises.count(), 0)
