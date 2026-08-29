import tempfile
from datetime import date
from io import StringIO

from django.core.management import call_command
from django.core.management.base import CommandError
from openpyxl import Workbook, load_workbook

from apps.crisis.models import Crisis
from apps.entry.models import Figure
from apps.report.management.commands.import_reports import Command as ImportReportsCommand
from apps.report.management.commands.import_reports import ReportImportSerializer
from apps.report.models import Report
from apps.report.serializers import ReportSerializer
from apps.users.enums import USER_ROLE
from utils.factories import CountryFactory, CrisisFactory, TagFactory, UserFactory, ViolenceSubTypeFactory
from utils.tests import HelixTestCase, create_user_with_role


def write_sheet(headers, rows, sheet_name="Data"):
    """Write a temporary .xlsx with the given headers + rows and return its path."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    workbook.save(tmp.name)
    return tmp.name


class TestImportReportsCommand(HelixTestCase):
    def setUp(self):
        super().setUp()
        # ReportSerializer.has_permission_for_report grants edit to several roles and to a
        # REPORTING_TEAM member for their own reports; a user with no portfolio is a guest and has
        # none. create_user_with_role only builds a portfolio for ADMIN and REPORTING_TEAM.
        self.editor = create_user_with_role(USER_ROLE.ADMIN.name)
        self.outsider = UserFactory.create()
        self.country = CountryFactory.create(iso3="NPL")
        self.other_country = CountryFactory.create(iso3="IND")

    def _report(self, name="Probe Report", countries=None):
        # A report carries a figure date filter; the serializer requires both bounds.
        report = Report.objects.create(
            name=name,
            filter_figure_start_after=date(2020, 1, 1),
            filter_figure_end_before=date(2020, 12, 31),
        )
        report.filter_figure_countries.set(countries or [self.country])
        return report

    # ----- create path -----

    def test_a_row_without_an_id_creates_a_report(self):
        path = write_sheet(
            ["id", "name", "filter_figure_countries", "filter_figure_start_after", "filter_figure_end_before"],
            [
                {
                    "id": None,
                    "name": "Fresh Report",
                    "filter_figure_countries": "NPL",
                    "filter_figure_start_after": "2020-01-01",
                    "filter_figure_end_before": "2020-12-31",
                }
            ],
        )
        out = StringIO()
        call_command("import_reports", path, user_email=self.editor.email, stdout=out)

        self.assertIn("Created 1, updated 0.", out.getvalue())
        report = Report.objects.get(name="Fresh Report")
        self.assertEqual(list(report.filter_figure_countries.values_list("iso3", flat=True)), ["NPL"])
        # A create emits no changelog line: there is no previous value to record.
        self.assertNotIn("ROW_UPDATED", out.getvalue())

    def test_several_rows_without_ids_are_not_duplicates_of_each_other(self):
        dates = {"filter_figure_start_after": "2020-01-01", "filter_figure_end_before": "2020-12-31"}
        path = write_sheet(
            ["id", "name", "filter_figure_start_after", "filter_figure_end_before"],
            [{"id": None, "name": "Alpha Report", **dates}, {"id": None, "name": "Beta Report", **dates}],
        )
        out = StringIO()
        call_command("import_reports", path, user_email=self.editor.email, stdout=out)

        self.assertIn("Created 2, updated 0.", out.getvalue())
        self.assertNotIn("also appears on row", out.getvalue())

    # ----- update path -----

    def test_a_row_with_an_id_patches_the_report(self):
        report = self._report()
        path = write_sheet(
            ["id", "name", "analysis"],
            [{"id": report.id, "name": "Renamed Report", "analysis": "Fresh analysis."}],
        )
        out = StringIO()
        call_command("import_reports", path, user_email=self.editor.email, stdout=out)
        output = out.getvalue()

        report.refresh_from_db()
        self.assertEqual(report.name, "Renamed Report")
        self.assertEqual(report.analysis, "Fresh analysis.")
        self.assertEqual(Report.objects.count(), 1)
        self.assertIn(f"ROW_UPDATED\treport={report.id}\trow=2", output)
        self.assertIn("name=Probe Report->Renamed Report", output)

    def test_a_user_without_permission_cannot_edit(self):
        report = self._report()
        path = write_sheet(["id", "name"], [{"id": report.id, "name": "Should Not Land"}])
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_reports", path, user_email=self.outsider.email, stdout=out)

        self.assertIn("You do not have permission to edit report.", out.getvalue())
        report.refresh_from_db()
        self.assertEqual(report.name, "Probe Report")

    def test_a_blank_cell_leaves_the_field_unchanged(self):
        report = self._report()
        report.analysis = "Keep me."
        report.save()

        path = write_sheet(["id", "name", "analysis"], [{"id": report.id, "name": "Renamed", "analysis": None}])
        call_command("import_reports", path, user_email=self.editor.email)

        report.refresh_from_db()
        self.assertEqual(report.name, "Renamed")
        self.assertEqual(report.analysis, "Keep me.")

    # ----- lookups -----

    def test_m2m_by_name_and_its_changelog(self):
        report = self._report()
        path = write_sheet(
            ["id", "filter_figure_countries"],
            [{"id": report.id, "filter_figure_countries": "IND"}],
        )
        out = StringIO()
        call_command("import_reports", path, user_email=self.editor.email, stdout=out)
        output = out.getvalue()

        report.refresh_from_db()
        self.assertEqual(list(report.filter_figure_countries.values_list("iso3", flat=True)), ["IND"])
        # A many-to-many is rendered as its sorted ids, before -> after.
        self.assertIn(f"filter_figure_countries=[{self.country.id}]->[{self.other_country.id}]", output)

    def test_an_enum_array_column_resolves_from_a_delimited_cell(self):
        report = self._report()
        path = write_sheet(
            ["id", "filter_figure_categories"],
            [{"id": report.id, "filter_figure_categories": "IDPS;RETURN"}],
        )
        call_command("import_reports", path, user_email=self.editor.email)

        report.refresh_from_db()
        self.assertEqual(
            sorted(report.filter_figure_categories),
            sorted([Figure.FIGURE_CATEGORY_TYPES.IDPS.value, Figure.FIGURE_CATEGORY_TYPES.RETURN.value]),
        )

    def test_m2m_by_id_resolves_high_cardinality_references(self):
        report = self._report()
        crisis = CrisisFactory.create(crisis_type=Crisis.CRISIS_TYPE.DISASTER)
        path = write_sheet(["id", "filter_figure_crises"], [{"id": report.id, "filter_figure_crises": str(crisis.id)}])
        call_command("import_reports", path, user_email=self.editor.email)

        report.refresh_from_db()
        self.assertEqual(list(report.filter_figure_crises.values_list("id", flat=True)), [crisis.id])

    def test_an_unresolvable_name_fails_the_row(self):
        report = self._report()
        path = write_sheet(
            ["id", "filter_figure_tags"],
            [{"id": report.id, "filter_figure_tags": "NoSuchTag"}],
        )
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_reports", path, user_email=self.editor.email, stdout=out)

        self.assertIn("filter_figure_tags", out.getvalue())
        report.refresh_from_db()
        self.assertEqual(report.filter_figure_tags.count(), 0)

    def test_a_resolvable_tag_is_set(self):
        report = self._report()
        tag = TagFactory.create(name="ProbeTag")
        path = write_sheet(["id", "filter_figure_tags"], [{"id": report.id, "filter_figure_tags": "ProbeTag"}])
        call_command("import_reports", path, user_email=self.editor.email)

        report.refresh_from_db()
        self.assertEqual(list(report.filter_figure_tags.values_list("id", flat=True)), [tag.id])

    def test_a_cell_the_serializer_empties_is_reported_as_ignored(self):
        # validate_figure_crisis_type clears the violence sub-types unless CONFLICT is among the
        # crisis types, so this cell cannot land. The row is still updated; the discarded cell must
        # not vanish in silence.
        report = self._report()
        sub_type = ViolenceSubTypeFactory.create(name="ProbeViolenceSubType")
        path = write_sheet(
            ["id", "filter_figure_crisis_types", "filter_figure_violence_sub_types"],
            [
                {
                    "id": report.id,
                    "filter_figure_crisis_types": "DISASTER",
                    "filter_figure_violence_sub_types": "ProbeViolenceSubType",
                }
            ],
        )
        out = StringIO()
        call_command("import_reports", path, user_email=self.editor.email, stdout=out)
        output = out.getvalue()

        report.refresh_from_db()
        self.assertEqual(report.filter_figure_violence_sub_types.count(), 0)
        self.assertNotIn(sub_type.id, list(report.filter_figure_violence_sub_types.values_list("id", flat=True)))
        self.assertIn("CELL_IGNORED\trow=2\tfilter_figure_violence_sub_types=ProbeViolenceSubType", output)
        self.assertIn("cell(s) were ignored", output)

    def test_a_gidd_report_normalisation_is_not_reported_as_a_change(self):
        # ReportSerializer assigns datetime(year,1,1) to a DateField. The stored value is already
        # that date, so nothing moves; the changelog must not invent a change from the type alone.
        report = Report.objects.create(
            name="GIDD Probe",
            is_gidd_report=True,
            gidd_report_year=2020,
            filter_figure_start_after=date(2020, 1, 1),
            filter_figure_end_before=date(2020, 12, 31),
        )
        path = write_sheet(["id"], [{"id": report.id}])
        out = StringIO()
        call_command("import_reports", path, user_email=self.editor.email, stdout=out)
        output = out.getvalue()

        # The dates never moved in the database, so they must not appear.
        self.assertNotIn("filter_figure_start_after", output)
        self.assertNotIn("filter_figure_end_before", output)
        # The row is still reported: the GIDD branch does make real changes (empty arrays,
        # is_public), and suppressing a type difference must not suppress those.
        self.assertIn("ROW_UPDATED", output)
        self.assertIn("is_public=False->True", output)

    # ----- shared framework behaviour -----

    def test_an_unknown_id_fails_the_row(self):
        path = write_sheet(["id", "name"], [{"id": 987654, "name": "X"}])
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_reports", path, user_email=self.editor.email, stdout=out)

        self.assertIn("no Report found with id 987654", out.getvalue())

    def test_a_duplicate_id_fails_naming_both_rows(self):
        report = self._report()
        path = write_sheet(
            ["id", "name"],
            [{"id": report.id, "name": "First"}, {"id": report.id, "name": "Second"}],
        )
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_reports", path, user_email=self.editor.email, stdout=out)

        self.assertIn(f"Row 3: id: Report {report.id} also appears on row 2", out.getvalue())
        report.refresh_from_db()
        self.assertEqual(report.name, "Probe Report")

    def test_one_invalid_row_leaves_every_other_row_unapplied(self):
        report = self._report()
        path = write_sheet(
            ["id", "name"],
            [{"id": report.id, "name": "Renamed"}, {"id": 987654, "name": "X"}],
        )
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_reports", path, user_email=self.editor.email, stdout=out)

        self.assertIn("nothing committed", out.getvalue())
        report.refresh_from_db()
        self.assertEqual(report.name, "Probe Report")

    def test_dry_run_rolls_back(self):
        report = self._report()
        path = write_sheet(["id", "name"], [{"id": report.id, "name": "Renamed"}])
        out = StringIO()
        call_command("import_reports", path, user_email=self.editor.email, dry_run=True, stdout=out)

        self.assertIn("DRY RUN", out.getvalue())
        report.refresh_from_db()
        self.assertEqual(report.name, "Probe Report")

    def test_a_row_that_changes_nothing_reports_no_changelog_line(self):
        report = self._report()
        path = write_sheet(["id", "name"], [{"id": report.id, "name": "Probe Report"}])
        out = StringIO()
        call_command("import_reports", path, user_email=self.editor.email, stdout=out)
        output = out.getvalue()

        self.assertNotIn("ROW_UPDATED", output)
        self.assertIn("1 of the updated rows had no effective change.", output)

    def test_a_sheet_naming_an_unknown_column_is_rejected(self):
        report = self._report()
        path = write_sheet(["id", "not_a_column"], [{"id": report.id, "not_a_column": "x"}])
        with self.assertRaises(CommandError) as caught:
            call_command("import_reports", path, user_email=self.editor.email, stdout=StringIO())

        self.assertIn("Unknown column(s): not_a_column", str(caught.exception))

    def test_an_unknown_user_email_is_refused(self):
        report = self._report()
        path = write_sheet(["id", "name"], [{"id": report.id, "name": "Renamed"}])
        with self.assertRaises(CommandError):
            call_command("import_reports", path, user_email="nobody@example.com", stdout=StringIO())

    # ----- GIDD visibility -----

    PFA_COLUMNS = [
        "id",
        "name",
        "filter_figure_countries",
        "filter_figure_start_after",
        "filter_figure_end_before",
        "filter_figure_categories",
        "filter_figure_crisis_types",
        "is_public",
        "is_pfa_visible_in_gidd",
    ]

    def _pfa_row(self, **overrides):
        row = {
            "id": None,
            "name": "GRID 2021 - Nepal (ND) - C",
            "filter_figure_countries": "NPL",
            "filter_figure_start_after": "2020-01-01",
            "filter_figure_end_before": "2020-12-31",
            "filter_figure_categories": "NEW_DISPLACEMENT",
            "filter_figure_crisis_types": "CONFLICT",
            "is_public": "Yes",
            "is_pfa_visible_in_gidd": "Yes",
        }
        row.update(overrides)
        return row

    def test_a_created_report_that_qualifies_becomes_visible_in_gidd(self):
        path = write_sheet(self.PFA_COLUMNS, [self._pfa_row()])
        call_command("import_reports", path, user_email=self.editor.email, stdout=StringIO())

        report = Report.objects.get(name="GRID 2021 - Nepal (ND) - C")
        self.assertTrue(report.is_pfa_visible_in_gidd)

    def test_a_created_report_that_does_not_qualify_is_silenced_not_refused(self):
        # No category: a PFA total is defined for exactly one of IDPs / Internal Displacements.
        path = write_sheet(
            self.PFA_COLUMNS,
            [self._pfa_row(name="No Category Report", filter_figure_categories=None)],
        )
        out = StringIO()
        call_command("import_reports", path, user_email=self.editor.email, stdout=out)

        self.assertIn("Created 1, updated 0.", out.getvalue())
        report = Report.objects.get(name="No Category Report")
        self.assertFalse(report.is_pfa_visible_in_gidd)

    def test_a_second_country_cannot_be_named_in_the_country_cell(self):
        # filter_figure_countries is declared list_values=False, so the sheet cannot express the
        # multi-country case the PFA rules reject - the cell fails to resolve first.
        path = write_sheet(
            self.PFA_COLUMNS,
            [self._pfa_row(name="Two Country Report", filter_figure_countries="NPL,IND")],
        )
        with self.assertRaises(CommandError):
            call_command("import_reports", path, user_email=self.editor.email, stdout=StringIO())
        self.assertFalse(Report.objects.filter(name="Two Country Report").exists())

    def test_a_non_public_report_does_not_become_visible_in_gidd(self):
        path = write_sheet(self.PFA_COLUMNS, [self._pfa_row(name="Private Report", is_public="No")])
        call_command("import_reports", path, user_email=self.editor.email, stdout=StringIO())

        report = Report.objects.get(name="Private Report")
        self.assertFalse(report.is_pfa_visible_in_gidd)

    def test_a_part_year_report_does_not_become_visible_in_gidd(self):
        path = write_sheet(
            self.PFA_COLUMNS,
            [self._pfa_row(name="Half Year Report", filter_figure_end_before="2020-06-30")],
        )
        call_command("import_reports", path, user_email=self.editor.email, stdout=StringIO())

        report = Report.objects.get(name="Half Year Report")
        self.assertFalse(report.is_pfa_visible_in_gidd)

    def test_an_update_can_turn_visibility_on(self):
        report = self._report(name="Later A PFA")
        report.filter_figure_categories = [Figure.FIGURE_CATEGORY_TYPES.IDPS]
        report.filter_figure_crisis_types = [Crisis.CRISIS_TYPE.DISASTER]
        report.is_public = True
        report.save()
        self.assertFalse(report.is_pfa_visible_in_gidd)

        path = write_sheet(["id", "is_pfa_visible_in_gidd"], [{"id": report.id, "is_pfa_visible_in_gidd": "Yes"}])
        call_command("import_reports", path, user_email=self.editor.email, stdout=StringIO())

        report.refresh_from_db()
        self.assertTrue(report.is_pfa_visible_in_gidd)

    def test_the_visibility_flag_is_writable_only_from_the_sheet(self):
        # ReportSerializer's fields become the report mutation inputs, where the flag would bypass
        # setPfaVisibleInGidd's admin permission.
        self.assertNotIn("is_pfa_visible_in_gidd", ReportSerializer().fields)
        self.assertIn("is_pfa_visible_in_gidd", ReportImportSerializer().fields)
        self.assertIn("is_pfa_visible_in_gidd", ImportReportsCommand().import_columns())

    # ----- template -----

    def test_make_template_writes_the_expected_columns(self):
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        call_command("import_reports", make_template=tmp.name, stdout=StringIO())

        workbook = load_workbook(tmp.name)
        self.assertIn("Data", workbook.sheetnames)
        headers = [cell.value for cell in workbook["Data"][1]]
        self.assertEqual(headers, ImportReportsCommand().import_columns())
        # Single match key, so it leads the sheet.
        self.assertEqual(headers[0], "id")
        self.assertEqual(ImportReportsCommand().match_column_names, ("id",))
