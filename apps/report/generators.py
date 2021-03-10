from openpyxl import Workbook


def generate_report_excel():
    from apps.report.models import Report
    # FIXME remove me
    generation = Report.objects.get(id=2).generations.first()
    wb = Workbook()

    for sheet_name, data in generation.get_excel_sheets_data().items():
        headers, data, formulae, aggregation = data

        ws = wb.create_sheet(sheet_name)
        # primary headers and data
        for idx, (header_key, header_val) in enumerate(headers.items()):
            ws.cell(column=idx + 1, row=1, value=header_val)
            for idy, datum in enumerate(data):
                ws.cell(column=idx + 1, row=idy + 2, value=datum[header_key])
        # secondary headers and data
        idx2 = 0
        for idx2, (header_key, formula) in enumerate(formulae.items()):
            # column starts at 1, hence idx+idx2+2
            ws.cell(column=idx + idx2 + 2, row=1, value=header_key)
            # list indexing starts at 0, hence idx+idx2+1
            for row, cell in enumerate(list(ws.columns)[idx + idx2 + 1], 1):
                if row == 1:
                    continue
                cell.value = formula.format(row=row)
        # add a gap
        column_at = idx + idx2 + 2
        ws.cell(column=column_at, row=1, value='')

        agg_headers = aggregation['headers']
        agg_data = aggregation['data']
        agg_formulae = aggregation['formulae']

        # primary headers and data
        for idx, (header_key, header_val) in enumerate(agg_headers.items()):
            ws.cell(column=column_at + idx + 1, row=1, value=header_val)
            for idy, datum in enumerate(agg_data):
                ws.cell(column=column_at + idx + 1, row=idy + 2, value=datum[header_key])
        # secondary headers and data
        for idx2, (header_key, formula) in enumerate(agg_formulae.items()):
            # column starts at 1, hence idx+idx2+2
            ws.cell(column=column_at + idx + idx2 + 2, row=1, value=header_key)
            # list indexing starts at 0, hence idx+idx2+1
            for row, cell in enumerate(list(ws.columns)[column_at + idx + idx2 + 1], 1):
                if row == 1:
                    continue
                cell.value = formula.format(row=row)

    wb.save(filename='temp.xlsx')
