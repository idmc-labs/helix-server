import itertools
import logging
import re
import time
from tempfile import NamedTemporaryFile

from django.core.files.base import File
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

# from helix.settings import QueuePriority
from helix.celery import app as celery_app

REPORT_TIMEOUT = 20 * 60 * 1000

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def generate_excel_file(excel_sheet_data):
    wb = Workbook()
    active = wb.active.title
    del wb[active]

    for sheet_name, sheet_data in excel_sheet_data:
        headers = sheet_data['headers']
        data = sheet_data['data']
        formulae = sheet_data['formulae']
        aggregation = sheet_data.get('aggregation', None)

        ws = wb.create_sheet(sheet_name)
        # primary headers and data
        for idx, (header_key, header_val) in enumerate(headers.items()):
            ws.cell(column=idx + 1, row=1, value=header_val)
            for idy, datum in enumerate(data):
                ws.cell(column=idx + 1, row=idy + 2, value=datum.get(header_key, ''))
        # secondary headers and data
        idx2 = 0
        for idx2, (header_key, formula) in enumerate(formulae.items()):
            # column starts at 1, hence idx+idx2+2
            ws.cell(column=idx + idx2 + 2, row=1, value=header_key)
            # list indexing starts at 0, hence idx+idx2+1
            for row, cell in enumerate(list(ws.columns)[idx + idx2 + 1], 1):
                if row == 1:
                    continue
                cell.value = formula.format(row=row)
        # add a gap
        column_at = idx + idx2 + 3
        ws.cell(column=column_at, row=1, value='')

        if not aggregation:
            continue
        agg_headers = aggregation['headers']
        agg_data = aggregation['data']
        agg_formulae = aggregation['formulae']

        # primary headers and data
        for idx, (header_key, header_val) in enumerate(agg_headers.items()):
            ws.cell(column=column_at + idx + 1, row=1, value=header_val)
            for idy, datum in enumerate(agg_data):
                ws.cell(column=column_at + idx + 1, row=idy + 2, value=datum.get(header_key, ''))
        # secondary headers and data
        for idx2, (header_key, formula) in enumerate(agg_formulae.items()):
            # column starts at 1, hence idx+idx2+2
            ws.cell(column=column_at + idx + idx2 + 2, row=1, value=header_key)
            # list indexing starts at 0, hence idx+idx2+1
            for row, cell in enumerate(list(ws.columns)[column_at + idx + idx2 + 1], 1):
                if row == 1:
                    continue
                cell.value = formula.format(row=row)

    return wb


def generate_report_excel(generation_id):
    from apps.report.models import ReportGeneration
    generation = ReportGeneration.objects.get(id=generation_id)
    excel_sheet_data = generation.get_excel_sheets_data().items()
    return generate_excel_file(excel_sheet_data)


def generate_report_snapshot(generation_id):
    from apps.report.models import ReportGeneration
    generation = ReportGeneration.objects.get(id=generation_id)
    snapshot = generation.get_snapshot()
    wb = Workbook(write_only=True)
    for sheet_name, sheet_data in snapshot.items():
        ws = wb.create_sheet(sheet_name)
        headers = list(set(itertools.chain.from_iterable(sheet_data)))
        ws.append(headers)
        for elements in sheet_data:
            ws.append([re.sub(ILLEGAL_CHARACTERS_RE, '', str(elements.get(header))) for header in headers])

    return wb


@celery_app.task(time_limit=REPORT_TIMEOUT)
def trigger_report_generation(generation_id):
    from apps.report.models import ReportGeneration
    generation = ReportGeneration.objects.get(id=generation_id)
    generation.started_at = timezone.now()
    generation.status = ReportGeneration.REPORT_GENERATION_STATUS.IN_PROGRESS
    generation.save()
    try:
        with transaction.atomic():
            then = time.time()
            path = f'{generation.report.name}.xlsx'

            logger.warn('Starting report generation...')
            workbook = generate_report_excel(generation_id)
            with NamedTemporaryFile(dir='/tmp') as tmp:
                workbook.save(tmp.name)
                workbook.close()
                file = File(tmp)
                generation.full_report.save(path, file)
                del workbook
            logger.warn(f'Completed report generation in {time.time() - then}')
            then = time.time()

            workbook = generate_report_snapshot(generation_id)
            with NamedTemporaryFile(dir='/tmp') as tmp:
                workbook.save(tmp.name)
                workbook.close()
                file = File(tmp)
                generation.snapshot.save(path, file)
                del workbook
            logger.warn(f'Completed snapshot generation {time.time() - then}')

            generation.status = ReportGeneration.REPORT_GENERATION_STATUS.COMPLETED
            generation.completed_at = timezone.now()
            generation.save()
    except Exception as e:  # NOQA E722
        logger.error('Report Generation Failed', exc_info=True)
        generation.status = ReportGeneration.REPORT_GENERATION_STATUS.FAILED
        generation.save(update_fields=['status'])
