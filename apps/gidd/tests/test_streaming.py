import json
import tempfile
from datetime import date
from unittest import mock

from django.core.files import File
from django.core.files.storage import FileSystemStorage
from django.test import SimpleTestCase
from openpyxl import Workbook, load_workbook

from apps.crisis.models import Crisis
from apps.entry.models import Figure
from apps.gidd import cache as gidd_cache
from apps.gidd.models import GiddEvent, GiddFigure, PublicFigureAnalysis
from apps.gidd.views import DisaggregationViewSet
from utils.factories import CountryFactory
from utils.streaming import spool_to_temp_file, stream_json_array, stream_json_object_with_array
from utils.tests import HelixTestCase


class StreamJsonArrayTest(SimpleTestCase):
    def _collect(self, items) -> list:
        return json.loads(b"".join(stream_json_array(items)))

    def test_matches_monolithic_json(self):
        items = [{"id": i, "name": f"row {i}"} for i in range(3)]

        self.assertEqual(self._collect(iter(items)), items)

    def test_empty_items(self):
        self.assertEqual(self._collect(iter([])), [])

    def test_encodes_dates_via_django_encoder(self):
        result = self._collect(iter([{"start_date": date(2023, 5, 6)}]))

        self.assertEqual(result, [{"start_date": "2023-05-06"}])

    def test_items_consumed_lazily(self):
        consumed = []

        def items():
            for i in range(3):
                consumed.append(i)
                yield {"id": i}

        chunks = stream_json_array(items())
        next(chunks)  # array-open
        self.assertEqual(consumed, [])
        next(chunks)  # first item
        self.assertEqual(consumed, [0])


class StreamJsonObjectWithArrayTest(SimpleTestCase):
    def _collect(self, **kwargs) -> dict:
        return json.loads(b"".join(stream_json_object_with_array(**kwargs)))

    def test_matches_monolithic_json(self):
        scalar = {"type": "FeatureCollection", "readme": "hi", "lastUpdated": "2024-01-01"}
        features = [{"type": "Feature", "id": i} for i in range(3)]

        result = self._collect(scalar_fields=scalar, array_key="features", items=iter(features))

        self.assertEqual(result, {**scalar, "features": features})

    def test_empty_items(self):
        scalar = {"type": "FeatureCollection"}

        result = self._collect(scalar_fields=scalar, array_key="features", items=iter([]))

        self.assertEqual(result, {"type": "FeatureCollection", "features": []})

    def test_empty_scalar_fields(self):
        result = self._collect(scalar_fields={}, array_key="features", items=iter([{"a": 1}]))

        self.assertEqual(result, {"features": [{"a": 1}]})

    def test_encodes_dates_via_django_encoder(self):
        # DjangoJSONEncoder is the default; it must serialise date/datetime.
        result = self._collect(
            scalar_fields={"lastUpdated": date(2024, 1, 2)},
            array_key="features",
            items=iter([{"start_date": date(2023, 5, 6)}]),
        )

        self.assertEqual(result["lastUpdated"], "2024-01-02")
        self.assertEqual(result["features"][0]["start_date"], "2023-05-06")

    def test_only_one_item_held_in_memory(self):
        # Later elements raise, so only lazy consumption gets through the join.
        consumed = []

        def items():
            for i in range(3):
                consumed.append(i)
                yield {"id": i}

        chunks = stream_json_object_with_array(
            scalar_fields={"type": "FeatureCollection"},
            array_key="features",
            items=items(),
        )
        # Nothing consumed before iteration starts.
        next(chunks)  # header
        next(chunks)  # array-open marker
        next(chunks)  # array bracket
        self.assertEqual(consumed, [])
        next(chunks)  # first item
        self.assertEqual(consumed, [0])


class SpoolToTempFileTest(SimpleTestCase):
    def test_round_trip(self):
        with spool_to_temp_file(iter([b"hello ", b"streamed ", b"world"])) as tmp:
            self.assertEqual(tmp.read(), b"hello streamed world")
            # Yielded rewound: a consumer must be able to read from the start.
            tmp.seek(0)
            self.assertEqual(tmp.read(6), b"hello ")
            name = tmp.name

        # Cleaned up on exit.
        with self.assertRaises(FileNotFoundError):
            open(name, "rb")

    def test_saved_through_django_storage(self):
        # Exercise the real integration point: Storage.save() wraps the object
        # in a File and drains it, the same path S3Boto3Storage and
        # FileSystemStorage take.
        payload = list(
            stream_json_object_with_array(
                scalar_fields={"type": "FeatureCollection", "lastUpdated": "2024-01-01"},
                array_key="features",
                items=iter([{"id": 1}, {"id": 2}]),
            )
        )
        expected = b"".join(payload)

        with tempfile.TemporaryDirectory() as tmp_dir:
            storage = FileSystemStorage(location=tmp_dir)
            with spool_to_temp_file(iter(payload)) as tmp:
                name = storage.save("export.geojson", File(tmp))
            with storage.open(name, "rb") as fp:
                written = fp.read()

        self.assertEqual(written, expected)
        self.assertEqual(json.loads(written)["features"], [{"id": 1}, {"id": 2}])


class DisaggregationExportSmokeTest(HelixTestCase):
    def setUp(self):
        super().setUp()
        self.country = CountryFactory.create(iso3="NPL", idmc_short_name="Nepal")
        self.gidd_event = GiddEvent.objects.create(
            name="Nepal Flood 2024",
            cause=Crisis.CRISIS_TYPE.DISASTER,
            event_codes=[],
            event_codes_type=[],
            event_codes_iso3=[],
        )
        self.gidd_figure = GiddFigure.objects.create(
            iso3="NPL",
            country_name="Nepal",
            country=self.country,
            year=2024,
            figure_raw_id=101,
            total_figures=100,
            reported=100,
            unit=Figure.UNIT.PERSON,
            cause=Crisis.CRISIS_TYPE.DISASTER,
            gidd_event=self.gidd_event,
            locations_coordinates=["12.3, 45.6"],
            locations_names=["Kathmandu"],
            locations_accuracy=[],
            locations_type=[],
        )

    def _figure_qs(self):
        return GiddFigure.objects.select_related("gidd_event").order_by("-year", "iso3", "id")

    def test_streamed_export_is_valid_geojson(self):
        chunks = DisaggregationViewSet()._export_disaggregated_geojson("export", self._figure_qs())
        doc = json.loads(b"".join(chunks))

        self.assertEqual(doc["type"], "FeatureCollection")
        self.assertEqual(len(doc["features"]), 1)
        feature = doc["features"][0]
        self.assertEqual(feature["geometry"], {"type": "MultiPoint", "coordinates": [[45.6, 12.3]]})
        self.assertEqual(feature["properties"]["ID"], self.gidd_figure.figure_raw_id)
        self.assertEqual(feature["properties"]["Country"], "Nepal")

    def test_excel_export_is_a_loadable_workbook(self):
        wb = DisaggregationViewSet()._export_disaggregated_excel(
            "export", self._figure_qs(), PublicFigureAnalysis.objects.none()
        )
        self.assertIsInstance(wb, Workbook)

        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            wb.close()
            loaded = load_workbook(tmp.name, read_only=True)
        sheet = loaded[loaded.sheetnames[0]]
        rows = list(sheet.iter_rows(max_row=2, values_only=True))
        self.assertEqual(rows[0][:3], ("ID", "ISO3", "Country"))
        self.assertEqual(rows[1][:3], (101, "NPL", "Nepal"))


class WorkbookThroughCacheTest(HelixTestCase):
    def test_workbook_generator_is_spooled_to_storage(self):
        # The cache must route Workbook payloads through a temp file — never
        # `save_virtual_workbook`-style whole-xlsx bytes in memory.
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Main")
        ws.append(["a", "b"])
        ws.append([1, 2])

        with tempfile.TemporaryDirectory() as tmp_dir:
            storage = FileSystemStorage(location=tmp_dir)
            with mock.patch.object(gidd_cache, "external_storage", storage):
                cache_key = gidd_cache.GiddExportCache._get_or_create(
                    gidd_cache.GiddExportCache.Key.DISAGGREGATION_EXPORT,
                    {"year": 2024},
                    "export.xlsx",
                    lambda: wb,
                )
            loaded = load_workbook(storage.path(cache_key), read_only=True)

        self.assertEqual(
            list(loaded["Main"].iter_rows(values_only=True)),
            [("a", "b"), (1, 2)],
        )
