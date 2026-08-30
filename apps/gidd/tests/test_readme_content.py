"""The README sheets are published prose, so they are pinned like any other published value.

Nothing read a README until these tests, which is how four hand-maintained copies of the same
document drifted apart and how the disaster workbook came to describe a tab it does not carry.
The assertions here are about the text the exports ship, not about the data underneath it.
"""

import io
import json
import re

import openpyxl
from django.test import TestCase

from apps.gidd.models import GiddDisplacement, GiddFigure, IdpsSaddEstimate, PublicFigureAnalysis
from apps.gidd.readme_revisions import SEPTEMBER_2026_NOTE
from apps.gidd.views import DisaggregationViewSet, DisasterViewSet, DisplacementDataViewSet

# A README that names a sheet spells it exactly as the tab is spelled.
SHEET_REFERENCE = re.compile(r"\b\d_[A-Za-z][A-Za-z0-9_]*")

# Cells of a revision table are rendered one per column, joined the way the GeoJSON dump joins them.
COLUMN_SEPARATOR = "\t"


def _readme_lines(workbook):
    """The README as one string per row, plus the tab names of the workbook carrying it.

    Rows are joined the way the GeoJSON dump renders the same rows, so a workbook README and the
    GeoJSON README compare as equals.
    """
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    loaded = openpyxl.load_workbook(buffer, read_only=True)
    lines = [
        COLUMN_SEPARATOR.join("" if cell is None else str(cell) for cell in row) if row else ""
        for row in loaded["README"].values
    ]
    return lines, loaded.sheetnames


def _between(lines, heading, next_heading_prefix):
    """The non-empty lines between a heading and the next one."""
    first = lines.index(heading)
    last = next(i for i in range(first + 1, len(lines)) if lines[i].startswith(next_heading_prefix))
    return [line for line in lines[first + 1 : last] if line]


def _revision_tables(lines):
    """round heading -> the table rows that round publishes."""
    tables, heading = {}, None
    for line in lines:
        if line.startswith("FIGURES REVIEWED IN "):
            heading = line
            tables[heading] = []
        elif heading:
            cells = line.split(COLUMN_SEPARATOR)
            if len(cells) == 6 and cells[0] != "ISO3":
                tables[heading].append(tuple(cells))
    return tables


class GiddReadmeTestCase(TestCase):
    """Every GIDD README, built off empty querysets -- the prose does not depend on the rows."""

    @classmethod
    def setUpTestData(cls):
        cls.disaster, cls.disaster_sheets = _readme_lines(DisasterViewSet()._export(GiddDisplacement.objects.none()))
        cls.displacement, cls.displacement_sheets = _readme_lines(
            DisplacementDataViewSet()._export(
                GiddDisplacement.objects.none(),
                PublicFigureAnalysis.objects.none(),
                IdpsSaddEstimate.objects.none(),
                None,
            )
        )
        cls.disaggregation, cls.disaggregation_sheets = _readme_lines(
            DisaggregationViewSet()._export_disaggregated_excel(
                "probe.xlsx", GiddFigure.objects.none(), PublicFigureAnalysis.objects.none()
            )
        )
        payload = json.loads(
            b"".join(DisaggregationViewSet()._export_disaggregated_geojson("probe", GiddFigure.objects.none())).decode()
        )
        cls.geojson = payload["readme"].split("\n")

    def legacy_readmes(self):
        """The three exports publishing the unversioned document.

        The displacement export is README version 4, which deliberately revised this prose for
        itself; its own blocks are compared where they are still shared.
        """
        return {"disaster": self.disaster, "disaggregation": self.disaggregation, "geojson": self.geojson}


class TestSharedReadmeBlocks(GiddReadmeTestCase):
    """A block several exports publish must be one text, not several that resemble each other."""

    def test_the_description_is_one_text_across_the_exports_that_share_it(self):
        sections = {
            name: _between(lines, "DESCRIPTION:", "KEY DEFINITIONS:") for name, lines in self.legacy_readmes().items()
        }
        self.assertEqual(len(sections["disaster"]), 4, sections["disaster"])
        self.assertEqual(sections["disaster"], sections["disaggregation"])
        self.assertEqual(sections["disaster"], sections["geojson"])

    def test_the_key_definitions_are_one_text_across_the_exports_that_share_them(self):
        sections = {
            name: _between(lines, "KEY DEFINITIONS:", "USE LICENSE:") for name, lines in self.legacy_readmes().items()
        }
        self.assertEqual(len(sections["disaggregation"]), 5, sections["disaggregation"])
        self.assertEqual(sections["disaggregation"], sections["geojson"])
        # The disaster workbook reports disaster flows alone, so its block is the shared one minus
        # the conflict trigger -- not a separately worded block that happens to be shorter.
        self.assertEqual(
            sections["disaster"],
            [line for line in sections["disaggregation"] if not line.startswith("Conflict displacement:")],
        )

    def test_the_use_license_is_one_text_across_the_exports_that_share_it(self):
        licenses = {
            name: [line for line in lines if line.startswith("USE LICENSE:")]
            for name, lines in self.legacy_readmes().items()
        }
        self.assertEqual([len(value) for value in licenses.values()], [1, 1, 1], licenses)
        self.assertEqual(len({tuple(value) for value in licenses.values()}), 1, licenses)

    def test_the_coverage_is_one_text_across_the_exports_that_share_it(self):
        coverage = {
            name: [line for line in lines if line.startswith("COVERAGE:")][0]
            for name, lines in self.legacy_readmes().items()
        }
        self.assertEqual(coverage["disaggregation"], coverage["geojson"])
        # The disaster workbook claims no conflict coverage, so this one is legitimately its own.
        self.assertNotIn("conflict", coverage["disaster"])

    def test_the_citation_names_the_database_each_export_actually_publishes(self):
        citations = {
            name: [line for line in lines if line.startswith("All derived work")][0]
            for name, lines in self.legacy_readmes().items()
        }
        self.assertEqual(citations["disaggregation"], citations["geojson"])
        self.assertIn("Global Internal Displacement Database - Disasters.", citations["disaster"])
        # The two spellings differ in the database name alone, not in the sentence around it.
        self.assertEqual(citations["disaster"].replace(" - Disasters.", "."), citations["disaggregation"])

    def test_the_revision_tables_are_one_set_across_the_exports_that_share_them(self):
        displacement = _revision_tables(self.displacement)
        disaggregation = _revision_tables(self.disaggregation)
        geojson = _revision_tables(self.geojson)
        self.assertTrue(displacement, "no revision round was published at all")
        self.assertEqual(displacement, disaggregation)
        self.assertEqual(displacement, geojson)

    def test_the_disaster_revision_tables_are_the_subset_that_workbook_can_report(self):
        full = _revision_tables(self.displacement)
        disaster = _revision_tables(self.disaster)
        self.assertEqual(sorted(full), sorted(disaster))
        for heading, rows in full.items():
            expected = [row for row in rows if row[4] == "Disaster" and row[5] == "Internal Displacements"]
            self.assertEqual(disaster[heading], expected, heading)


class TestReadmeNamesOnlyItsOwnSheets(GiddReadmeTestCase):
    """The general form of the defect where the disaster README pointed at a tab it lacks."""

    def test_no_readme_names_a_sheet_its_own_workbook_does_not_contain(self):
        workbooks = {
            "disaster": (self.disaster, self.disaster_sheets),
            "displacement": (self.displacement, self.displacement_sheets),
            "disaggregation": (self.disaggregation, self.disaggregation_sheets),
        }
        for name, (lines, sheetnames) in workbooks.items():
            named = set(SHEET_REFERENCE.findall("\n".join(lines)))
            self.assertTrue(named, f"the {name} README names no sheet at all, so this proves nothing")
            self.assertEqual(named - set(sheetnames), set(), f"the {name} README names sheets it does not carry")

    def test_the_geojson_readme_points_at_no_workbook_tab(self):
        # A GeoJSON dump is a single document, so a reader has no tab to be sent to.
        self.assertNotIn("2_Context_Displacement_data", "\n".join(self.geojson))


class TestIdpsNoteScope(GiddReadmeTestCase):
    """The September 2026 note explains an absence of IDPs rows, so it needs an IDPs metric."""

    def test_the_idps_note_appears_only_where_an_idps_metric_exists(self):
        for name, lines in (
            ("displacement", self.displacement),
            ("disaggregation", self.disaggregation),
            ("geojson", self.geojson),
        ):
            self.assertIn(SEPTEMBER_2026_NOTE, lines, f"the {name} README reports IDPs and must carry the note")
            self.assertTrue(
                [row for rows in _revision_tables(lines).values() for row in rows if row[5] == "IDPs"],
                f"the {name} README carries the note but lists no IDPs revision, so the note explains nothing",
            )

        self.assertNotIn(SEPTEMBER_2026_NOTE, self.disaster)
        self.assertFalse(
            [row for rows in _revision_tables(self.disaster).values() for row in rows if row[5] == "IDPs"],
            "the disaster workbook carries no IDPs metric, so it must list no IDPs revision",
        )
