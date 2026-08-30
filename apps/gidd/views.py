import typing
from datetime import datetime
from pathlib import Path

from django.core.exceptions import FieldDoesNotExist
from django.db import models
from django.db.models import Case, F, Q, Sum, When
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import (
    extend_schema,
    extend_schema_view,
)
from openpyxl import Workbook
from openpyxl.cell import Cell as OpCell
from rest_framework import filters, mixins, renderers, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.permissions import AllowAny

from apps.common.utils import (
    EXTERNAL_ARRAY_SEPARATOR,
    EXTERNAL_FIELD_SEPARATOR,
    get_enum_label,
)
from apps.contrib.commons import DATE_ACCURACY
from apps.country.models import Country
from apps.crisis.models import Crisis
from apps.entry.models import ExternalApiDump, Figure, FigureLocation
from apps.event.models import EventCode
from utils.common import client_id, get_valid_xml_string, track_gidd
from utils.db import rounded_figure_expr, tiebreak_fields
from utils.streaming import stream_json_object_with_array

from . import readme_blocks
from .cache import GiddExportCache
from .models import (
    GiddDisplacement,
    GiddEventDisplacement,
    GiddFigure,
    IdpsSaddEstimate,
    PublicFigureAnalysis,
    StatusLog,
)
from .paginations import GiddLimitOffsetPagination
from .readme_revisions import historical_revisions_block
from .rest_filters import (
    DisaggregationFilterSet,
    DisaggregationPublicFigureAnalysisFilterSet,
    IdpsSaddEstimateFilter,
    PublicFigureAnalysisFilterSet,
    RestConflictFilterSet,
    RestDisasterFilterSet,
    RestDisplacementDataFilterSet,
)
from .serializers import (
    ConflictSerializer,
    CountrySerializer,
    DisasterSerializer,
    DisplacementDataSerializer,
    PublicFigureAnalysisSerializer,
)


class XlsxRenderer(renderers.BaseRenderer):
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    format = "xlsx"
    charset = None
    render_style = "binary"

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data


def _get_location_accuracy_label(accuracy):
    if accuracy is None:
        return None
    return get_enum_label(FigureLocation.ACCURACY.get(accuracy))


def _get_pcode_accuracy_label(accuracy):
    if accuracy is None:
        return None
    return get_enum_label(FigureLocation.PCODE_ACCURACY.get(accuracy))


def _get_location_type_label(type):
    if type is None:
        return None
    return get_enum_label(FigureLocation.IDENTIFIER.get(type))


def _get_event_code_label(key: str):
    if key is None:
        return None
    return get_enum_label(EventCode.EVENT_CODE_TYPE.get(int(key)))


def _get_location_accuracy_labels(location_accuracy: typing.List[typing.Tuple[int]]) -> str:
    return _join_keeping_gaps([_get_location_accuracy_label(accuracy) for accuracy in location_accuracy])


def _join_keeping_gaps(values) -> str:
    """Join with a slot for every element, so each value stays under its own row element.

    The location columns and the source columns are positionally paired, and `string_join` drops
    `None`: one missing p-code or organisation kind would shift every later value one position
    left, silently pairing it with the wrong location or source.
    """
    return EXTERNAL_ARRAY_SEPARATOR.join("" if value is None else str(value) for value in values or [])


def _get_pcode_accuracy_labels(pcode_accuracy) -> str:
    return _join_keeping_gaps([_get_pcode_accuracy_label(accuracy) for accuracy in pcode_accuracy or []])


def _get_location_type_labels(location_type: typing.List[typing.Tuple[int]]) -> str:
    return _join_keeping_gaps([_get_location_type_label(type) for type in location_type])


def get_hyperlink(ws, url, text):
    # NOTE: 0, 0 will be updated by append
    # - https://openpyxl.readthedocs.io/en/3.1.3/_modules/openpyxl/worksheet/worksheet.html#Worksheet.append
    cell = OpCell(ws, 0, 0)
    clean_text = get_valid_xml_string(text.replace('"', '""'))
    cell.value = f'=HYPERLINK("{url}", "{clean_text}")'
    cell.style = "Hyperlink"
    return cell


def string_join(
    separator: str,
    data: typing.List[str],
) -> str:
    return separator.join([str(item) for item in data if item is not None])


def remove_null_from_dict(data: dict) -> dict:
    return {key: value for key, value in data.items() if value is not None}


class GiddOrderingFilter(filters.OrderingFilter):
    """Ordering bounded by what the serializer exposes, and always total.

    OFFSET paging over a sort with ties has no stable page boundary, so a row can repeat or be
    skipped across requests.

    TODO: index the sortable columns. None carries an index today, and these endpoints are
    unauthenticated, so a deep page pays a full sort.
    """

    def filter_queryset(self, request, queryset, view):
        if getattr(view, "action", None) in getattr(view, "ORDERING_UNSUPPORTED_ACTIONS", ()):
            if request.query_params.get(self.ordering_param):
                raise ValidationError({self.ordering_param: ["This action returns a fixed order and cannot be sorted."]})
            return queryset
        return super().filter_queryset(request, queryset, view)

    def get_ordering(self, request, queryset, view):
        ordering = super().get_ordering(request, queryset, view)
        if not ordering:
            # Exports iterate the queryset whole, so sorting one here would reorder a file nobody
            # asked to be sorted.
            if getattr(view, "paginator", None) is None:
                return ordering
            # `filter_queryset` applies this through `order_by()`, which REPLACES whatever the
            # queryset carries, so an endpoint's own default ordering is restated to survive.
            ordering = [key for key in queryset.query.order_by if isinstance(key, str)]
            if not ordering:
                return tiebreak_fields(queryset)
        return [*ordering, *tiebreak_fields(queryset, ordering)]

    def get_valid_fields(self, queryset, view, context={}):
        return [(term, term) for term in self._term_to_source(queryset, view, context)]

    def _term_to_source(self, queryset, view, context={}):
        """Accepted ordering term -> the ORM path it sorts on.

        DRF's own derivation is not reused because it drops sources that are model properties, and
        django_enumfield installs one per `EnumField`, which would leave `figure_cause` unsortable.
        The serializer field name is accepted alongside its source, since `country_name` is the
        only spelling a caller knows for `Country.idmc_short_name`.
        """
        serializer_class = view.get_serializer_class()
        mapping = {}
        for field_name, field in serializer_class(context=context).fields.items():
            if getattr(field, "write_only", False) or field.source == "*":
                continue
            source = field.source.replace(".", "__") or field_name
            try:
                queryset.model._meta.get_field(source)
            except FieldDoesNotExist:
                if source not in queryset.query.annotations:
                    continue
            mapping[source] = source
            mapping.setdefault(field_name, source)
        for term, source in getattr(serializer_class, "ORDERING_SOURCES", {}).items():
            try:
                queryset.model._meta.get_field(source)
            except FieldDoesNotExist:
                if source not in queryset.query.annotations:
                    continue
            mapping.setdefault(term, source)
        return mapping

    def remove_invalid_fields(self, queryset, fields, view, request):
        """Resolve aliases, and refuse an unknown term rather than dropping it.

        DRF discards what it cannot resolve and still answers 200, so a caller cannot tell the
        sort was skipped. The GraphQL lists already raise on the same input.
        """
        if getattr(view, "ordering_fields", self.ordering_fields) is not None:
            valid = super().remove_invalid_fields(queryset, fields, view, request)
            unknown = [field for field in fields if field not in valid]
            mapping = None
        else:
            mapping = self._term_to_source(queryset, view)
            valid, unknown = [], []
            for field in fields:
                prefix, bare = ("-", field[1:]) if field.startswith("-") else ("", field)
                if bare in mapping:
                    valid.append(prefix + mapping[bare])
                else:
                    unknown.append(field)
        if unknown:
            raise ValidationError({"ordering": [f"Invalid ordering field: {field.lstrip('-')}" for field in unknown]})
        return valid

    def get_schema_operation_parameters(self, view):
        """Publish the accepted sort keys, and omit the parameter where the action refuses it.

        The keys come from `_term_to_source`, the same map the request is validated against, so
        the documented set cannot drift from the accepted one. `track_gidd` skips a
        `swagger_fake_view`, so building the queryset here records no API usage.
        """
        if getattr(view, "action", None) in getattr(view, "ORDERING_UNSUPPORTED_ACTIONS", ()):
            return []
        parameters = super().get_schema_operation_parameters(view)
        terms = sorted(self._term_to_source(view.get_queryset(), view))
        if terms and parameters:
            parameters[0]["description"] = (
                f"{parameters[0].get('description', '').rstrip()} "
                f"Accepted keys: {', '.join(terms)}. "
                "Prefix a key with `-` to reverse it, and separate several with commas. "
                "An unrecognised key is rejected rather than ignored."
            ).strip()
        return parameters


# `ordering_fields` stays unset so the keys come from the serializer: "__all__" would admit every
# model column, including internal ones like `event_raw_id` that no response carries.
# `SearchFilter` is absent: no GIDD viewset defines `search_fields`, so it returned the queryset
# untouched while still publishing `search` in the schema.
GIDD_LIST_FILTER_BACKENDS = (DjangoFilterBackend, GiddOrderingFilter)


@client_id
class ListOnlyViewSetMixin(mixins.ListModelMixin, viewsets.GenericViewSet):
    filter_backends = GIDD_LIST_FILTER_BACKENDS

    def get(self, request, *args, **kwargs):
        return self.list(request, *args, **kwargs)

    def handle_exception(self, exc):
        """Answer an error as JSON even where the action's only renderer emits a spreadsheet.

        `XlsxRenderer` passes its data through untouched, so an error payload would otherwise reach
        the caller as the repr of a dict under a spreadsheet content type.
        """
        response = super().handle_exception(exc)
        if isinstance(getattr(self.request, "accepted_renderer", None), XlsxRenderer):
            self.request.accepted_renderer = renderers.JSONRenderer()
            self.request.accepted_media_type = renderers.JSONRenderer.media_type
        return response


@extend_schema_view(
    list=extend_schema(
        responses=CountrySerializer(many=True),
        tags=["GIDD"],
    ),
)
class CountryViewSet(ListOnlyViewSetMixin):
    serializer_class = CountrySerializer
    lookup_field = "iso3"
    filterset_fields = ["id"]
    pagination_class = GiddLimitOffsetPagination

    def get_queryset(self):
        track_gidd(
            self.request.GET.get("client_id"),
            ExternalApiDump.ExternalApiType.GIDD_COUNTRY_REST,
            viewset=self,
        )
        return Country.objects.all()


@extend_schema_view(
    list=extend_schema(
        responses=ConflictSerializer(many=True),
        tags=["GIDD"],
    ),
)
class ConflictViewSet(ListOnlyViewSetMixin):
    serializer_class = ConflictSerializer
    filterset_class = RestConflictFilterSet
    pagination_class = GiddLimitOffsetPagination

    def get_queryset(self):
        track_gidd(
            self.request.GET.get("client_id"),
            ExternalApiDump.ExternalApiType.GIDD_CONFLICT_REST,
            viewset=self,
        )
        # Rounded figures are not summable: the serializer recomputes them from these raw sums.
        return (
            GiddDisplacement.objects.filter(cause=Crisis.CRISIS_TYPE.CONFLICT)
            .values("iso3", "country_name", "year")
            .annotate(
                new_displacement=Sum("new_displacement"),
                total_displacement=Sum("total_displacement"),
            )
            .order_by("iso3", "year")
        )


@extend_schema_view(
    list=extend_schema(
        description=Path("docs/disaster/main-description.md").read_text(),
        responses=DisasterSerializer(many=True),
        tags=["GIDD"],
    ),
)
class DisasterViewSet(ListOnlyViewSetMixin):
    serializer_class = DisasterSerializer
    filterset_class = RestDisasterFilterSet
    pagination_class = GiddLimitOffsetPagination

    def get_queryset(self):
        api_type = ExternalApiDump.ExternalApiType.GIDD_DISASTER_REST
        if self.action == "export":
            api_type = ExternalApiDump.ExternalApiType.GIDD_DISASTER_EXPORT_REST

        track_gidd(
            self.request.GET.get("client_id"),
            api_type,
            viewset=self,
        )
        # The cross-country event codes this dump publishes are read from stored columns rather
        # than aggregated per request, which keeps this a plain streamable queryset.
        qs = GiddEventDisplacement.objects.filter(cause=Crisis.CRISIS_TYPE.DISASTER).order_by("iso3", "year", "event_raw_id")
        return qs

    @staticmethod
    def get_displacement_status(displacement_occurred: typing.List[int]) -> str:
        if not displacement_occurred:
            return ""
        elif Figure.DISPLACEMENT_OCCURRED.BEFORE.value in displacement_occurred:
            return "Displacement reporting preventive evacuations"
        return "Displacement without preventive evacuations reported"

    def _export(self, qs):
        """
        Export disaster
        """
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("1_Disaster_Displacement_data")
        ws.append(
            [
                "ISO3",
                "Country / Territory",
                "Year",
                "Event Name",
                "Date of Event (start)",
                "Disaster Internal Displacements",
                "Disaster Internal Displacements (Raw)",
                "Hazard Category",
                "Hazard Type",
                "Hazard Sub Type",
                "Event Codes (Code:Type)",
                "Event ID",
                "Displacement occurred",
            ]
        )

        for disaster in qs.iterator(chunk_size=2000):
            ws.append(
                [
                    disaster.iso3,
                    disaster.country_name,
                    disaster.year,
                    disaster.event_name,
                    disaster.start_date,
                    disaster.new_displacement_rounded,
                    disaster.new_displacement,
                    disaster.hazard_category_name,
                    disaster.hazard_type_name,
                    disaster.hazard_sub_type_name,
                    EXTERNAL_ARRAY_SEPARATOR.join(
                        [
                            f"{key}{EXTERNAL_FIELD_SEPARATOR}{value}"
                            for key, value in zip(disaster.all_country_event_codes, disaster.all_country_event_codes_type)
                        ]
                    ),
                    disaster.event_raw_id,
                    self.get_displacement_status(disaster.displacement_occurred),
                ]
            )

        ws2 = wb.create_sheet("README")
        readme_text = [
            *readme_blocks.preamble_block(
                title="Global Internal Displacement Database (GIDD) - Disasters",
                filename="IDMC_GIDD_Disasters_Internal_Displacement_Data",
                extracted_on=datetime.now().strftime("%B %d, %Y"),
                last_update=StatusLog.last_release_date(),
                description=readme_blocks.description_rows(),
                # This workbook holds disaster flows alone, so it neither defines conflict
                # displacement nor claims the conflict coverage the other exports do.
                definitions=readme_blocks.definition_rows(covers_conflict=False),
                coverage=readme_blocks.COVERAGE_DISASTERS_ONLY,
                citation=readme_blocks.citation_row(readme_blocks.DATABASE_NAME_DISASTERS),
            ),
            *readme_blocks.data_description_block("1_Disaster_Displacement_data"),
        ]

        for item in readme_text:
            ws2.append(item)

        readme_text_2 = [
            ["ISO3: Represents the ISO 3166-1 alpha-3 code. The code 'AB9' is assigned to the Abyei Area."],
            ["Country / Territory: Short name of the country or territory."],
            ["Year: Indicates the year for which displacement data are reported."],
            [
                "Event Name: Common or official event name for the event, if available. Otherwise, events are coded "
                "based on the country, type of hazard, location, and event start date."
            ],
            ["Date of Event (Start): Approximate start date of the event."],
            [
                "Disaster Internal Displacements: Total number of internal displacements reported (rounded figures at "
                "national level), as a result of disasters over the reporting year. Units are recorded as 'internal "
                "displacement flows' or 'internal displacement movements'."
            ],
            [
                "Disaster Internal Displacements raw: Total number of internal displacements reported (not rounded), "
                "as a result of disasters over the reporting year. Units are recorded as 'internal displacement flows' "
                "or 'internal displacement movements'."
            ],
            ["Hazard Category: Hazard category based on the CRED EM-DAT classification."],
            ["Hazard Type: Hazard type as categorized by CRED EM-DAT."],
            ["Hazard Sub-Type: Specific sub-type of the hazard based on the CRED EM-DAT classification."],
            [
                "Event Codes (Code:Type): Unique codes such as the GLIDE number and other database-specific codes used "
                "to identify and track specific events across various databases."
            ],
            ["Event ID: Unique identifier for events as assigned by IDMC."],
            [
                "Displacement Occurred: This field contains values that represent if preventive evacuations were reported. "
                "These evacuations are the result of existing early warning systems."
            ],
        ]

        for item in readme_text_2:
            ws2.append(item)

        readme_text_3 = [
            [],
            # No 2_Context_Displacement_data sheet here, so the intro must not point at one.
            *historical_revisions_block(cause="Disaster", category="Internal Displacements"),
        ]

        for item in readme_text_3:
            ws2.append(item)
        return wb

    @extend_schema(
        description=Path("docs/disaster/xlsx-export-description.md").read_text(),
        responses={
            (200, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"): OpenApiTypes.BINARY,
        },
        filters=True,
        tags=["GIDD"],
    )
    @action(
        detail=False,
        methods=["get"],
        url_path="disaster-export",
        permission_classes=[AllowAny],
        pagination_class=None,
        renderer_classes=[XlsxRenderer],
    )
    def export(self, request):
        qs = self.filter_queryset(self.get_queryset())
        filename = "IDMC_GIDD_Disasters_Internal_Displacement_Data.xlsx"

        return GiddExportCache.get_or_create(
            filename,
            request,
            [self.filterset_class],
            GiddExportCache.Key.DISASTER_EXPORT,
            lambda: self._export(qs),
            s3_parameters={
                "ResponseContentDisposition": f"attachment; filename={filename}",
                "ResponseContentType": "application/octet-stream",
            },
        )


@extend_schema_view(
    list=extend_schema(
        description=Path("docs/displacement/main-description.md").read_text(),
        responses=DisplacementDataSerializer(many=True),
        tags=["GIDD"],
    )
)
class DisplacementDataViewSet(ListOnlyViewSetMixin):
    serializer_class = DisplacementDataSerializer
    filterset_class = RestDisplacementDataFilterSet
    pagination_class = GiddLimitOffsetPagination
    # `export()` re-sorts for the sheet layout, so a requested ordering cannot survive it: the
    # parameter is refused there rather than accepted and replaced.
    ORDERING_UNSUPPORTED_ACTIONS = frozenset({"export"})

    def get_queryset(self):
        api_type = ExternalApiDump.ExternalApiType.GIDD_DISPLACEMENT_REST
        if self.action == "export":
            api_type = ExternalApiDump.ExternalApiType.GIDD_DISPLACEMENT_EXPORT_REST

        track_gidd(
            self.request.GET.get("client_id"),
            api_type,
            viewset=self,
        )
        # A cause with no rows sums to NULL, not 0, which is what this endpoint publishes.
        # Rounded figures are not summable: the serializer recomputes them for the list endpoint,
        # and `export()` derives them in SQL.
        return (
            GiddDisplacement.objects.values("iso3", "country_name", "year")
            .annotate(
                conflict_new_displacement=Sum("new_displacement", filter=Q(cause=Crisis.CRISIS_TYPE.CONFLICT)),
                conflict_total_displacement=Sum("total_displacement", filter=Q(cause=Crisis.CRISIS_TYPE.CONFLICT)),
                disaster_new_displacement=Sum("new_displacement", filter=Q(cause=Crisis.CRISIS_TYPE.DISASTER)),
                disaster_total_displacement=Sum("total_displacement", filter=Q(cause=Crisis.CRISIS_TYPE.DISASTER)),
            )
            .order_by("iso3", "year")
        )

    def export_conflicts(self, ws, qs):
        ws.append(
            [
                "ISO3",
                "Name",
                "Year",
                "Conflict Stock Displacement",
                "Conflict Stock Displacement (Raw)",
                "Conflict Internal Displacements",
                "Conflict Internal Displacements (Raw)",
            ]
        )
        for item in qs.iterator(chunk_size=2000):
            ws.append(
                [
                    item["iso3"],
                    item["country_name"],
                    item["year"],
                    item["conflict_total_displacement_rounded"],
                    item["conflict_total_displacement"],
                    item["conflict_new_displacement_rounded"],
                    item["conflict_new_displacement"],
                ]
            )

    def export_disasters(self, ws, qs):
        ws.append(
            [
                "ISO3",
                "Name",
                "Year",
                "Disaster Internal Displacements",
                "Disaster Internal Displacements (Raw)",
                "Disaster Stock Displacement",
                "Disaster Stock Displacement (Raw)",
            ]
        )
        for item in qs.iterator(chunk_size=2000):
            ws.append(
                [
                    item["iso3"],
                    item["country_name"],
                    item["year"],
                    item["disaster_new_displacement_rounded"],
                    item["disaster_new_displacement"],
                    item["disaster_total_displacement_rounded"],
                    item["disaster_total_displacement"],
                ]
            )

    def export_displacements(self, ws, qs):
        ws.append(
            [
                "ISO3",
                "Name",
                "Year",
                "Conflict Stock Displacement",
                "Conflict Stock Displacement (Raw)",
                "Conflict Internal Displacements",
                "Conflict Internal Displacements (Raw)",
                "Disaster Internal Displacements",
                "Disaster Internal Displacements (Raw)",
                "Disaster Stock Displacement",
                "Disaster Stock Displacement (Raw)",
            ]
        )
        for item in qs.iterator(chunk_size=2000):
            ws.append(
                [
                    item["iso3"],
                    item["country_name"],
                    item["year"],
                    item["conflict_total_displacement_rounded"],
                    item["conflict_total_displacement"],
                    item["conflict_new_displacement_rounded"],
                    item["conflict_new_displacement"],
                    item["disaster_new_displacement_rounded"],
                    item["disaster_new_displacement"],
                    item["disaster_total_displacement_rounded"],
                    item["disaster_total_displacement"],
                ]
            )

    def _export(self, qs, pfa_qs, idps_sadd_qs, request_cause):
        """
        Export displacements, conflict and disaster
        """

        wb = Workbook(write_only=True)
        # Tab 1
        ws = wb.create_sheet("1_Displacement_data")

        if request_cause and request_cause.lower() == "conflict":
            self.export_conflicts(ws, qs)
        elif request_cause and request_cause.lower() == "disaster":
            self.export_disasters(ws, qs)
        else:
            self.export_displacements(ws, qs)
        # Tab 2
        ws2 = wb.create_sheet("2_Context_Displacement_data")
        ws2.append(
            [
                "ISO3",
                "Year",
                "Figure cause",
                "Figure category",
                "Description",
                "Figures",
                "Figures rounded",
            ]
        )

        for item in pfa_qs.iterator(chunk_size=2000):
            ws2.append(
                [
                    item.iso3,
                    item.year,
                    get_enum_label(item.figure_cause),
                    get_enum_label(item.figure_category),
                    item.description,
                    item.figures,
                    item.figures_rounded,
                ]
            )
        # Tab 3
        ws3 = wb.create_sheet("3_IDPs_SADD_estimates")
        ws3.append(
            [
                "ISO3",
                "Country",
                "Year",
                "Sex",
                "Cause",
                "0-4",
                "5-11",
                "12-17",
                "18-59",
                "60+",
            ]
        )

        for item in idps_sadd_qs.iterator(chunk_size=2000):
            ws3.append(
                [
                    item.iso3,
                    item.country_name,
                    item.year,
                    item.sex,
                    get_enum_label(item.cause),
                    item.zero_to_four,
                    item.five_to_eleven,
                    item.twelve_to_seventeen,
                    item.eighteen_to_fiftynine,
                    item.sixty_plus,
                ]
            )
        # Tab 4
        ws4 = wb.create_sheet("README")
        readme_text = [
            *readme_blocks.preamble_block(
                title="Global Internal Displacement Database (GIDD)",
                filename="IDMC_Internal_Displacement_Conflict-Violence_Disasters",
                extracted_on=datetime.now().strftime("%B %d, %Y"),
                last_update=StatusLog.last_release_date(),
                # README version 4 revised the shared prose for this export alone: it reports
                # through 2024, states the year-end IDPs rule, and documents the API.
                version="4",
                description=[[], [readme_blocks.DISPLACEMENT_DESCRIPTION]],
                definitions=readme_blocks.DISPLACEMENT_DEFINITIONS,
                coverage=readme_blocks.COVERAGE_ALL_CAUSES,
                license_row=readme_blocks.DISPLACEMENT_LICENSE_ROW,
            ),
            *readme_blocks.data_description_block("1_Displacement_data table"),
        ]

        for item in readme_text:
            ws4.append(item)

        readme_text_2 = [
            [
                "Annually validated data on internal displacement caused by disasters, conflicts, "
                "and other situations of violence, as compiled and reported by IDMC."
            ],
            [],
            ["ISO3: ISO 3166-1 alpha-3 code. AB9 = Abyei Area."],
            ["Name: Short name of the country or territory."],
            ["Year: Year for which displacement data are reported."],
            [
                "Conflict Stock Displacement: Total number of IDPs (rounded, national level), conflict "
                "and violence, end of reporting year. Units: People."
            ],
            [
                "Conflict Stock Displacement (Raw): Total number of IDPs (not rounded), conflict and "
                "violence, end of reporting year. Units: People."
            ],
            [
                "Conflict Internal Displacements: Total internal displacements reported (rounded, "
                "national level), conflict and violence, over the reporting year. "
                "Units: flows / movements."
            ],
            [
                "Conflict Internal Displacements (Raw): Total internal displacements (not rounded), "
                "conflict and violence, over the reporting year. Units: flows / movements."
            ],
            [
                "Disaster Internal Displacements: Total internal displacements reported (rounded, "
                "national level), disasters, over the reporting year. "
                "Units: flows / movements."
            ],
            [
                "Disaster Internal Displacements (Raw): Total internal displacements (not rounded), "
                "disasters, over the reporting year. Units: flows / movements."
            ],
            [
                "Disaster Stock Displacement: Total number of IDPs (rounded, national level), disasters, "
                "end of reporting year. Units: People."
            ],
            [
                "Disaster Stock Displacement (Raw): Total number of IDPs (not rounded), disasters, end of "
                "reporting year. Units: People."
            ],
        ]
        for item in readme_text_2:
            ws4.append(item)
        for item in readme_blocks.data_description_block("2_Context_Displacement_data table"):
            ws4.append(item)
        readme_text_3 = [
            [
                "Contextual information and analysis documented by IDMC analysts. Captures flags related "
                "to methodology, caveats, sources, and challenges identified for each metric, reporting year, "
                "and country."
            ],
            [],
            ["ISO3: ISO 3166-1 alpha-3 code. AB9 = Abyei Area."],
            ["Year: Year for which displacement data are reported."],
            ["Figure cause: Trigger of displacement: Conflict or Disaster."],
            ["Figure category: Type of metric: Internal Displacements (flows) or IDPs (stocks)."],
            ["Description: Contextual information including sources, data limitations, methodology, and caveats."],
            ["Figures: Total number of internal displacements or IDPs."],
            [
                "Figures rounded: Rounded figures matching values reported in the Global Report on "
                "Internal Displacement (GRID)."
            ],
        ]

        for item in readme_text_3:
            ws4.append(item)
        for item in readme_blocks.data_description_block("3_IDPs_SADD_estimates table"):
            ws4.append(item)
        ws4.append(
            [
                "Sex and Age Disaggregated Data (SADD) is often scarce. IDMC employs UN Population "
                "Estimates and Projections to break down internally displaced people by sex and age. "
                "Methodology and limitations: https://www.internal-displacement.org/monitoring-tools"
            ]
        )
        ws4.append([])
        readme_text_4 = [
            ["ISO3: ISO 3166-1 alpha-3 code. AB9 = Abyei Area."],
            ["Country: Short name of the country or territory."],
            ["Year: Year for which displacement figures are reported."],
            ["Sex: Female / Male / Both Sexes (UN DESA classification)."],
            ["Cause: Trigger of displacement: Conflict or Disaster."],
            ["Age 0-4: Newborns to 4 years old."],
            ["Age 5-11: Children aged 5 to 11."],
            ["Age 12-17: Adolescents aged 12 to 17."],
            ["Age 18-59: Adults aged 18 to 59."],
            ["Age 60+: Population aged 60 and older."],
        ]
        for item in readme_text_4:
            ws4.append(item)

        readme_text_5 = [
            ["SADD METHODOLOGY & LIMITATIONS"],
            [],
            ["OVERVIEW"],
            [
                "Sex and Age Disaggregated Data (SADD) for displacement associated with conflict or "
                "disasters is often scarce. One way to estimate it is to use SADD available at the "
                "national level. IDMC employs United Nations Population Estimates and Projections to "
                "break down the number of internally displaced people (IDPs) by sex and age."
            ],
            [],
            ["DATASETS USED"],
            [
                "Population data: Obtained from the United Nations Population Estimates and Projections "
                "(World Population Prospects). Two datasets were used: \n"
                "    • Population on 01 January, by single age (1950–2021)\n"
                "    • Population on 01 January, by single age (2022–2100)"
            ],
            [
                "Displacement data: The number of internally displaced people (IDPs) was extracted from IDMC's "
                "Global Database on Internal Displacement (GIDD)."
            ],
            [],
            ["LIMITATIONS"],
            [
                "Accuracy & representativeness: Using absolute population values to estimate SADD for IDPs could "
                "lead to issues of accuracy and representativeness, as displaced populations may differ structurally "
                "from the national population."
            ],
            [
                "Static population assumption: This method does not capture the dynamic nature of displacement; "
                "it assumes a static population structure over time and does not account for shifts in demographics "
                "caused by the displacement itself."
            ],
            [
                "National-level statistics: Using national-level statistics derived from census data may not fully "
                "reflect the experiences and needs of different groups — such as men and women, or different age "
                "cohorts — who may be disproportionately affected by conflict or disaster situations."
            ],
            [
                "Interpretive caution: Results should be interpreted with caution. Other sources of "
                "information — including qualitative data and local knowledge — should be consulted to gain a more "
                "comprehensive understanding of the displacement situation."
            ],
        ]
        ws4.append([])
        for item in readme_text_5:
            ws4.append(item)

        readme_text_6 = historical_revisions_block(context_tab="2_Context_Displacement_data")
        ws4.append([])
        for item in readme_text_6:
            ws4.append(item)

        ws4.append([])
        readme_text_7 = [
            ["CHANGELOG"],
            [],
            ["Version", "Date", "Notes"],
            [
                "4",
                "September 1, 2026",
                "This release includes updates from the Historical data revision project. All IDP stock figures "
                "for 2008–2016 have been updated in the September 2026 release. Because every IDPs figure was "
                "updated, only Internal Displacements changes are listed in Figures reviewed.",
            ],
            [
                "3.1",
                "May 12, 2026",
                "Added Q2 2026 revision round (13 figures). Standardized Disasters -> Disaster in revision tables.",
            ],
            ["3.0", "May 13, 2025", "Added May 2025 revision round (61 figures)."],
            ["2.0", "Jan 2025", "Added January 2025 revision round (24 figures)."],
        ]
        for item in readme_text_7:
            ws4.append(item)

        return wb

    @extend_schema(
        description=Path("docs/displacement/xlsx-export-description.md").read_text(),
        responses={
            (200, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"): OpenApiTypes.BINARY,
        },
        filters=True,
        tags=["GIDD"],
    )
    @action(
        detail=False,
        methods=["get"],
        url_path="displacement-export",
        permission_classes=[AllowAny],
        pagination_class=None,
        renderer_classes=[XlsxRenderer],
    )
    def export(self, request):
        # Track export
        # Rounded in SQL rather than in python: the sheet builders stream this with `.iterator()`,
        # which needs it to stay a queryset.
        # `order_by` below replaces any requested ordering, so `ordering` is refused by the filter
        # backend rather than accepted and silently dropped.
        qs = (
            self.filter_queryset(self.get_queryset())
            .annotate(
                conflict_new_displacement_rounded=rounded_figure_expr("conflict_new_displacement"),
                conflict_total_displacement_rounded=rounded_figure_expr("conflict_total_displacement"),
                disaster_new_displacement_rounded=rounded_figure_expr("disaster_new_displacement"),
                disaster_total_displacement_rounded=rounded_figure_expr("disaster_total_displacement"),
            )
            .order_by("-year", "iso3")
        )

        request_cause = request.GET.get("cause")

        pfa_qs = PublicFigureAnalysisFilterSet(
            data=self.request.query_params, queryset=PublicFigureAnalysis.objects.all()
        ).qs.order_by("iso3", "year")

        idps_sadd_qs = IdpsSaddEstimateFilter(
            data=self.request.query_params,
            queryset=IdpsSaddEstimate.objects.all(),
        ).qs.order_by("iso3", "year")

        filename = "IDMC_Internal_Displacement_Conflict-Violence_Disasters.xlsx"
        return GiddExportCache.get_or_create(
            filename,
            request,
            [self.filterset_class, PublicFigureAnalysisFilterSet, IdpsSaddEstimateFilter],
            GiddExportCache.Key.DISPLACEMENT_EXPORT,
            lambda: self._export(qs, pfa_qs, idps_sadd_qs, request_cause),
            s3_parameters={
                "ResponseContentDisposition": f"attachment; filename={filename}",
                "ResponseContentType": "application/octet-stream",
            },
        )


# Only the columns the disaggregation exports read — `.values()` skips model
# hydration and the wide unused columns (calculation_logic, source_excerpt, ...).
DISAGGREGATION_EXPORT_VALUES = (
    "figure_raw_id",
    "iso3",
    "country_name",
    "geographical_region_name",
    "cause",
    "year",
    "category",
    "unit",
    "reported",
    "household_size",
    "total_figures",
    "disaster_category_name",
    "disaster_sub_category_name",
    "disaster_type_name",
    "disaster_sub_type_name",
    "violence_name",
    "other_sub_type_name",
    "start_date",
    "start_date_accuracy",
    "end_date",
    "end_date_accuracy",
    "stock_date",
    "stock_date_accuracy",
    "stock_reporting_date",
    "publishers",
    "sources",
    "sources_type",
    "is_housing_destruction",
    "locations_coordinates",
    "locations_names",
    "locations_accuracy",
    "locations_type",
    "locations_pcode",
    "locations_pcode_accuracy",
    "locations_pcode_source",
    "displacement_occurred",
    "event_main_trigger",
    "gidd_event__event_raw_id",
    "gidd_event__name",
    "gidd_event__cause",
    "gidd_event__start_date",
    "gidd_event__end_date",
    "gidd_event__start_date_accuracy",
    "gidd_event__end_date_accuracy",
    "gidd_event__event_codes",
    "gidd_event__event_codes_type",
    "gidd_event__event_codes_iso3",
)


@client_id
class DisaggregationViewSet(viewsets.GenericViewSet):
    # `@client_id` is declared here rather than inherited: it documents the required query
    # parameter, and `ListOnlyViewSetMixin` was the only thing carrying it. Both actions demand a
    # registered client at runtime via `track_gidd`, so a schema that omits the parameter describes
    # an endpoint nobody can call.
    #
    # Only the two export actions are routed (helix/external_urls.py). This was a
    # ListOnlyViewSetMixin with pagination_class = None — an UNPAGINATED list over the whole
    # GiddFigure table, unrouted but one router.register away from shipping; drop the list
    # action instead of leaving the footgun.
    # Ordered with the pk as tiebreak: the exports stream this queryset whole, and an unordered
    # scan hands back rows in plan order, so the same data published twice differs by moved rows.
    queryset = GiddFigure.objects.all().order_by("iso3", "year", "id")
    filter_backends = (DjangoFilterBackend,)
    filterset_class = DisaggregationFilterSet

    def _generate_export_filename(self):
        filename_map = {
            Crisis.CRISIS_TYPE.CONFLICT.name.lower(): "IDMC_GIDD_Conflict_Internal_Displacement_Disaggregated",
            Crisis.CRISIS_TYPE.DISASTER.name.lower(): "IDMC_GIDD_Disasters_Internal_Displacement_Disaggregated",
        }

        filter_cause = self.request.query_params.get("cause", "").lower()
        return filename_map.get(filter_cause, "IDMC_GIDD_Internal_Displacement_Disaggregated")

    def _get_category(self, category) -> typing.Optional[str]:
        if category is None:
            return
        return get_enum_label(Figure.FIGURE_CATEGORY_TYPES.get(category))

    def _get_cause(self, cause) -> typing.Optional[str]:
        if cause is None:
            return
        return get_enum_label(Crisis.CRISIS_TYPE.get(cause))

    def _get_date_accuracy(self, accuracy) -> typing.Optional[str]:
        if accuracy is None:
            return
        return get_enum_label(DATE_ACCURACY.get(accuracy))

    def _get_displacement_occurred(self, displacement_occurred) -> str:
        if displacement_occurred is not None:
            return DisasterViewSet.get_displacement_status([displacement_occurred])
        return ""

    def _get_unit(self, unit) -> typing.Optional[str]:
        if unit is None:
            return None
        return get_enum_label(Figure.UNIT.get(unit))

    def extract_event_data(
        self,
        event_code: typing.List[typing.Tuple[str]],
        event_code_type: typing.List[typing.Tuple[int]],
        event_code_iso3: typing.List[typing.Tuple[str]],
        filter_iso3: str,
    ) -> str:
        event_code_components = [event_code, event_code_type, event_code_iso3]
        transposed_components = zip(*event_code_components)

        return EXTERNAL_ARRAY_SEPARATOR.join(
            EXTERNAL_FIELD_SEPARATOR.join([loc[0], _get_event_code_label(loc[1])])
            for loc in transposed_components
            if loc[2] == filter_iso3
        )

    def extract_event_data_raw(
        self,
        event_code: typing.List[typing.Tuple[str]],
        event_code_type: typing.List[typing.Tuple[int]],
        event_code_iso3: typing.List[typing.Tuple[str]],
        filter_iso3: str,
    ) -> str:
        event_code_components = [event_code, event_code_type, event_code_iso3]
        transposed_components = zip(*event_code_components)

        return [[loc[0], _get_event_code_label(loc[1])] for loc in transposed_components if loc[2] == filter_iso3]

    def _export_disaggregated_geojson(self, filename, qs):
        def format_coordinate(coordinate: str) -> typing.Tuple[float, float]:
            lat, lng = coordinate.split(", ")
            return (float(lng), float(lat))

        def format_coordinates(coordinates: typing.List[str]):
            return [format_coordinate(x) for x in coordinates]

        qs = qs.exclude(Q(locations_coordinates__isnull=True) | Q(locations_coordinates=[])).annotate(
            event_main_trigger=Case(
                When(gidd_event__cause=Crisis.CRISIS_TYPE.CONFLICT, then=F("gidd_event__violence__name")),
                When(gidd_event__cause=Crisis.CRISIS_TYPE.DISASTER, then=F("gidd_event__disaster_sub_type__name")),
                When(gidd_event__cause=Crisis.CRISIS_TYPE.OTHER, then=F("gidd_event__other_sub_type__name")),
                output_field=models.CharField(),
            ),
        )
        now = timezone.now().strftime("%B %d, %Y")

        readme_text = readme_blocks.rows_to_text(
            [
                *readme_blocks.preamble_block(
                    title="Disasters Global Internal Displacement Database (GIDD)",
                    filename=filename,
                    extracted_on=now,
                    last_update=StatusLog.last_release_date(),
                    description=readme_blocks.description_rows(),
                    definitions=readme_blocks.definition_rows(),
                    coverage=readme_blocks.COVERAGE_ALL_CAUSES,
                ),
                *readme_blocks.data_description_block("1_Disaggregated_Data table"),
                *readme_blocks.DISAGGREGATION_FIELD_ROWS,
                [],
                [],
                # A GeoJSON dump is a single document with no tabs, so the intro must not send a
                # reader to the sheet the companion workbook carries the change account on.
                *historical_revisions_block(),
            ]
        )
        scalar_fields = {
            "type": "FeatureCollection",
            "readme": readme_text,
            "lastUpdated": StatusLog.last_release_date(format="%Y-%m-%d"),
        }

        def feature_iterator():
            # ``.iterator()`` streams rows from a server-side cursor so the whole
            # queryset is never held in memory; ``stream_json_object_with_array``
            # then encodes one feature at a time.
            for item in qs.values(*DISAGGREGATION_EXPORT_VALUES).iterator(chunk_size=2000):
                yield {
                    "type": "Feature",
                    "geometry": {
                        "type": "MultiPoint",
                        "coordinates": format_coordinates(item["locations_coordinates"]),
                    },
                    "properties": remove_null_from_dict(
                        {
                            "ID": item["figure_raw_id"],
                            "ISO3": item["iso3"],
                            "Country": item["country_name"],
                            "Geographical region": item["geographical_region_name"],
                            "Figure cause": self._get_cause(item["cause"]),
                            "Year": item["year"],
                            "Figure category": self._get_category(item["category"]),
                            "Figure unit": self._get_unit(item["unit"]),
                            "Reported figures": item["reported"],
                            "Household size": item["household_size"],
                            "Total figures": item["total_figures"],
                            "Hazard category": item["disaster_category_name"],
                            "Hazard sub category": item["disaster_sub_category_name"],
                            "Hazard type": item["disaster_type_name"],
                            "Hazard sub type": item["disaster_sub_type_name"],
                            "Violence type": item["violence_name"],
                            "Other event sub type": item["other_sub_type_name"],
                            "Start date": item["start_date"],
                            "Start date accuracy": self._get_date_accuracy(item["start_date_accuracy"]),
                            "End date": item["end_date"],
                            "End date accuracy": self._get_date_accuracy(item["end_date_accuracy"]),
                            "Stock date": item["stock_date"],
                            "Stock date accuracy": self._get_date_accuracy(item["stock_date_accuracy"]),
                            "Stock reporting date": item["stock_reporting_date"],
                            "Publishers": item["publishers"],
                            "Sources": item["sources"],
                            "Sources type": item["sources_type"],
                            "Event ID": item["gidd_event__event_raw_id"],
                            "Event name": item["gidd_event__name"],
                            "Event cause": self._get_cause(item["gidd_event__cause"]),
                            "Event main trigger": item["event_main_trigger"],
                            "Event start date": item["gidd_event__start_date"],
                            "Event end date": item["gidd_event__end_date"],
                            "Event start date accuracy": self._get_date_accuracy(item["gidd_event__start_date_accuracy"]),
                            "Event end date accuracy": self._get_date_accuracy(item["gidd_event__end_date_accuracy"]),
                            "Is housing destruction": "Yes" if item["is_housing_destruction"] else "No",
                            "Event codes (Code:Type)": self.extract_event_data_raw(
                                item["gidd_event__event_codes"],
                                item["gidd_event__event_codes_type"],
                                item["gidd_event__event_codes_iso3"],
                                item["iso3"],
                            ),
                            "Locations name": item["locations_names"],
                            "Locations accuracy": [_get_location_accuracy_label(x) for x in item["locations_accuracy"]],
                            "Locations type": [_get_location_type_label(x) for x in item["locations_type"]],
                            "Pcode": item["locations_pcode"],
                            "Pcode accuracy": [_get_pcode_accuracy_label(x) for x in item["locations_pcode_accuracy"]],
                            "Pcode source": item["locations_pcode_source"],
                            "Displacement occurred": self._get_displacement_occurred(item["displacement_occurred"]),
                        }
                    ),
                }

        return stream_json_object_with_array(
            scalar_fields=scalar_fields,
            array_key="features",
            items=feature_iterator(),
        )

    def _export_disaggregated_excel(self, filename, qs, pfa_qs):
        wb = Workbook(write_only=True)

        ws = wb.create_sheet("1_Disaggregated_Data")
        ws.append(
            [
                "ID",
                "ISO3",
                "Country",
                "Geographical region",
                "Figure cause",
                "Year",
                "Figure category",
                "Figure unit",
                "Reported figures",
                "Household size",
                "Total figures",
                "Hazard category",
                "Hazard sub category",
                "Hazard type",
                "Hazard sub type",
                "Violence type",
                "Other event sub type",
                "Start date",
                "Start date accuracy",
                "End date",
                "End date accuracy",
                "Stock date",
                "Stock date accuracy",
                "Stock reporting date",
                "Publishers",
                "Sources",
                "Sources type",
                "Event ID",
                "Event name",
                "Event cause",
                "Event main trigger",
                "Event start date",
                "Event end date",
                "Event start date accuracy",
                "Event end date accuracy",
                "Is housing destruction",
                "Event codes (Code:Type)",
                "Locations coordinates",
                "Locations name",
                "Locations accuracy",
                "Locations type",
                "Pcode",
                "Pcode accuracy",
                "Pcode source",
                "Displacement occurred",
            ]
        )

        # Tab 2
        ws2 = wb.create_sheet("2_Context_Displacement_data")
        ws2.append(
            [
                "ISO3",
                "Year",
                "Figure cause",
                "Figure category",
                "Description",
                "Figures",
                "Figures rounded",
            ]
        )

        for item in pfa_qs.iterator(chunk_size=2000):
            ws2.append(
                [
                    item.iso3,
                    item.year,
                    get_enum_label(item.figure_cause),
                    get_enum_label(item.figure_category),
                    item.description,
                    item.figures,
                    item.figures_rounded,
                ]
            )

        # README TAB
        ws3 = wb.create_sheet("README")
        readme_text = [
            *readme_blocks.preamble_block(
                title="Disasters Global Internal Displacement Database (GIDD)",
                filename=filename,
                extracted_on=timezone.now().strftime("%B %d, %Y"),
                last_update=StatusLog.last_release_date(),
                description=readme_blocks.description_rows(),
                definitions=readme_blocks.definition_rows(),
                coverage=readme_blocks.COVERAGE_ALL_CAUSES,
            ),
            *readme_blocks.data_description_block("1_Disaggregated_Data table"),
            *readme_blocks.DISAGGREGATION_FIELD_ROWS,
            *readme_blocks.data_description_block("2_Context_Displacement_data table"),
        ]

        for item in readme_text:
            ws3.append(item)
        ws3.append(
            [
                "This dataset provides contextual information and analysis documented by IDMC analysts. It captures flags "
                "related to methodology, caveats, sources, and challenges identified for each metric, reporting year, and "
                "country."
            ]
        )
        ws3.append([])

        data_description_2 = [
            ["ISO3: Represents the ISO 3166-1 alpha-3 code. The code 'AB9' is assigned to the Abyei Area."],
            ["Year: Indicates the year for which displacement data are reported."],
            ["Figure cause: Identifies the trigger of displacement, such as conflict or disasters."],
            [
                "Figure category: Categorizes the type of displacement metric. It details values for Internal "
                "Displacements (internal displacement flows) and Total Number of IDPs (internal displacement stocks), "
                "as defined earlier in this document."
            ],
            [
                "Description: Provides contextual information about the data, including sources and data limitations. "
                "It is essential for representing the analysis conducted by IDMC analysts. This field also details the "
                "methodology used, descriptions of sources, and outlines any caveats and challenges identified with "
                "the displacement figures reported."
            ],
            [
                "Figures: Represents the total number of internal displacements or IDPs. For internal displacements, "
                "units are recorded as 'internal displacement flows' or 'internal displacement movements'. For the total "
                "number of IDPs, units reflect the total number of people living in displacement."
            ],
            [
                "Figures rounded: Displays rounded figures to provide a simplified view of the data that matches the "
                "figures reported in the Global Report on Internal Displacement (GRID)."
            ],
        ]
        for item in data_description_2:
            ws3.append(item)

        data_description_3 = [
            [],
            [],
            *historical_revisions_block(context_tab="2_Context_Displacement_data"),
        ]

        for item in data_description_3:
            ws3.append(item)

        qs = qs.filter(
            locations_coordinates__isnull=False,
        ).annotate(
            event_main_trigger=Case(
                When(gidd_event__cause=Crisis.CRISIS_TYPE.CONFLICT, then=F("gidd_event__violence__name")),
                When(gidd_event__cause=Crisis.CRISIS_TYPE.DISASTER, then=F("gidd_event__disaster_sub_type__name")),
                When(gidd_event__cause=Crisis.CRISIS_TYPE.OTHER, then=F("gidd_event__other_sub_type__name")),
                output_field=models.CharField(),
            ),
        )

        for item in qs.values(*DISAGGREGATION_EXPORT_VALUES).iterator(chunk_size=2000):
            ws.append(
                [
                    item["figure_raw_id"],
                    item["iso3"],
                    item["country_name"],
                    item["geographical_region_name"],
                    self._get_cause(item["cause"]),
                    item["year"],
                    self._get_category(item["category"]),
                    self._get_unit(item["unit"]),
                    item["reported"],
                    item["household_size"],
                    item["total_figures"],
                    item["disaster_category_name"],
                    item["disaster_sub_category_name"],
                    item["disaster_type_name"],
                    item["disaster_sub_type_name"],
                    item["violence_name"],
                    item["other_sub_type_name"],
                    item["start_date"],
                    self._get_date_accuracy(item["start_date_accuracy"]),
                    item["end_date"],
                    self._get_date_accuracy(item["end_date_accuracy"]),
                    item["stock_date"],
                    self._get_date_accuracy(item["stock_date_accuracy"]),
                    item["stock_reporting_date"],
                    string_join(EXTERNAL_ARRAY_SEPARATOR, item["publishers"]),
                    _join_keeping_gaps(item["sources"]),
                    _join_keeping_gaps(item["sources_type"]),
                    item["gidd_event__event_raw_id"],
                    item["gidd_event__name"],
                    self._get_cause(item["gidd_event__cause"]),
                    item["event_main_trigger"],
                    item["gidd_event__start_date"],
                    item["gidd_event__end_date"],
                    self._get_date_accuracy(item["gidd_event__start_date_accuracy"]),
                    self._get_date_accuracy(item["gidd_event__end_date_accuracy"]),
                    "Yes" if item["is_housing_destruction"] else "No",
                    self.extract_event_data(
                        item["gidd_event__event_codes"],
                        item["gidd_event__event_codes_type"],
                        item["gidd_event__event_codes_iso3"],
                        item["iso3"],
                    ),
                    _join_keeping_gaps(item["locations_coordinates"]),
                    _join_keeping_gaps(item["locations_names"]),
                    _get_location_accuracy_labels(item["locations_accuracy"]),
                    _get_location_type_labels(item["locations_type"]),
                    _join_keeping_gaps(item["locations_pcode"]),
                    _get_pcode_accuracy_labels(item["locations_pcode_accuracy"]),
                    _join_keeping_gaps(item["locations_pcode_source"]),
                    self._get_displacement_occurred(item["displacement_occurred"]),
                ]
            )

        return wb

    @extend_schema(
        description=Path("docs/disaggregation/geojson-export-description.md").read_text(),
        responses={
            # FIXME: Handle proper accept header
            # (302, "application/geo+json"): OpenApiTypes.STR,
            (302, "application/json"): OpenApiTypes.STR,
        },
        filters=True,
        tags=["GIDD"],
    )
    @action(
        detail=False,
        methods=["get"],
        url_path="disaggregated-geojson",
        permission_classes=[AllowAny],
        pagination_class=None,
    )
    def export_disaggregated_geojson(self, request):
        """
        Export the disaggregated data in geojson format file
        """
        track_gidd(
            self.request.GET.get("client_id"),
            ExternalApiDump.ExternalApiType.GIDD_DISAGGREGATION_EXPORT_GEOJSON,
            viewset=self,
        )
        queryset = GiddFigure.objects.select_related("gidd_event").order_by(
            "-year",
            "iso3",
            "id",
        )
        qs = self.filter_queryset(queryset)

        filename = self._generate_export_filename()
        return GiddExportCache.get_or_create(
            f"{filename}.geojson",
            request,
            [self.filterset_class],
            GiddExportCache.Key.DISAGGREGATION_EXPORT_GEOJSON,
            lambda: self._export_disaggregated_geojson(filename, qs),
            s3_parameters={
                "ResponseContentDisposition": f"attachment; filename={filename}.geojson",
                "ResponseContentType": "application/json",
            },
        )

    @extend_schema(
        description=Path("docs/disaggregation/xlsx-export-description.md").read_text(),
        responses={
            # FIXME: Handle proper accept header
            # (302, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"): OpenApiTypes.BINARY,
            (302, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"): None,
        },
        filters=True,
        tags=["GIDD"],
    )
    @action(
        detail=False,
        methods=["get"],
        url_path="disaggregated-export",
        permission_classes=[AllowAny],
        pagination_class=None,
        # FIXME: Handle proper accept header
        # renderer_classes=[XlsxRenderer],
    )
    def export_disaggregated(self, request):
        """
        Export the disaggregated data in excel format file
        """
        track_gidd(
            self.request.GET.get("client_id"),
            ExternalApiDump.ExternalApiType.GIDD_DISAGGREGATION_EXPORT_EXCEL,
            viewset=self,
        )
        queryset = GiddFigure.objects.select_related("gidd_event").order_by(
            "-year",
            "iso3",
            "id",
        )
        qs: models.QuerySet[GiddFigure] = self.filter_queryset(queryset)

        pfa_qs: models.QuerySet[PublicFigureAnalysis] = DisaggregationPublicFigureAnalysisFilterSet(
            data=self.request.query_params
        ).qs.order_by("iso3", "year", "id")

        filename = self._generate_export_filename()
        return GiddExportCache.get_or_create(
            f"{filename}.xlsx",
            request,
            [self.filterset_class, DisaggregationPublicFigureAnalysisFilterSet],
            GiddExportCache.Key.DISAGGREGATION_EXPORT,
            lambda: self._export_disaggregated_excel(filename, qs, pfa_qs),
            s3_parameters={
                "ResponseContentDisposition": f"attachment; filename={filename}.xlsx",
                "ResponseContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            },
        )


@extend_schema_view(
    list=extend_schema(
        description=Path("docs/public-figure-analyses/main-description.md").read_text(),
        responses=PublicFigureAnalysisSerializer(many=True),
        tags=["GIDD"],
    ),
)
class PublicFigureAnalysisViewSet(ListOnlyViewSetMixin):
    serializer_class = PublicFigureAnalysisSerializer
    filterset_class = PublicFigureAnalysisFilterSet
    pagination_class = GiddLimitOffsetPagination

    def get_queryset(self):
        track_gidd(
            self.request.GET.get("client_id"),
            ExternalApiDump.ExternalApiType.GIDD_PUBLIC_FIGURE_ANALYSIS_REST,
            viewset=self,
        )
        return PublicFigureAnalysis.objects.all()
