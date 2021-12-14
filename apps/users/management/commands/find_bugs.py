from django.core.management.base import BaseCommand
from django.db.models import F, Q
from openpyxl.workbook import Workbook
from apps.entry.models import Figure
from apps.event.models import Event
from apps.report.models import Report

# TODO:
# 1. Add stats for triangulation
# 2. Add stats for reports and stock figures
# 3. Add stats for GRID and MYU reports

largest_date = '2022-01-01'
smallest_date = '1995-01-01'


def get_fact_url(id):
    if not id:
        return ''
    return f'https://helix.idmcdb.org/facts/{id}'


def get_event_url(id):
    if not id:
        return ''
    return f'https://helix.idmcdb.org/events/{id}'


def get_new_event_url(id):
    if not id:
        return ''
    return f'https://helix-alpha.idmcdb.org/events/{id}'


def add_row(workspace, row, *args):
    for index, arg in enumerate(args):
        if isinstance(arg, str) and (arg.startswith('https://') or arg.startswith('http://')):
            workspace.cell(row=row, column=index + 1).hyperlink = arg
        else:
            workspace.cell(row=row, column=index + 1).value = arg


settings = {
    'ws0': {
        'title': 'Summary',
        'code': 'Summary',
    },
    'ws1': {
        'title': f'Events with small/large event dates ({smallest_date} to {largest_date})',
        'code': 'E1',
        'remarks': '',
    },
    'ws2': {
        'title': 'Recommended stock/flow figures without start date',
        'code': 'E2',
        'remarks': '',
    },
    'ws3': {
        'title': 'Recommended flow figures without end date',
        'code': 'E3',
        'remarks': '',
    },
    'ws4': {
        'title': 'Recommended stock figures without stock reporting date',
        'code': 'E4',
        'remarks': 'There are zero figures because we are automatically setting the stock reporting date',
    },
    'ws5': {
        'title': 'Recommended flow figures where start date greater than end date',
        'code': 'E5',
        'remarks': '',
    },
    'ws6': {
        'title': 'Recommended stock figures where start date greater than stock reporting date',
        'code': 'E6',
        'remarks': 'We are generating the stock reporting date using groups',
    },
    'ws7': {
        'title': f'Recommended figures with small/large start/end dates ({smallest_date} to {largest_date})',
        'code': 'E7',
        'remarks': '',
    },
    'ws8': {
        'title': 'Recommended new displacement figures not included in reports (but included in masterfacts)',
        'code': 'E8',
        'remarks': '',
    },
    'ws9': {
        'title': 'Recommended new displacement figures added in reports (but not included in masterfacts)',
        'code': 'E9',
        'remarks': '',
    },

    'ws10': {
        'title': 'Recommended idps stock figures not included in reports (but included in masterfacts)',
        'code': 'E10',
        'remarks': 'Not calculated',
    },
    'ws11': {
        'title': 'Recommended idps stock figures added in reports (but not included in masterfacts)',
        'code': 'E11',
        'remarks': 'Not calculated',
    },
}


class Command(BaseCommand):

    def handle(self, *args, **options):
        wb = Workbook()

        ws0 = wb.active

        # Events with small or large dates
        ws1 = wb.create_sheet(settings['ws1']['code'])
        ws1.append([settings['ws1']['title']])
        ws1.append(["Old ID", "Old URL", "ID", "URL"])

        # NOTE: start_date and end_date are required fields on database
        small_and_large_event_date_qs = Event.objects.filter(
            Q(start_date__gt=largest_date) |
            Q(start_date__lt=smallest_date) |
            Q(end_date__lt=smallest_date) |
            Q(end_date__gt=largest_date),
        ).distinct()

        events = small_and_large_event_date_qs.values('old_id', 'id')
        for row, event in enumerate(events):
            add_row(
                ws1,
                row + 3,
                event["old_id"],
                get_event_url(event["old_id"]),
                event["id"],
                get_new_event_url(event["id"]),
            )

        # Recommended stock and flow figures without start date
        ws2 = wb.create_sheet(settings['ws2']['code'])
        ws2.append([settings['ws2']['title']])
        ws2.append(["Fact ID", "Fact URL"])

        start_date_null_figures_qs = Figure.objects.filter(
            start_date__isnull=True,
            role=Figure.ROLE.RECOMMENDED
        )

        old_ids = start_date_null_figures_qs.values_list('old_id', flat=True)
        for row, id in enumerate(old_ids):
            add_row(
                ws2,
                row + 3,
                id,
                get_fact_url(id),
            )

        # Recommended flow figures without end date
        ws3 = wb.create_sheet(settings['ws3']['code'])
        ws3.append([settings['ws3']['title']])
        ws3.append(["Fact ID", "Fact URL"])

        flow_figures_without_end_date_qs = Figure.objects.filter(
            end_date__isnull=True,
            category__type='Flow',
            role=Figure.ROLE.RECOMMENDED
        )

        old_ids = flow_figures_without_end_date_qs.values_list('old_id', flat=True)
        for row, id in enumerate(old_ids):
            add_row(
                ws3,
                row + 3,
                id,
                get_fact_url(id),
            )

        # Recommended stock figures without end date
        ws4 = wb.create_sheet(settings['ws4']['code'])
        ws4.append([settings['ws4']['title']])
        ws4.append(["Fact ID", "Fact URL"])

        stock_figures_without_end_date_qs = Figure.objects.filter(
            end_date__isnull=True,
            category__type='Stock',
            role=Figure.ROLE.RECOMMENDED
        )

        old_ids = stock_figures_without_end_date_qs.values_list('old_id', flat=True)
        for row, id in enumerate(old_ids):
            add_row(
                ws4,
                row + 3,
                id,
                get_fact_url(id),
            )

        # Recommended flow figures where start date is greater than end date
        ws5 = wb.create_sheet(settings['ws5']['code'])
        ws5.append([settings['ws5']['title']])
        ws5.append(["Fact ID", "Fact URL"])

        flow_figures_with_start_date_gt_end_date_qs = Figure.objects.filter(
            start_date__isnull=False,
            end_date__isnull=False,
            start_date__gt=F('end_date'),
            category__type='Flow',
            role=Figure.ROLE.RECOMMENDED
        )

        old_ids = flow_figures_with_start_date_gt_end_date_qs.values_list('old_id', flat=True)
        for row, id in enumerate(old_ids):
            add_row(
                ws5,
                row + 3,
                id,
                get_fact_url(id),
            )

        # Recommended stock figures where start date is greater than end date
        ws6 = wb.create_sheet(settings['ws6']['code'])
        ws6.append([settings['ws6']['title']])
        ws6.append(["Fact ID", "Fact URL"])

        stock_figures_with_start_date_gt_end_date_qs = Figure.objects.filter(
            start_date__isnull=False,
            end_date__isnull=False,
            start_date__gt=F('end_date'),
            category__type='Stock',
            role=Figure.ROLE.RECOMMENDED
        )

        old_ids = stock_figures_with_start_date_gt_end_date_qs.values_list('old_id', flat=True)
        for row, id in enumerate(old_ids):
            add_row(
                ws6,
                row + 3,
                id,
                get_fact_url(id),
            )

        # Recommended stock and flow figures with small/large start/end dates
        ws7 = wb.create_sheet(settings['ws7']['code'])
        ws7.append([settings['ws7']['title']])
        ws7.append(["Fact ID", "Fact URL"])

        small_and_large_figure_date_qs = Figure.objects.filter(
            Q(start_date__gt=largest_date) |
            Q(start_date__lt=smallest_date) |
            Q(end_date__gt=largest_date) |
            Q(end_date__lt=smallest_date),
            role=Figure.ROLE.RECOMMENDED
        ).distinct()

        old_ids = small_and_large_figure_date_qs.values_list('old_id', flat=True)
        for row, id in enumerate(old_ids):
            add_row(
                ws7,
                row + 3,
                id,
                get_fact_url(id),
            )

        # Recommended flow figures not included in reports
        ws8 = wb.create_sheet(settings['ws8']['code'])
        ws8.append([settings['ws8']['title']])
        ws8.append(["Masterfact ID", "Masterface URL", "Subfact ID", "Subfact URL"])

        # Recommended flow figures added in reports
        ws9 = wb.create_sheet(settings['ws9']['code'])
        ws9.append([settings['ws9']['title']])
        ws9.append(["Masterfact ID", "Masterface URL", "Subfact ID", "Subfact URL"])

        missing_row = 0
        added_row = 0
        all_reports = Report.objects.all()
        for report in all_reports:
            if not report.old_id.isnumeric():
                continue

            linked_figures = report.attached_figures.filter(
                role=Figure.ROLE.RECOMMENDED,
                category__type='Flow',
                category__name='New Displacement',
            ).values_list('old_id', flat=True)
            extracted_figures = report.extract_report_figures.filter(
                role=Figure.ROLE.RECOMMENDED,
                category__type='Flow',
                category__name='New Displacement',
            ).values_list('old_id', flat=True)

            missing = set(linked_figures) - set(extracted_figures)
            for id in missing:
                add_row(
                    ws8,
                    missing_row + 3,
                    report.old_id,
                    get_fact_url(report.old_id),
                    id,
                    get_fact_url(id),
                )
                missing_row = missing_row + 1

            added = set(extracted_figures) - set(linked_figures)
            for id in added:
                add_row(
                    ws9,
                    added_row + 3,
                    report.old_id,
                    get_fact_url(report.old_id),
                    id,
                    get_fact_url(id),
                )
                added_row = added_row + 1

        # Summary page
        ws0.title = "summary"
        ws0.append(["Code", "Title", "Count", "Remarks"])
        ws0.append([settings['ws1']['code'], settings['ws1']['title'], small_and_large_event_date_qs.count(), settings['ws1']['remarks']])
        ws0.append([settings['ws2']['code'], settings['ws2']['title'], start_date_null_figures_qs.count(), settings['ws2']['remarks']])
        ws0.append([settings['ws3']['code'], settings['ws3']['title'], flow_figures_without_end_date_qs.count(), settings['ws3']['remarks']])
        ws0.append([settings['ws4']['code'], settings['ws4']['title'], stock_figures_without_end_date_qs.count(), settings['ws4']['remarks']])
        ws0.append([settings['ws5']['code'], settings['ws5']['title'], flow_figures_with_start_date_gt_end_date_qs.count(), settings['ws5']['remarks']])
        ws0.append([settings['ws6']['code'], settings['ws6']['title'], stock_figures_with_start_date_gt_end_date_qs.count(), settings['ws6']['remarks']])
        ws0.append([settings['ws7']['code'], settings['ws7']['title'], small_and_large_figure_date_qs.count(), settings['ws7']['remarks']])
        ws0.append([settings['ws8']['code'], settings['ws8']['title'], missing_row, settings['ws8']['remarks']])
        ws0.append([settings['ws9']['code'], settings['ws9']['title'], added_row, settings['ws9']['remarks']])
        ws0.append([settings['ws10']['code'], settings['ws10']['title'], '', settings['ws10']['remarks']])
        ws0.append([settings['ws11']['code'], settings['ws11']['title'], '', settings['ws11']['remarks']])

        wb.save(filename="data-errors.xlsx")
