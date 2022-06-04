import os
from django.core.management.base import BaseCommand
from django.db.models import F, Q, Count
from openpyxl.workbook import Workbook
from apps.entry.models import Figure
from apps.event.models import Event
from apps.report.models import Report
from apps.helixmigration.models import Documents, Facts, Events as OldEvents

largest_date = '2023-01-01'
smallest_date = '1995-01-01'


def get_fact_url(id):
    if not id:
        return ''
    return f'https://helix.idmcdb.org/facts/{id}'


def get_event_url(id):
    if not id:
        return ''
    return f'https://helix.idmcdb.org/events/{id}'


def get_new_event_url(id):
    if not id:
        return ''
    return f'https://helix-alpha.idmcdb.org/events/{id}'


def get_report_url(id):
    if not id:
        return ''
    return f'https://helix-alpha.idmcdb.org/reports/{id}/'


def add_row(workspace, row, *args):
    for index, arg in enumerate(args):
        if isinstance(arg, str) and (arg.startswith('https://') or arg.startswith('http://')):
            workspace.cell(row=row, column=index + 1).hyperlink = arg
        else:
            workspace.cell(row=row, column=index + 1).value = arg


settings = {
    'ws0': {
        'title': 'Summary',
        'code': 'Summary',
    },
    'ws1': {
        'title': f'Events with small/large event dates ({smallest_date} to {largest_date})',
        'code': 'E1',
        'remarks': '',
    },
    'ws2': {
        'title': 'Recommended stock/flow figures without start date',
        'code': 'E2',
        'remarks': '',
    },
    'ws3': {
        'title': 'Recommended flow figures without end date',
        'code': 'E3',
        'remarks': '',
    },
    'ws4': {
        'title': 'Recommended stock figures without stock reporting date',
        'code': 'E4',
        'remarks': 'There are zero figures because we are automatically setting the stock reporting date',
    },
    'ws5': {
        'title': 'Recommended flow figures where start date greater than end date',
        'code': 'E5',
        'remarks': '',
    },
    'ws6': {
        'title': 'Recommended stock figures where start date greater than stock reporting date',
        'code': 'E6',
        'remarks': 'We are generating the stock reporting date using groups from facts/masterfacts',
    },
    'ws7': {
        'title': f'Recommended figures with small/large start/end dates ({smallest_date} to {largest_date})',
        'code': 'E7',
        'remarks': '',
    },
    'ws8': {
        'title': 'Recommended new displacement figures not included in reports (but included in masterfacts)',
        'code': 'E8',
        'remarks': 'Figures are not included in reports when figure type does not match',
    },
    'ws9': {
        'title': 'Recommended new displacement figures added in reports (but not included in masterfacts)',
        'code': 'E9',
        'remarks': '',
    },

    'ws10': {
        'title': 'Recommended idps stock figures not included in reports (but included in masterfacts)',
        'code': 'E10',
        'remarks': 'Figures are not included in reports when figure type does not match',
    },
    'ws11': {
        'title': 'Recommended idps stock figures added in reports (but not included in masterfacts)',
        'code': 'E11',
        'remarks': '',
    },
    'ws12': {
        'title': 'Triangulation stock/flow figures without start date',
        'code': 'W1',
        'remarks': '',
    },
    'ws13': {
        'title': 'Triangulation flow figures without end date',
        'code': 'W2',
        'remarks': '',
    },
    'ws14': {
        'title': 'Triangulation stock figures without stock reporting date',
        'code': 'W3',
        'remarks': 'We are generating the stock reporting date using groups from facts/masterfacts',
    },
    'ws15': {
        'title': 'Triangulation flow figures where start date greater than end date',
        'code': 'W4',
        'remarks': '',
    },
    'ws16': {
        'title': 'Triangulation stock figures where start date greater than stock reporting date',
        'code': 'W5',
        'remarks': 'We are generating the stock reporting date using groups from facts/masterfacts',
    },
    'ws17': {
        'title': f'Triangulation figures with small/large start/end dates ({smallest_date} to {largest_date})',
        'code': 'W6',
        'remarks': '',
    },
    'ws18': {
        'title': 'Facts with titles that contain "DELETE", "NOT INCLUDE" or "IGNORE"',
        'code': 'W7',
        'remarks': '',
    },
    'ws19': {
        'title': 'Documents with titles that contain "DELETE", "NOT INCLUDE" or "IGNORE"',
        'code': 'W8',
        'remarks': '',
    },
    'ws20': {
        'title': 'Events with title that contain "DELETE", "NOT INCLUDE" or "IGNORE"',
        'code': 'W9',
        'remarks': '',
    },
    'ws21': {
        'title': 'Masterfacts with title that start with GRID/MYU but without "GRID" or "MYU" groups',
        'code': 'E12',
        'remarks': '',
    },
    'ws22': {
        'title': 'Subfacts without any documents / Masterfacts as subfact of other masterfact',
        'code': 'E13',
        'remarks': 'The document is not defined, but the parent is defined in the database',
    },
}


class Command(BaseCommand):

    def handle(self, *args, **options):

        wb = Workbook()

        ws0 = wb.active
        # Events with small or large dates
        ws1 = wb.create_sheet(settings['ws1']['code'])
        ws1.append([settings['ws1']['title']])
        ws1.append(["Old ID", "Old URL", "ID", "URL"])
        # NOTE: start_date and end_date are required fields on database
        small_and_large_event_date_qs = Event.objects.filter(
            Q(start_date__gt=largest_date) |
            Q(start_date__lt=smallest_date) |
            Q(end_date__lt=smallest_date) |
            Q(end_date__gt=largest_date),
            Q(old_id__isnull=False)
        ).distinct()

        events = small_and_large_event_date_qs.values('old_id', 'id')
        for row, event in enumerate(events):
            add_row(
                ws1,
                row + 3,
                event["old_id"],
                get_event_url(event["old_id"]),
                event["id"],
                get_new_event_url(event["id"]),
            )

        # Recommended stock and flow figures without start date
        ws2 = wb.create_sheet(settings['ws2']['code'])
        ws2.append([settings['ws2']['title']])
        ws2.append(["Fact ID", "Fact URL"])

        start_date_null_figures_qs = Figure.objects.filter(
            start_date__isnull=True,
            role=Figure.ROLE.RECOMMENDED
        )

        old_ids = start_date_null_figures_qs.values_list('old_id', flat=True)
        for row, id in enumerate(old_ids):
            add_row(
                ws2,
                row + 3,
                id,
                get_fact_url(id),
            )
        # Recommended flow figures without end date
        ws3 = wb.create_sheet(settings['ws3']['code'])
        ws3.append([settings['ws3']['title']])
        ws3.append(["Fact ID", "Fact URL"])

        flow_figures_without_end_date_qs = Figure.objects.filter(
            end_date__isnull=True,
            category__in=Figure.flow_list(),
            role=Figure.ROLE.RECOMMENDED
        )

        old_ids = flow_figures_without_end_date_qs.values_list('old_id', flat=True)
        for row, id in enumerate(old_ids):
            add_row(
                ws3,
                row + 3,
                id,
                get_fact_url(id),
            )

        # Recommended stock figures without end date
        ws4 = wb.create_sheet(settings['ws4']['code'])
        ws4.append([settings['ws4']['title']])
        ws4.append(["Fact ID", "Fact URL"])

        stock_figures_without_end_date_qs = Figure.objects.filter(
            end_date__isnull=True,
            category__in=Figure.stock_list(),
            role=Figure.ROLE.RECOMMENDED
        )

        old_ids = stock_figures_without_end_date_qs.values_list('old_id', flat=True)
        for row, id in enumerate(old_ids):
            add_row(
                ws4,
                row + 3,
                id,
                get_fact_url(id),
            )

        # Recommended flow figures where start date is greater than end date
        ws5 = wb.create_sheet(settings['ws5']['code'])
        ws5.append([settings['ws5']['title']])
        ws5.append(["Fact ID", "Fact URL"])
        flow_figures_with_start_date_gt_end_date_qs = Figure.objects.filter(
            start_date__isnull=False,
            end_date__isnull=False,
            start_date__gt=F('end_date'),
            category__in=Figure.flow_list(),
            role=Figure.ROLE.RECOMMENDED
        )

        old_ids = flow_figures_with_start_date_gt_end_date_qs.values_list('old_id', flat=True)
        for row, id in enumerate(old_ids):
            add_row(
                ws5,
                row + 3,
                id,
                get_fact_url(id),
            )

        # Recommended stock figures where start date is greater than end date
        ws6 = wb.create_sheet(settings['ws6']['code'])
        ws6.append([settings['ws6']['title']])
        ws6.append(["Fact ID", "Fact URL"])

        stock_figures_with_start_date_gt_end_date_qs = Figure.objects.filter(
            start_date__isnull=False,
            end_date__isnull=False,
            start_date__gt=F('end_date'),
            category__in=Figure.stock_list(),
            role=Figure.ROLE.RECOMMENDED
        )

        old_ids = stock_figures_with_start_date_gt_end_date_qs.values_list('old_id', flat=True)
        for row, id in enumerate(old_ids):
            add_row(
                ws6,
                row + 3,
                id,
                get_fact_url(id),
            )

        # Recommended stock and flow figures with small/large start/end dates
        ws7 = wb.create_sheet(settings['ws7']['code'])
        ws7.append([settings['ws7']['title']])
        ws7.append(["Fact ID", "Fact URL"])
        small_and_large_figure_date_qs = Figure.objects.filter(
            Q(start_date__gt=largest_date) |
            Q(start_date__lt=smallest_date) |
            Q(end_date__gt=largest_date) |
            Q(end_date__lt=smallest_date),
            role=Figure.ROLE.RECOMMENDED
        ).distinct()

        old_ids = small_and_large_figure_date_qs.values_list('old_id', flat=True)
        for row, id in enumerate(old_ids):
            add_row(
                ws7,
                row + 3,
                id,
                get_fact_url(id),
            )

        all_reports = Report.objects.all()
        # Recommended flow figures not included in reports
        ws8 = wb.create_sheet(settings['ws8']['code'])
        ws8.append([settings['ws8']['title']])
        ws8.append(["Masterfact ID", "Masterface URL", "Subfact ID", "Subfact URL", "Report ID", "Report URL"])

        # Recommended flow figures added in reports
        ws9 = wb.create_sheet(settings['ws9']['code'])
        ws9.append([settings['ws9']['title']])
        ws9.append(["Masterfact ID", "Masterface URL", "Subfact ID", "Subfact URL", "Report ID", "Report URL"])
        missing_flow_row = 0
        added_flow_row = 0
        for report in all_reports:
            if not report.old_id.isnumeric():
                continue

            linked_figures = report.attached_figures.filter(
                role=Figure.ROLE.RECOMMENDED,
                category__in=Figure.flow_list(),
                category=Figure.FIGURE_CATEGORY_TYPES.NEW_DISPLACEMENT.value,
            ).values_list('old_id', flat=True)
            extracted_figures = report.extract_report_figures.filter(
                role=Figure.ROLE.RECOMMENDED,
                category__in=Figure.flow_list(),
                category=Figure.FIGURE_CATEGORY_TYPES.NEW_DISPLACEMENT.value,
            ).values_list('old_id', flat=True)

            missing = set(linked_figures) - set(extracted_figures)
            for id in missing:
                add_row(
                    ws8,
                    missing_flow_row + 3,
                    report.old_id,
                    get_fact_url(report.old_id),
                    id,
                    get_fact_url(id),
                    report.id,
                    get_report_url(report.id),
                )
                missing_flow_row = missing_flow_row + 1

            added = set(extracted_figures) - set(linked_figures)
            for id in added:
                add_row(
                    ws9,
                    added_flow_row + 3,
                    report.old_id,
                    get_fact_url(report.old_id),
                    id,
                    get_fact_url(id),
                    report.id,
                    get_report_url(report.id),
                )
                added_flow_row = added_flow_row + 1

        # Recommended flow figures not included in reports
        ws10 = wb.create_sheet(settings['ws10']['code'])
        ws10.append([settings['ws10']['title']])
        ws10.append(["Masterfact ID", "Masterface URL", "Subfact ID", "Subfact URL", "Report ID", "Report URL"])
        # Recommended flow figures added in reports
        ws11 = wb.create_sheet(settings['ws11']['code'])
        ws11.append([settings['ws11']['title']])
        ws11.append(["Masterfact ID", "Masterface URL", "Subfact ID", "Subfact URL", "Report ID", "Report URL"])

        missing_stock_row = 0
        added_stock_row = 0
        for report in all_reports:
            if not report.old_id.isnumeric():
                continue

            linked_figures = report.attached_figures.filter(
                role=Figure.ROLE.RECOMMENDED,
                category__in=Figure.stock_list(),
                category=Figure.FIGURE_CATEGORY_TYPES.IDPS.value,
            ).values_list('old_id', flat=True)
            extracted_figures = report.extract_report_figures.filter(
                role=Figure.ROLE.RECOMMENDED,
                category__in=Figure.stock_list(),
                category=Figure.FIGURE_CATEGORY_TYPES.IDPS.value,
            ).values_list('old_id', flat=True)

            missing = set(linked_figures) - set(extracted_figures)
            for id in missing:
                add_row(
                    ws10,
                    missing_stock_row + 3,
                    report.old_id,
                    get_fact_url(report.old_id),
                    id,
                    get_fact_url(id),
                    report.id,
                    get_report_url(report.id),
                )
                missing_stock_row = missing_stock_row + 1

            added = set(extracted_figures) - set(linked_figures)
            for id in added:
                add_row(
                    ws11,
                    added_stock_row + 3,
                    report.old_id,
                    get_fact_url(report.old_id),
                    id,
                    get_fact_url(id),
                    report.id,
                    get_report_url(report.id),
                )
                added_stock_row = added_stock_row + 1

        # Triangulation stock and flow figures without start date
        ws12 = wb.create_sheet(settings['ws12']['code'])
        ws12.append([settings['ws12']['title']])
        ws12.append(["Fact ID", "Fact URL"])
        triangulation_start_date_null_figures_qs = Figure.objects.filter(
            start_date__isnull=True,
            role=Figure.ROLE.TRIANGULATION
        )

        old_ids = triangulation_start_date_null_figures_qs.values_list('old_id', flat=True)
        for row, id in enumerate(old_ids):
            add_row(
                ws12,
                row + 3,
                id,
                get_fact_url(id),
            )

        # Triangulation flow figures without end date
        ws13 = wb.create_sheet(settings['ws13']['code'])
        ws13.append([settings['ws13']['title']])
        ws13.append(["Fact ID", "Fact URL"])

        triangulation_flow_figures_without_end_date_qs = Figure.objects.filter(
            end_date__isnull=True,
            category__in=Figure.flow_list(),
            role=Figure.ROLE.TRIANGULATION
        )

        old_ids = triangulation_flow_figures_without_end_date_qs.values_list('old_id', flat=True)
        for row, id in enumerate(old_ids):
            add_row(
                ws13,
                row + 3,
                id,
                get_fact_url(id),
            )

        # Triangulation stock figures without end date
        ws14 = wb.create_sheet(settings['ws14']['code'])
        ws14.append([settings['ws14']['title']])
        ws14.append(["Fact ID", "Fact URL"])
        triangulation_stock_figures_without_end_date_qs = Figure.objects.filter(
            end_date__isnull=True,
            category__in=Figure.stock_list(),
            role=Figure.ROLE.TRIANGULATION
        )

        old_ids = triangulation_stock_figures_without_end_date_qs.values_list('old_id', flat=True)
        for row, id in enumerate(old_ids):
            add_row(
                ws14,
                row + 3,
                id,
                get_fact_url(id),
            )

        # Triangulation flow figures where start date is greater than end date
        ws15 = wb.create_sheet(settings['ws15']['code'])
        ws15.append([settings['ws15']['title']])
        ws15.append(["Fact ID", "Fact URL"])

        triangulation_flow_figures_with_start_date_gt_end_date_qs = Figure.objects.filter(
            start_date__isnull=False,
            end_date__isnull=False,
            start_date__gt=F('end_date'),
            category__in=Figure.flow_list(),
            role=Figure.ROLE.TRIANGULATION
        )

        old_ids = triangulation_flow_figures_with_start_date_gt_end_date_qs.values_list('old_id', flat=True)
        for row, id in enumerate(old_ids):
            add_row(
                ws15,
                row + 3,
                id,
                get_fact_url(id),
            )

        # Triangulation stock figures where start date is greater than end date
        ws16 = wb.create_sheet(settings['ws16']['code'])
        ws16.append([settings['ws16']['title']])
        ws16.append(["Fact ID", "Fact URL"])

        triangulation_stock_figures_with_start_date_gt_end_date_qs = Figure.objects.filter(
            start_date__isnull=False,
            end_date__isnull=False,
            start_date__gt=F('end_date'),
            category__in=Figure.stock_list(),
            role=Figure.ROLE.TRIANGULATION
        )

        old_ids = triangulation_stock_figures_with_start_date_gt_end_date_qs.values_list('old_id', flat=True)
        for row, id in enumerate(old_ids):
            add_row(
                ws16,
                row + 3,
                id,
                get_fact_url(id),
            )

        # Triangulation stock and flow figures with small/large start/end dates
        ws17 = wb.create_sheet(settings['ws17']['code'])
        ws17.append([settings['ws17']['title']])
        ws17.append(["Fact ID", "Fact URL"])
        triangulation_small_and_large_figure_date_qs = Figure.objects.filter(
            Q(start_date__gt=largest_date) |
            Q(start_date__lt=smallest_date) |
            Q(end_date__gt=largest_date) |
            Q(end_date__lt=smallest_date),
            role=Figure.ROLE.TRIANGULATION
        ).distinct()

        old_ids = triangulation_small_and_large_figure_date_qs.values_list('old_id', flat=True)
        for row, id in enumerate(old_ids):
            add_row(
                ws17,
                row + 3,
                id,
                get_fact_url(id),
            )

        all_reports = Report.objects.all()

        # Old facts Delete and do not include items
        ws18 = wb.create_sheet(settings['ws18']['code'])
        ws18.append([settings['ws18']['title']])
        ws18.append(["Fact ID", "Fact URL", "Fact Name"])
        old_facts_qs = Facts.objects.using('helixmigration').filter(
            Q(name__icontains="ignore") |
            Q(name__icontains="not include") |
            Q(name__icontains="delete")
        )
        for row, fact_obj in enumerate(old_facts_qs):
            add_row(
                ws18,
                row + 3,
                fact_obj.id,
                get_fact_url(fact_obj.id),
                fact_obj.name
            )

        # Old documents Delete and do not include items
        ws19 = wb.create_sheet(settings['ws19']['code'])
        ws19.append([settings['ws19']['title']])
        ws19.append(["Fact ID", "Fact URL", "Fact Name"])

        old_documents_qs = Documents.objects.using('helixmigration').filter(
            Q(name__icontains="ignore") |
            Q(name__icontains="not include") |
            Q(name__icontains="delete")
        )
        for row, document_obj in enumerate(old_documents_qs):
            add_row(
                ws19,
                row + 3,
                document_obj.id,
                get_fact_url(document_obj.id),
                document_obj.name
            )

        # Old events Delete and do not include items
        ws20 = wb.create_sheet(settings['ws20']['code'])
        ws20.append([settings['ws20']['title']])
        ws20.append(["Event ID", "Event URL", "Event Name"])

        old_events_qs = OldEvents.objects.using('helixmigration').filter(
            Q(name__icontains="ignore") |
            Q(name__icontains="not include") |
            Q(name__icontains="delete")
        )

        for row, event_obj in enumerate(old_events_qs):
            add_row(
                ws20,
                row + 3,
                event_obj.id,
                get_event_url(event_obj.id),
                event_obj.name,
            )

        # Old Named GRID and MYU but not tagged as one
        # (https://github.com/idmc-labs/Helix2.0/issues/243#issuecomment-974207507)
        ws21 = wb.create_sheet(settings['ws21']['code'])
        ws21.append([settings['ws21']['title']])
        ws21.append(["Fact ID", "Fact URL", "Fact Name"])

        masterfact_stats_qs = Facts.objects.using(
            'helixmigration'
        ).order_by().values('parent_id').annotate(count=Count('*')).filter(
            count__gt=0,
        ).values_list('parent_id')

        old_grid_and_myu_qs = Facts.objects.using('helixmigration').filter(
            document_id__isnull=True,
            parent_id__isnull=True,
            displacement_type='Disaster',
            data__isnull=False,
            groups__len=0,
            id__in=masterfact_stats_qs,
        ).filter(
            Q(name__contains='MYU') |
            Q(name__contains='GRID')
        ).annotate(subfact=F('parent_id'))

        for row, fact_obj in enumerate(old_grid_and_myu_qs):
            add_row(
                ws21,
                row + 3,
                fact_obj.id,
                get_fact_url(fact_obj.id),
                fact_obj.name
            )

        # Old Facts that are both masterfact and subfact at the same time
        ws22 = wb.create_sheet(settings['ws22']['code'])
        ws22.append([settings['ws22']['title']])
        ws22.append(["Fact ID", "Fact URL"])

        master_facts_qs = Facts.objects.using('helixmigration').filter(
            parent_id__isnull=False,
            document_id__isnull=True
        )
        ids = master_facts_qs.values_list('id', flat=True)
        for row, id in enumerate(ids):
            add_row(
                ws22,
                row + 3,
                id,
                get_fact_url(id),
            )

        # Summary page
        ws0.title = "summary"
        ws0.append(["Code", "Title", "Count", "Remarks"])
        ws0.append(
            [
                settings['ws1']['code'], settings['ws1']['title'],
                small_and_large_event_date_qs.count(), settings['ws1']['remarks']
            ]
        )

        ws0.append(
            [
                settings['ws2']['code'], settings['ws2']['title'],
                start_date_null_figures_qs.count(), settings['ws2']['remarks']
            ]
        )
        ws0.append(
            [
                settings['ws3']['code'], settings['ws3']['title'],
                flow_figures_without_end_date_qs.count(), settings['ws3']['remarks']
            ]
        )
        ws0.append(
            [
                settings['ws4']['code'], settings['ws4']['title'],
                stock_figures_without_end_date_qs.count(), settings['ws4']['remarks']
            ]
        )
        ws0.append(
            [
                settings['ws5']['code'], settings['ws5']['title'],
                flow_figures_with_start_date_gt_end_date_qs.count(), settings['ws5']['remarks']
            ]
        )
        ws0.append(
            [
                settings['ws6']['code'], settings['ws6']['title'],
                stock_figures_with_start_date_gt_end_date_qs.count(), settings['ws6']['remarks']
            ]
        )
        ws0.append(
            [
                settings['ws7']['code'], settings['ws7']['title'],
                small_and_large_figure_date_qs.count(), settings['ws7']['remarks']
            ]
        )
        ws0.append(
            [settings['ws8']['code'], settings['ws8']['title'], missing_flow_row, settings['ws8']['remarks']]
        )
        ws0.append([settings['ws9']['code'], settings['ws9']['title'], added_flow_row, settings['ws9']['remarks']])

        ws0.append(
            [settings['ws10']['code'], settings['ws10']['title'], missing_stock_row, settings['ws10']['remarks']]
        )
        ws0.append(
            [settings['ws11']['code'], settings['ws11']['title'], added_stock_row, settings['ws11']['remarks']]
        )
        ws0.append(
            [
                settings['ws12']['code'], settings['ws12']['title'],
                triangulation_start_date_null_figures_qs.count(), settings['ws12']['remarks']
            ]
        )
        ws0.append(
            [
                settings['ws13']['code'], settings['ws13']['title'],
                triangulation_flow_figures_without_end_date_qs.count(), settings['ws13']['remarks']
            ]
        )
        ws0.append(
            [
                settings['ws14']['code'], settings['ws14']['title'],
                triangulation_stock_figures_without_end_date_qs.count(), settings['ws14']['remarks']
            ]
        )
        ws0.append(
            [
                settings['ws15']['code'], settings['ws15']['title'],
                triangulation_flow_figures_with_start_date_gt_end_date_qs.count(), settings['ws15']['remarks']
            ]
        )
        ws0.append(
            [
                settings['ws16']['code'], settings['ws16']['title'],
                triangulation_stock_figures_with_start_date_gt_end_date_qs.count(), settings['ws16']['remarks']
            ]
        )
        ws0.append(
            [
                settings['ws17']['code'], settings['ws17']['title'],
                triangulation_small_and_large_figure_date_qs.count(), settings['ws17']['remarks']
            ]
        )
        ws0.append(
            [
                settings['ws18']['code'], settings['ws18']['title'],
                old_facts_qs.count(), settings['ws18']['remarks']
            ]
        )
        ws0.append(
            [
                settings['ws19']['code'], settings['ws19']['title'],
                old_documents_qs.count(), settings['ws19']['remarks']
            ]
        )
        ws0.append(
            [
                settings['ws20']['code'], settings['ws20']['title'],
                old_events_qs.count(), settings['ws20']['remarks']
            ]
        )
        ws0.append(
            [
                settings['ws21']['code'], settings['ws21']['title'],
                old_grid_and_myu_qs.count(), settings['ws21']['remarks']
            ]
        )
        ws0.append(
            [
                settings['ws22']['code'], settings['ws22']['title'],
                master_facts_qs.count(), settings['ws22']['remarks']
            ]
        )
        # Make sure generated directory exists
        try:
            os.makedirs("generated")
        except FileExistsError:
            pass
        wb.save(filename="generated/data-errors.xlsx")
