import json

from apps.crisis.models import Crisis
from apps.entry.models import (
    Figure,
    FigureLocation,
)
from apps.users.enums import USER_ROLE
from utils.factories import (
    ContextOfViolenceFactory,
    CountryFactory,
    EntryFactory,
    EventFactory,
    FigureFactory,
    FigureLocationFactory,
    OrganizationFactory,
    TagFactory,
)
from utils.permissions import PERMISSION_DENIED_MESSAGE
from utils.tests import HelixGraphQLTestCase, create_user_with_role


class TestEntryCreation(HelixGraphQLTestCase):
    def setUp(self) -> None:
        self.country = CountryFactory.create(iso2="lo", iso3="lol")
        self.country_id = str(self.country.id)
        self.event = EventFactory.create(event_type=Crisis.CRISIS_TYPE.CONFLICT.value)
        self.event.countries.add(self.country)
        self.fig_cat = Figure.FIGURE_CATEGORY_TYPES.NEW_DISPLACEMENT
        self.editor = create_user_with_role(USER_ROLE.MONITORING_EXPERT.name)
        self.mutation = """
            mutation CreateEntry($input: EntryCreateInputType!) {
                createEntry(data: $input) {
                    ok
                    errors
                    result {
                        id
                        figures {
                            id
                            createdBy{
                                id
                                fullName
                            }
                        }
                        createdBy{
                            id
                            fullName
                        }
                    }
                }
            }
        """
        self.input = {
            "url": "https://yoko-onos-blog.com",
            "articleTitle": "title 1",
            "publishers": [str(OrganizationFactory.create().id)],
            "publishDate": "2020-09-09",
            "idmcAnalysis": "analysis one",
            "isConfidential": True,
        }
        self.force_login(self.editor)
        self.tag1 = TagFactory.create()
        self.tag2 = TagFactory.create()
        self.tag3 = TagFactory.create()
        self.context_of_violence = ContextOfViolenceFactory.create()

    def test_valid_create_entry(self):
        response = self.query(self.mutation, input_data=self.input)
        content = json.loads(response.content)

        self.assertResponseNoErrors(response)
        self.assertTrue(content["data"]["createEntry"]["ok"], content)
        self.assertIsNone(content["data"]["createEntry"]["errors"], content)
        self.assertIsNotNone(content["data"]["createEntry"]["result"]["id"])

    def test_invalid_guest_entry_create(self):
        guest = create_user_with_role(role=USER_ROLE.GUEST.name)
        self.force_login(guest)
        response = self.query(self.mutation, input_data=self.input)
        content = json.loads(response.content)
        self.assertIn(PERMISSION_DENIED_MESSAGE, content["errors"][0]["message"])


class TestEntryUpdate(HelixGraphQLTestCase):
    def setUp(self) -> None:
        self.country = CountryFactory.create(iso2="np")
        self.country_id = str(self.country.id)
        self.fig_cat = Figure.FIGURE_CATEGORY_TYPES.NEW_DISPLACEMENT
        self.editor = create_user_with_role(USER_ROLE.MONITORING_EXPERT)
        self.admin = create_user_with_role(USER_ROLE.ADMIN.name)
        self.event = EventFactory.create(name="myevent", event_type=Crisis.CRISIS_TYPE.CONFLICT.value)
        self.event.countries.add(self.country)
        self.entry = EntryFactory.create(
            created_by=self.editor,
        )
        self.mutation = """
        mutation MyMutation($input: EntryUpdateInputType!) {
          updateEntry(data: $input) {
            ok
            errors
            result {
              id
              createdAt
              articleTitle
              createdBy {
                  id
                  fullName
              }
            }
          }
        }
        """
        self.input = {
            "id": self.entry.id,
            "articleTitle": "updated-bla",
        }

    def test_valid_update_entry(self):
        self.force_login(self.admin)
        response = self.query(self.mutation, input_data=self.input)
        content = json.loads(response.content)

        self.assertResponseNoErrors(response)
        self.assertTrue(content["data"]["updateEntry"]["ok"], content)


class TestEntryDelete(HelixGraphQLTestCase):
    def setUp(self) -> None:
        self.editor = create_user_with_role(USER_ROLE.MONITORING_EXPERT.name)
        self.entry = EntryFactory.create(created_by=self.editor)
        self.mutation = """
            mutation DeleteEntry($id: ID!) {
                deleteEntry(id: $id) {
                    ok
                    errors
                    result {
                        id
                        url
                        createdAt
                    }
                }
            }
        """
        self.variables = {
            "id": self.entry.id,
        }

    def test_valid_delete_entry(self):
        self.force_login(self.editor)
        response = self.query(self.mutation, variables=self.variables)
        content = json.loads(response.content)

        self.assertResponseNoErrors(response)
        self.assertTrue(content["data"]["deleteEntry"]["ok"], content)
        self.assertEqual(content["data"]["deleteEntry"]["result"]["url"], self.entry.url)

    def test_valid_entry_delete_by_admins(self):
        admin = create_user_with_role(USER_ROLE.ADMIN.name)
        self.force_login(admin)
        response = self.query(self.mutation, variables=self.variables)
        content = json.loads(response.content)

        self.assertResponseNoErrors(response)
        self.assertTrue(content["data"]["deleteEntry"]["ok"], content)
        self.assertEqual(content["data"]["deleteEntry"]["result"]["url"], self.entry.url)


class TestExportEntry(HelixGraphQLTestCase):
    def setUp(self) -> None:
        self.editor = create_user_with_role(USER_ROLE.MONITORING_EXPERT.name)
        for _ in range(3):
            EntryFactory.create(created_by=self.editor)
        self.mutation = """
        mutation ExportEntries($filterFigureStartAfter: Date, $filterFigureEndBefore: Date){
            exportEntries(
                filters: {
                    filterFigureStartAfter: $filterFigureStartAfter
                    filterFigureEndBefore: $filterFigureEndBefore
                }
          ){
            errors
            ok
          }
        }

        """
        self.variables = {
            "filterFigureStartAfter": "2018-08-25",
            "filterFigureEndBefore": "2021-08-25",
        }

    def test_export_entry(self):
        self.force_login(self.editor)
        response = self.query(self.mutation, variables=self.variables)
        self.assertResponseNoErrors(response)


class TestExportFigures(HelixGraphQLTestCase):
    def setUp(self) -> None:
        self.editor = create_user_with_role(USER_ROLE.MONITORING_EXPERT.name)
        self.mutation = """
        mutation ExportFigures($metadata: ExportFiguresMetadataInputType){
            exportFigures(
                filters: {}
                metadata: $metadata
          ){
            errors
            ok
            result {
              id
              metadata
            }
          }
        }
        """

    def test_export_figures_without_metadata_is_unchanged(self):
        """ExportFigures callers without metadata should still work."""
        from apps.contrib.models import ExcelDownload

        self.force_login(self.editor)
        response = self.query(self.mutation, variables={})
        self.assertResponseNoErrors(response)
        content = json.loads(response.content)
        self.assertTrue(content["data"]["exportFigures"]["ok"], content)
        instance_id = content["data"]["exportFigures"]["result"]["id"]
        instance = ExcelDownload.objects.get(id=instance_id)
        self.assertIsNone(instance.metadata)

    def test_export_figures_persists_metadata(self):
        """ExportFigures mutation accepts the metadata argument and persists it."""
        from apps.contrib.models import ExcelDownload

        self.force_login(self.editor)
        response = self.query(
            self.mutation,
            variables={"metadata": {"explodeByLocations": True}},
        )
        self.assertResponseNoErrors(response)
        content = json.loads(response.content)
        self.assertTrue(content["data"]["exportFigures"]["ok"], content)
        instance_id = content["data"]["exportFigures"]["result"]["id"]
        instance = ExcelDownload.objects.get(id=instance_id)
        self.assertEqual(instance.metadata, {"explode_by_locations": True})

    def test_export_figures_rejects_unknown_metadata_key(self):
        """A typo in metadata key is rejected at the GraphQL boundary."""
        from apps.contrib.models import ExcelDownload

        self.force_login(self.editor)
        response = self.query(
            self.mutation,
            # NOTE: `explodeByLocation` is missing the trailing 's'
            variables={"metadata": {"explodeByLocation": True}},
        )
        content = json.loads(response.content)
        self.assertIn("errors", content, content)
        # Nothing should have been queued: the typo did not silently fall through.
        self.assertFalse(ExcelDownload.objects.exists())

    def test_export_figures_coerces_non_boolean_metadata_value(self):
        """Graphene's Boolean scalar *coerces* a non-bool value (e.g. the string 'yes') to a
        real Python bool rather than rejecting it. Unknown KEYS are rejected but value-type
        coercion follows GraphQL scalar rules. Check if export is successfully registered."""
        from apps.contrib.models import ExcelDownload

        self.force_login(self.editor)
        response = self.query(
            self.mutation,
            variables={"metadata": {"explodeByLocations": "yes"}},
        )
        self.assertResponseNoErrors(response)
        content = json.loads(response.content)
        self.assertTrue(content["data"]["exportFigures"]["ok"], content)
        instance = ExcelDownload.objects.get(id=content["data"]["exportFigures"]["result"]["id"])
        self.assertIs(instance.metadata["explode_by_locations"], True)


class TestFigureGetExcelSheetsData(HelixGraphQLTestCase):
    def _make_figure(self, *, editor, country, event, entry, total_figures, locations=()):
        # total_figures is normally computed in the figure serializer, which the factory
        # bypasses; the model has no save hook that derives it (default=0). So we set it
        # directly here to pin the value the export pipeline reads.
        return FigureFactory.create(
            entry=entry,
            event=event,
            country=country,
            created_by=editor,
            role=Figure.ROLE.RECOMMENDED,
            unit=Figure.UNIT.PERSON,
            reported=total_figures,
            total_figures=total_figures,
            category=Figure.FIGURE_CATEGORY_TYPES.IDPS,
            geo_locations=list(locations),
        )

    def setUp(self):
        super().setUp()
        self.editor = create_user_with_role(USER_ROLE.MONITORING_EXPERT.name)
        self.country = CountryFactory.create()
        self.event = EventFactory.create(created_by=self.editor)
        self.event.countries.set([self.country])
        self.entry = EntryFactory.create(created_by=self.editor)

    def _explode_rows(self):
        result = Figure.get_excel_sheets_data(
            user_id=self.editor.id,
            filters={},
            metadata={"explode_by_locations": True},
        )
        # FIXME: none of the usecase use the result so why not remove it instead
        return result, list(result["data"])

    def test_headers_shape_and_readme(self):
        """Check if Headers and Readme are correctly set up."""
        # Need at least one figure for the queryset path to exercise transformer-free codepaths.
        result = Figure.get_excel_sheets_data(
            user_id=self.editor.id,
            filters={},
            metadata={"explode_by_locations": True},
        )
        headers = result["headers"]
        for key in (
            "allocated_figure",
            "location_id",
            "location_display_name",
            "location_lat_lng",
            "location_accuracy",
            "location_identifier",
        ):
            self.assertIn(key, headers)
        for key in ("centroid", "centroid_lat", "centroid_lon", "locations"):
            self.assertNotIn(key, headers)
        keys = list(headers.keys())
        self.assertEqual(keys.index("allocated_figure") - keys.index("total_figures"), 1)

        # The explode rows are pre-transformed inside the generator, so the contract is
        # transformer=None; otherwise get_excel_sheet_content would double-transform them.
        self.assertIsNone(result["transformer"])

        readme_rows = result["readme_data"][0]["results"]["data"]
        readme_column_names = {row["column_name"] for row in readme_rows}
        self.assertIn("Allocated figure", readme_column_names)
        self.assertIn("About", readme_column_names)
        self.assertNotIn("Centroid", readme_column_names)

        # Drain generator to avoid keeping a half-consumed DB cursor open.
        list(result["data"])

    def test_r2_allocation_three_origins(self):
        """Test for T=100, N=3 -> [34, 33, 33] in id-ASC order."""
        locs = FigureLocationFactory.create_batch(3, identifier=FigureLocation.IDENTIFIER.ORIGIN)
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=100,
            locations=locs,
        )
        _, rows = self._explode_rows()
        self.assertEqual(len(rows), 3)
        self.assertEqual([r["allocated_figure"] for r in rows], [34, 33, 33])
        self.assertEqual([r["location_identifier"] for r in rows], ["Origin", "Origin", "Origin"])
        # Location ids in ASC order.
        sorted_loc_ids = sorted(loc.id for loc in locs)
        self.assertEqual([r["location_id"] for r in rows], sorted_loc_ids)

    def test_r2_allocation_single_origin(self):
        """Test for T=100, N=1 -> [100]."""
        loc = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.ORIGIN)
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=100,
            locations=[loc],
        )
        _, rows = self._explode_rows()
        self.assertEqual([r["allocated_figure"] for r in rows], [100])

    def test_r2_allocation_seven_over_four(self):
        """Test for T=7, N=4 -> [2, 2, 2, 1]."""
        locs = FigureLocationFactory.create_batch(4, identifier=FigureLocation.IDENTIFIER.ORIGIN)
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=7,
            locations=locs,
        )
        _, rows = self._explode_rows()
        self.assertEqual([r["allocated_figure"] for r in rows], [2, 2, 2, 1])

    def test_r2_allocation_one_over_three(self):
        """Test for T=1, N=3 -> [1, 0, 0]."""
        locs = FigureLocationFactory.create_batch(3, identifier=FigureLocation.IDENTIFIER.ORIGIN)
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=1,
            locations=locs,
        )
        _, rows = self._explode_rows()
        self.assertEqual([r["allocated_figure"] for r in rows], [1, 0, 0])

    def test_r2_allocation_ten_over_five(self):
        """Test for T=10, N=5 -> [2, 2, 2, 2, 2]."""
        locs = FigureLocationFactory.create_batch(5, identifier=FigureLocation.IDENTIFIER.ORIGIN)
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=10,
            locations=locs,
        )
        _, rows = self._explode_rows()
        self.assertEqual([r["allocated_figure"] for r in rows], [2, 2, 2, 2, 2])

    def test_o_and_d_single_location_yields_two_rows(self):
        """An O&D location alone produces one Origin row + one Destination row."""
        loc = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.ORIGIN_AND_DESTINATION)
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=100,
            locations=[loc],
        )
        _, rows = self._explode_rows()
        self.assertEqual(len(rows), 2)
        identifiers = [r["location_identifier"] for r in rows]
        self.assertEqual(identifiers, ["Origin", "Destination"])
        # Both rows get the full 100 because each bucket has N=1.
        self.assertEqual([r["allocated_figure"] for r in rows], [100, 100])
        # Same location_id appears on both rows.
        self.assertEqual(rows[0]["location_id"], rows[1]["location_id"])
        self.assertEqual(rows[0]["location_id"], loc.id)

    def test_o_and_d_mixed_with_origin_and_destination(self):
        """One Origin + one O&D + one Destination -> N_origin=2, N_dest=2."""
        loc_o = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.ORIGIN)
        loc_od = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.ORIGIN_AND_DESTINATION)
        loc_d = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.DESTINATION)
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=10,
            locations=[loc_o, loc_od, loc_d],
        )
        _, rows = self._explode_rows()
        # 2 origin rows + 2 destination rows.
        self.assertEqual(len(rows), 4)
        origin_rows = [r for r in rows if r["location_identifier"] == "Origin"]
        dest_rows = [r for r in rows if r["location_identifier"] == "Destination"]
        self.assertEqual(len(origin_rows), 2)
        self.assertEqual(len(dest_rows), 2)
        # Origin bucket is {loc_o, loc_od}; destination bucket is {loc_d, loc_od}.
        self.assertEqual(sorted(r["location_id"] for r in origin_rows), sorted([loc_o.id, loc_od.id]))
        self.assertEqual(sorted(r["location_id"] for r in dest_rows), sorted([loc_d.id, loc_od.id]))
        # T=10 split over N=2 each -> [5, 5] within each bucket.
        self.assertEqual(sum(r["allocated_figure"] for r in origin_rows), 10)
        self.assertEqual(sum(r["allocated_figure"] for r in dest_rows), 10)

    def test_exclude_total_figures_zero(self):
        """Figures with total_figures=0 are dropped."""
        loc = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.ORIGIN)
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=0,
            locations=[loc],
        )
        _, rows = self._explode_rows()
        self.assertEqual(rows, [])

    def test_exclude_figures_without_locations(self):
        """Figures with no geo_locations are dropped."""
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=50,
            locations=[],
        )
        _, rows = self._explode_rows()
        self.assertEqual(rows, [])

    def test_explode_respects_extraction_filters(self):
        """The explode export accepts the extraction filters."""
        other_country = CountryFactory.create()
        self.event.countries.add(other_country)

        loc_kept = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.ORIGIN)
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=10,
            locations=[loc_kept],
        )
        loc_filtered_out = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.ORIGIN)
        self._make_figure(
            editor=self.editor,
            country=other_country,
            event=self.event,
            entry=self.entry,
            total_figures=20,
            locations=[loc_filtered_out],
        )

        result = Figure.get_excel_sheets_data(
            user_id=self.editor.id,
            filters={"filter_figure_countries": [self.country.id]},
            metadata={"explode_by_locations": True},
        )
        rows = list(result["data"])
        # Only the figure in self.country survives the filter.
        self.assertEqual({r["location_id"] for r in rows}, {loc_kept.id})
        self.assertEqual([r["total_figures"] for r in rows], [10])

    def test_export_is_not_scoped_to_requesting_user(self):
        """The explode export is global: a figure created by one user still appears when a
        DIFFERENT user runs the export."""
        loc = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.ORIGIN)
        figure = self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=10,
            locations=[loc],
        )

        other_user = create_user_with_role(USER_ROLE.MONITORING_EXPERT.name)
        result = Figure.get_excel_sheets_data(
            user_id=other_user.id,
            filters={},
            metadata={"explode_by_locations": True},
        )
        rows = list(result["data"])
        self.assertEqual([r["id"] for r in rows], [figure.id])

    def test_row_order_within_figure(self):
        """Rows are created (identifier_label='Origin' first, then 'Destination'),
        each bucket sorted by location_id ASC, regardless of insertion order."""
        # Create locations in non-monotonic-id order by inserting destination first.
        loc_d_high = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.DESTINATION)
        loc_o_high = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.ORIGIN)
        loc_d_low = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.DESTINATION)
        loc_o_low = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.ORIGIN)

        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=4,
            locations=[loc_d_high, loc_o_high, loc_d_low, loc_o_low],
        )
        _, rows = self._explode_rows()
        # The lowest ids are loc_d_high (1st created) and loc_o_high (2nd created),
        # then loc_d_low (3rd), then loc_o_low (4th). So sorted Origin ids: [loc_o_high, loc_o_low].
        # Sorted Destination ids: [loc_d_high, loc_d_low].
        identifiers = [r["location_identifier"] for r in rows]
        # All origin first, then destination.
        self.assertEqual(identifiers, ["Origin", "Origin", "Destination", "Destination"])
        origin_ids = [r["location_id"] for r in rows if r["location_identifier"] == "Origin"]
        dest_ids = [r["location_id"] for r in rows if r["location_identifier"] == "Destination"]
        self.assertEqual(origin_ids, sorted(origin_ids))
        self.assertEqual(dest_ids, sorted(dest_ids))

    def test_total_figures_repeats_per_row(self):
        """All rows for the same figure carry the same total_figures."""
        locs = FigureLocationFactory.create_batch(3, identifier=FigureLocation.IDENTIFIER.ORIGIN)
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=42,
            locations=locs,
        )
        _, rows = self._explode_rows()
        self.assertEqual([r["total_figures"] for r in rows], [42, 42, 42])

    def test_per_identifier_sum_invariant(self):
        """For each identifier bucket, sum(allocated_figure) == total_figures."""
        origins = FigureLocationFactory.create_batch(3, identifier=FigureLocation.IDENTIFIER.ORIGIN)
        destinations = FigureLocationFactory.create_batch(2, identifier=FigureLocation.IDENTIFIER.DESTINATION)
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=100,
            locations=origins + destinations,
        )
        _, rows = self._explode_rows()
        origin_sum = sum(r["allocated_figure"] for r in rows if r["location_identifier"] == "Origin")
        dest_sum = sum(r["allocated_figure"] for r in rows if r["location_identifier"] == "Destination")
        self.assertEqual(origin_sum, 100)
        self.assertEqual(dest_sum, 100)

    def test_dropped_columns_absent_from_rows(self):
        """centroid, centroid_lat, centroid_lon, locations keys never appear on yielded rows."""
        loc = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.ORIGIN)
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=10,
            locations=[loc],
        )
        _, rows = self._explode_rows()
        self.assertEqual(len(rows), 1)
        for k in ("centroid", "centroid_lat", "centroid_lon", "locations"):
            self.assertNotIn(k, rows[0])

    def test_side_query_fires_once_per_chunk(self):
        """The side query against FigureLocation fires exactly ceil(n_figures / chunk_size) times."""
        from apps.entry import models as entry_models

        # Use a tiny chunk_size via monkey-patch so we don't need to create 4000+ figures.
        original = entry_models._EXPLODE_CHUNK_SIZE
        entry_models._EXPLODE_CHUNK_SIZE = 3

        try:
            # Make 7 figures -> ceil(7/3) = 3 chunks.
            for _ in range(7):
                loc = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.ORIGIN)
                self._make_figure(
                    editor=self.editor,
                    country=self.country,
                    event=self.event,
                    entry=self.entry,
                    total_figures=10,
                    locations=[loc],
                )

            # Patch FigureLocation.objects.filter to count calls with figures__id__in kwarg.
            from unittest.mock import patch

            real_filter = FigureLocation.objects.filter
            call_count = {"n": 0}

            def counting_filter(*args, **kwargs):
                if "figures__id__in" in kwargs:
                    call_count["n"] += 1
                return real_filter(*args, **kwargs)

            with patch.object(FigureLocation.objects, "filter", side_effect=counting_filter):
                _, rows = self._explode_rows()

            self.assertEqual(call_count["n"], 3)
            self.assertEqual(len(rows), 7)
        finally:
            entry_models._EXPLODE_CHUNK_SIZE = original

    def test_workbook_renders_real_rows(self):
        """feed the explode export through `get_excel_sheet_content`. The Main sheet
        renders without error, has expected headers, and contains real allocation rows."""
        import io

        from openpyxl import load_workbook

        from apps.contrib.tasks import get_excel_sheet_content

        locs = FigureLocationFactory.create_batch(3, identifier=FigureLocation.IDENTIFIER.ORIGIN)
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=100,
            locations=locs,
        )

        sheet_data = Figure.get_excel_sheets_data(
            user_id=self.editor.id,
            filters={},
            metadata={"explode_by_locations": True},
        )
        workbook = get_excel_sheet_content(**sheet_data)
        buf = io.BytesIO()
        workbook.save(buf)
        buf.seek(0)
        wb = load_workbook(buf)

        ws_main = wb["Main"]
        all_rows = list(ws_main.iter_rows(values_only=True))
        # 1 header row + 3 data rows.
        self.assertEqual(len(all_rows), 4)
        header_row = all_rows[0]
        self.assertIn("Allocated figure", header_row)
        self.assertNotIn("Centroid", header_row)

        alloc_col = header_row.index("Allocated figure")
        identifier_col = header_row.index("Location identifier")
        alloc_values = [row[alloc_col] for row in all_rows[1:]]
        identifier_values = [row[identifier_col] for row in all_rows[1:]]
        self.assertEqual(alloc_values, [34, 33, 33])
        self.assertEqual(identifier_values, ["Origin", "Origin", "Origin"])

        ws_readme = wb["Readme"]
        readme_rows = list(ws_readme.iter_rows(values_only=True))
        first_col_values = {row[0] for row in readme_rows}
        self.assertIn("Allocated figure", first_col_values)
        self.assertIn("About", first_col_values)

    def test_multiple_figures_grouping_isolation_and_order(self):
        """Multiple multi-location figures, split across chunk boundaries:
        - rows are grouped per figure (a figure's rows are contiguous, never interleaved),
        - no cross-figure location leaked (the batched side query groups correctly),
        - figures appear in `created_at` order,
        - within each figure, Origin block precedes Destination block, each location_id ASC,
        - each figure's per-identifier allocation sums to that figure's total_figures.
        """
        import itertools
        from datetime import timedelta

        from django.utils import timezone

        from apps.entry import models as entry_models

        # Figure A: 2 origins + 1 destination, T=10.
        a_o1, a_o2 = FigureLocationFactory.create_batch(2, identifier=FigureLocation.IDENTIFIER.ORIGIN)
        a_d1 = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.DESTINATION)
        fig_a = self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=10,
            locations=[a_o1, a_o2, a_d1],
        )
        # Figure B: a single O&D location, T=7 -> one Origin row + one Destination row.
        b_od = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.ORIGIN_AND_DESTINATION)
        fig_b = self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=7,
            locations=[b_od],
        )
        # Figure C: 3 origins, T=100.
        c_locs = FigureLocationFactory.create_batch(3, identifier=FigureLocation.IDENTIFIER.ORIGIN)
        fig_c = self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=100,
            locations=c_locs,
        )

        # Pin distinct created_at so cross-figure ordering is deterministic (A < B < C),
        # instead of relying on factory timestamps that can tie to microsecond precision.
        now = timezone.now()
        for offset, fig in enumerate((fig_a, fig_b, fig_c)):
            Figure.objects.filter(id=fig.id).update(created_at=now + timedelta(minutes=offset))

        # Force the 3 figures to span more than one chunk so chunk-boundary grouping is exercised.
        original_chunk = entry_models._EXPLODE_CHUNK_SIZE
        entry_models._EXPLODE_CHUNK_SIZE = 2
        try:
            _, rows = self._explode_rows()
        finally:
            entry_models._EXPLODE_CHUNK_SIZE = original_chunk

        # Contiguity + cross-figure order: collapse consecutive runs of figure id.
        collapsed = [figure_id for figure_id, _ in itertools.groupby(r["id"] for r in rows)]
        self.assertEqual(len(collapsed), len(set(collapsed)), "a figure's rows are not contiguous")
        self.assertEqual(collapsed, [fig_a.id, fig_b.id, fig_c.id])

        own_location_ids = {
            fig_a.id: {a_o1.id, a_o2.id, a_d1.id},
            fig_b.id: {b_od.id},
            fig_c.id: {loc.id for loc in c_locs},
        }
        expected_total = {fig_a.id: 10, fig_b.id: 7, fig_c.id: 100}
        for fig_id in collapsed:
            fig_rows = [r for r in rows if r["id"] == fig_id]
            # Isolation: only this figure's own locations appear in its rows.
            self.assertTrue(
                {r["location_id"] for r in fig_rows}.issubset(own_location_ids[fig_id]),
                f"cross-figure location leakage on figure {fig_id}",
            )
            # Origin block before Destination block.
            identifiers = [r["location_identifier"] for r in fig_rows]
            self.assertEqual(
                identifiers,
                sorted(identifiers, key=lambda label: 0 if label == "Origin" else 1),
            )
            # Within each identifier block, location_id ASC, and per-bucket sum == total_figures.
            for identifier in ("Origin", "Destination"):
                bucket = [r for r in fig_rows if r["location_identifier"] == identifier]
                if not bucket:
                    continue
                bucket_ids = [r["location_id"] for r in bucket]
                self.assertEqual(bucket_ids, sorted(bucket_ids))
                self.assertEqual(sum(r["allocated_figure"] for r in bucket), expected_total[fig_id])

    # FIXME: Shouldn't this be in TestExportFigures?
    def test_explode_false_returns_default_export(self):
        """Strict `is True`: an explicit explode_by_locations=False yields the DEFAULT figure
        export shape, not the explode shape."""
        loc = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.ORIGIN)
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=10,
            locations=[loc],
        )
        result = Figure.get_excel_sheets_data(
            user_id=self.editor.id,
            filters={},
            metadata={"explode_by_locations": False},
        )
        headers = result["headers"]
        # Default shape: aggregate location columns present, explode-only columns absent.
        self.assertIn("locations", headers)
        self.assertIn("centroid", headers)
        self.assertNotIn("allocated_figure", headers)
        self.assertNotIn("location_identifier", headers)
        # Default export attaches a transformer; the explode export sets it to None.
        self.assertIsNotNone(result["transformer"])

    def test_regular_export_still_aggregates_locations(self):
        """the `include_location_aggregate` refactor must NOT regress the default
        figure export, its aggregated `locations` column must still be built and rendered."""
        import io

        from openpyxl import load_workbook

        from apps.contrib.tasks import get_excel_sheet_content

        locs = FigureLocationFactory.create_batch(2, identifier=FigureLocation.IDENTIFIER.ORIGIN)
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=10,
            locations=locs,
        )
        sheet_data = Figure.get_excel_sheets_data(user_id=self.editor.id, filters={}, metadata=None)
        self.assertIn("locations", sheet_data["headers"])

        workbook = get_excel_sheet_content(**sheet_data)
        buf = io.BytesIO()
        workbook.save(buf)
        buf.seek(0)
        rows = list(load_workbook(buf)["Main"].iter_rows(values_only=True))
        header = rows[0]
        loc_label = "Locations (Name:Lat, Lon:Accuracy:Type)"
        self.assertIn(loc_label, header)
        # One figure -> one data row, with the aggregated locations cell populated by the transformer.
        self.assertEqual(len(rows), 2)
        self.assertTrue(rows[1][header.index(loc_label)], "aggregated locations cell is empty")

    def test_explode_headers_drop_aggregate_location_columns(self):
        """The explode sheet must NOT carry the aggregate-location columns; only the
        per-location columns remain. Guards against blank/redundant columns and matches the Readme."""
        loc = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.ORIGIN)
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=10,
            locations=[loc],
        )
        result, rows = self._explode_rows()
        for dropped in ("loc_lat_lon", "accuracy", "type_of_points", "locations"):
            self.assertNotIn(dropped, result["headers"])
            self.assertNotIn(dropped, rows[0])

    def test_every_header_key_is_populated_and_lat_lng_format(self):
        """No silently-blank columns: every explode header key is present on every yielded row.
        Also checks the `location_lat_lng` = 'lat, lng' formatting."""
        loc = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.ORIGIN)
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=10,
            locations=[loc],
        )
        result, rows = self._explode_rows()
        header_keys = set(result["headers"].keys())
        self.assertEqual(len(rows), 1)
        for row in rows:
            missing = header_keys - set(row.keys())
            self.assertEqual(missing, set(), f"header columns with no row value (blank): {missing}")

        loc.refresh_from_db()
        self.assertEqual(rows[0]["location_lat_lng"], f"{loc.lat}, {loc.lon}")

    def _drain_explode_query_count(self):
        """Run the explode flow to completion and return the number of DB queries it fired."""
        from django.db import connection
        from django.test.utils import CaptureQueriesContext

        with CaptureQueriesContext(connection) as ctx:
            result = Figure.get_excel_sheets_data(
                user_id=self.editor.id,
                filters={},
                metadata={"explode_by_locations": True},
            )
            list(result["data"])
        return len(ctx.captured_queries)

    def _make_n_single_origin_figures(self, n):
        for _ in range(n):
            loc = FigureLocationFactory.create(identifier=FigureLocation.IDENTIFIER.ORIGIN)
            self._make_figure(
                editor=self.editor,
                country=self.country,
                event=self.event,
                entry=self.entry,
                total_figures=10,
                locations=[loc],
            )

    def test_explode_query_count_scales_with_chunks_not_figures(self):
        """the explode flow's query count is determined by the number
        of CHUNKS, not the number of figures. Doubling figures within a single chunk must not
        change the query count; shrinking the chunk size adds exactly one side query
        per extra chunk."""
        from apps.entry import models as entry_models

        # Default chunk (2000): all figures fall in a single chunk regardless of count.
        self._make_n_single_origin_figures(4)
        q_4_figures = self._drain_explode_query_count()

        self._make_n_single_origin_figures(12)  # 16 figures total, still one chunk
        q_16_figures = self._drain_explode_query_count()

        # Flat in figure count -> no query-per-figure regression.
        self.assertEqual(
            q_4_figures,
            q_16_figures,
            "explode query count grew with figure count -> N+1 regression",
        )

        # Force 4 chunks over the same 16 figures; query count grows by exactly 3 (3 extra
        # side queries: 4 chunks vs the single chunk above).
        original_chunk = entry_models._EXPLODE_CHUNK_SIZE
        entry_models._EXPLODE_CHUNK_SIZE = 4
        try:
            q_4_chunks = self._drain_explode_query_count()
        finally:
            entry_models._EXPLODE_CHUNK_SIZE = original_chunk
        self.assertEqual(
            q_4_chunks - q_16_figures,
            3,
            "extra chunks did not add exactly one side query each",
        )

    def test_explode_query_parity_with_regular_export(self):
        """on identical data within a single chunk, the explode flow fires exactly ONE more query
        than the regular figure export; the per-chunk side query against FigureLocation
        and no additional per-figure/per-row SQL."""
        from django.db import connection
        from django.test.utils import CaptureQueriesContext

        self._make_n_single_origin_figures(5)  # one chunk

        with CaptureQueriesContext(connection) as regular_ctx:
            regular = Figure.get_excel_sheets_data(user_id=self.editor.id, filters={}, metadata=None)
            list(regular["data"])
        regular_queries = len(regular_ctx.captured_queries)

        explode_queries = self._drain_explode_query_count()

        self.assertEqual(
            explode_queries - regular_queries,
            1,
            f"explode added {explode_queries - regular_queries} queries over regular; "
            "expected exactly 1 (the single-chunk side query)",
        )

    def test_generate_excel_file_runs_explode_path_end_to_end(self):
        """DB-persisted metadata={'explode_by_locations': True} drives the
        explode branch, the download completes, and the saved workbook has the explode shape."""
        import io

        from openpyxl import load_workbook

        from apps.contrib.models import ExcelDownload
        from apps.contrib.tasks import generate_excel_file

        locs = FigureLocationFactory.create_batch(3, identifier=FigureLocation.IDENTIFIER.ORIGIN)
        self._make_figure(
            editor=self.editor,
            country=self.country,
            event=self.event,
            entry=self.entry,
            total_figures=100,
            locations=locs,
        )

        download = ExcelDownload.objects.create(
            created_by=self.editor,
            download_type=ExcelDownload.DOWNLOAD_TYPES.FIGURE,
            filters={},
            metadata={"explode_by_locations": True},
        )

        # Call the task body synchronously (no eager-celery config required).
        generate_excel_file(download.id, self.editor.id)

        download.refresh_from_db()
        self.assertEqual(download.status, ExcelDownload.EXCEL_GENERATION_STATUS.COMPLETED)
        self.assertTrue(download.file)
        self.assertGreater(download.file_size or 0, 0)

        with download.file.open("rb") as fh:
            wb = load_workbook(io.BytesIO(fh.read()))
        header_row = next(wb["Main"].iter_rows(values_only=True))
        self.assertIn("Allocated figure", header_row)
        self.assertNotIn("Centroid", header_row)


class TestFigureDelete(HelixGraphQLTestCase):
    def setUp(self) -> None:
        self.country = CountryFactory.create()
        self.country_id = str(self.country.id)
        self.editor = create_user_with_role(USER_ROLE.MONITORING_EXPERT.name)
        self.entry = EntryFactory.create(created_by=self.editor)
        self.event = EventFactory.create(
            event_type=Crisis.CRISIS_TYPE.OTHER.value,
        )
        self.event.countries.add(self.country)
        self.figure = FigureFactory.create(
            entry=self.entry,
            reported=101,
            role=Figure.ROLE.RECOMMENDED,
            unit=Figure.UNIT.PERSON,
            event=self.event,
        )
        self.mutation = """
            mutation DeleteFigure($id: ID!) {
                deleteFigure(id: $id) {
                    ok
                    errors
                    result {
                        id
                    }
                }
            }
        """
        self.variables = {
            "id": self.figure.id,
        }

    def test_can_delete_figure(self):
        self.force_login(self.editor)
        response = self.query(self.mutation, variables=self.variables)
        self.assertResponseNoErrors(response)

        content = json.loads(response.content)
        self.assertTrue(content["data"]["deleteFigure"]["ok"], content)
