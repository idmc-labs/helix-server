import tempfile

from django.core.management import call_command
from django.core.management.base import CommandError
from openpyxl import Workbook, load_workbook

from apps.entry.models import FigureLocation
from utils.factories import FigureLocationFactory
from utils.tests import HelixTestCase


def write_sheet(headers, rows, sheet_name="Data"):
    """Write a temporary .xlsx with the given headers + rows and return its path."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    workbook.save(tmp.name)
    return tmp.name


class TestImportFigureLocationsCommand(HelixTestCase):
    def test_backfill_updates_pcode_fields(self):
        location = FigureLocationFactory.create(display_name="Keep Me")
        path = write_sheet(
            ["id", "pcode", "pcode_source", "pcode_accuracy"],
            [{"id": location.id, "pcode": "ET0102", "pcode_source": "OCHA_COD", "pcode_accuracy": "ADM2"}],
        )
        call_command("import_figure_locations", path)

        location.refresh_from_db()
        self.assertEqual(location.pcode, "ET0102")
        self.assertEqual(location.pcode_source, "OCHA_COD")  # free text, stored verbatim
        self.assertEqual(location.pcode_accuracy, FigureLocation.PCODE_ACCURACY.ADM2.value)
        # Backfill leaves untouched fields alone.
        self.assertEqual(location.display_name, "Keep Me")
        self.assertEqual(FigureLocation.objects.count(), 1)  # updated, not created

    def test_pcode_source_is_free_text(self):
        # pcode_source is not an enum: any string (including future, non-canonical
        # sources) is accepted and stored verbatim.
        location = FigureLocationFactory.create()
        path = write_sheet(
            ["id", "pcode_source", "pcode_accuracy"],
            [{"id": location.id, "pcode_source": "SomeFutureSource", "pcode_accuracy": "ADM0"}],
        )
        call_command("import_figure_locations", path)

        location.refresh_from_db()
        self.assertEqual(location.pcode_source, "SomeFutureSource")
        self.assertEqual(location.pcode_accuracy, FigureLocation.PCODE_ACCURACY.ADM0.value)

    def test_blank_id_rejected_no_create(self):
        # update-only: a row without an id must fail loudly rather than create an orphan.
        path = write_sheet(
            ["id", "pcode", "pcode_source", "pcode_accuracy"],
            [{"pcode": "ET0102", "pcode_source": "OCHA_COD", "pcode_accuracy": "ADM2"}],
        )
        with self.assertRaises(CommandError):
            call_command("import_figure_locations", path)
        self.assertEqual(FigureLocation.objects.count(), 0)

    def test_bad_enum_rejected(self):
        location = FigureLocationFactory.create()
        path = write_sheet(
            ["id", "pcode_accuracy"],
            [{"id": location.id, "pcode_accuracy": "NOT_A_LEVEL"}],
        )
        with self.assertRaises(CommandError):
            call_command("import_figure_locations", path)
        location.refresh_from_db()
        self.assertIsNone(location.pcode_accuracy)  # nothing committed

    def test_missing_update_target_errors(self):
        path = write_sheet(["id", "pcode"], [{"id": 999999, "pcode": "ET0102"}])
        with self.assertRaises(CommandError):
            call_command("import_figure_locations", path)

    def test_blank_pcode_leaves_field_unchanged(self):
        location = FigureLocationFactory.create(pcode="ORIGINAL")
        path = write_sheet(
            ["id", "pcode", "pcode_accuracy"],
            [{"id": location.id, "pcode": "   ", "pcode_accuracy": "ADM3"}],  # blank pcode
        )
        call_command("import_figure_locations", path)

        location.refresh_from_db()
        self.assertEqual(location.pcode, "ORIGINAL")  # blank did NOT clear it
        self.assertEqual(location.pcode_accuracy, FigureLocation.PCODE_ACCURACY.ADM3.value)

    def test_clear_token_clears_pcode(self):
        location = FigureLocationFactory.create(pcode="ORIGINAL")
        path = write_sheet(["id", "pcode"], [{"id": location.id, "pcode": "<clear>"}])
        call_command("import_figure_locations", path)

        location.refresh_from_db()
        self.assertIsNone(location.pcode)

    def template_headers(self):
        out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        call_command("import_figure_locations", "--make-template", out.name)
        return [cell.value for cell in load_workbook(out.name)["Data"][1]]

    def test_template_exposes_only_the_pcode_columns(self):
        # Pinned in full and in order: the importer edits p-codes, so a sheet cannot carry a
        # location's coordinates or name and move it by accident.
        self.assertEqual(
            self.template_headers(),
            ["id", "pcode", "pcode_source", "pcode_accuracy"],
        )

    def test_template_readme_describes_update_only(self):
        out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        call_command("import_figure_locations", "--make-template", out.name)
        readme = load_workbook(out.name)["README"]
        readme_text = " ".join(str(cell.value) for row in readme.iter_rows() for cell in row if cell.value is not None)

        # An update-only importer must not tell the operator a blank id creates a record.
        self.assertNotIn("Leave 'id' blank to CREATE", readme_text)
        self.assertIn("This importer only UPDATES; it never creates records.", readme_text)
        self.assertIn("id of the row to update", readme_text)
        self.assertNotIn("leave blank to create", readme_text)
        self.assertNotIn("Required (create)", readme_text)

        # id is the one column the operator must supply, and the table says so.
        shape = {row[0].value: row[2].value for row in readme.iter_rows() if row[0].value in {"id", "pcode"}}
        self.assertEqual(shape["id"], "yes")
        self.assertEqual(shape["pcode"], "no")

    def test_template_hides_non_pcode_columns(self):
        headers = self.template_headers()
        for hidden in [
            "uuid",
            "geocoder_metadata",
            "moved",
            "created_by",
            "created_at",
            "old_id",
            "lat",
            "lon",
            "display_name",
            "country_code",
        ]:
            self.assertNotIn(hidden, headers)

    def test_a_column_outside_the_pcode_surface_is_rejected(self):
        location = FigureLocationFactory.create()
        original_uuid = location.uuid
        path = write_sheet(["id", "uuid"], [{"id": location.id, "uuid": "0" * 32}])
        with self.assertRaises(CommandError):
            call_command("import_figure_locations", path)
        location.refresh_from_db()
        self.assertEqual(location.uuid, original_uuid)
