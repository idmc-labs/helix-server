import tempfile
from datetime import date
from io import StringIO
from unittest import mock

from django.core.management import call_command
from django.core.management.base import CommandError
from openpyxl import Workbook, load_workbook

from apps.contrib.commons import DATE_ACCURACY
from apps.contrib.management.base.reader import read_rows
from apps.entry.management.commands.import_figures import Command as ImportFiguresCommand
from apps.entry.management.commands.import_figures import FigureRoleAndDatesSerializer
from apps.entry.models import Figure
from utils.factories import CountryFactory, EventFactory, FigureFactory
from utils.tests import HelixTestCase


def write_sheet(headers, rows, sheet_name="Data"):
    """Write a temporary .xlsx with the given headers + rows and return its path."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    workbook.save(tmp.name)
    return tmp.name


class TestImportFiguresCommand(HelixTestCase):
    def setUp(self):
        super().setUp()
        self.country = CountryFactory.create(iso3="NPL", iso2="NP")
        self.event = EventFactory.create(countries=[self.country])
        self.figure = self._figure()

    def _figure(self, **kwargs):
        defaults = dict(
            event=self.event,
            country=self.country,
            unit=Figure.UNIT.PERSON,
            category=Figure.FIGURE_CATEGORY_TYPES.IDPS,
            role=Figure.ROLE.RECOMMENDED,
            start_date=date(2020, 6, 1),
            start_date_accuracy=DATE_ACCURACY.DAY,
            end_date=date(2020, 6, 30),
            end_date_accuracy=DATE_ACCURACY.DAY,
            reported=100,
            total_figures=100,
            household_size=None,
        )
        defaults.update(kwargs)
        return FigureFactory.create(**defaults)

    # ----- the editable surface -----

    def test_only_the_role_and_date_fields_are_importable(self):
        self.assertEqual(
            ImportFiguresCommand().import_columns(),
            ["id", "uuid", "role", "start_date", "end_date", "start_date_accuracy", "end_date_accuracy"],
        )

    def test_a_column_outside_that_surface_is_rejected(self):
        # reported feeds total_figures, which this importer does not derive, so it is not offered.
        path = write_sheet(["id", "reported"], [{"id": self.figure.id, "reported": 250}])
        with self.assertRaises(CommandError) as caught:
            call_command("import_figures", path, stdout=StringIO())

        self.assertIn("Unknown column(s): reported", str(caught.exception))
        self.figure.refresh_from_db()
        self.assertEqual(self.figure.reported, 100)

    def test_the_serializer_does_not_inherit_figure_cross_field_validation(self):
        # The reason this importer has its own serializer: CommonFigureValidationMixin re-checks a
        # figure's whole stored state, which blocks edits to figures with invalid history.
        from apps.entry.serializers import CommonFigureValidationMixin

        self.assertFalse(issubclass(FigureRoleAndDatesSerializer, CommonFigureValidationMixin))

    # ----- core edit path -----

    def test_a_row_patches_the_named_fields(self):
        path = write_sheet(
            ["id", "role", "start_date", "end_date"],
            [
                {
                    "id": self.figure.id,
                    "role": "TRIANGULATION",
                    "start_date": date(2020, 5, 1),
                    "end_date": date(2020, 5, 31),
                }
            ],
        )
        call_command("import_figures", path)

        self.figure.refresh_from_db()
        self.assertEqual(self.figure.role, Figure.ROLE.TRIANGULATION.value)
        self.assertEqual(self.figure.start_date, date(2020, 5, 1))
        self.assertEqual(self.figure.end_date, date(2020, 5, 31))
        self.assertEqual(Figure.objects.count(), 1)  # updated, not created

    def test_a_date_formatted_cell_is_accepted(self):
        # A spreadsheet has no date type distinct from datetime, so openpyxl hands back a datetime.
        path = write_sheet(["id", "start_date"], [{"id": self.figure.id, "start_date": date(2020, 6, 15)}])
        call_command("import_figures", path)

        self.figure.refresh_from_db()
        self.assertEqual(self.figure.start_date, date(2020, 6, 15))

    def test_a_date_written_as_text_is_accepted(self):
        path = write_sheet(["id", "start_date"], [{"id": self.figure.id, "start_date": "2020-06-20"}])
        call_command("import_figures", path)

        self.figure.refresh_from_db()
        self.assertEqual(self.figure.start_date, date(2020, 6, 20))

    def test_a_blank_cell_leaves_the_field_unchanged(self):
        path = write_sheet(
            ["id", "role", "start_date"],
            [{"id": self.figure.id, "role": "TRIANGULATION", "start_date": None}],
        )
        call_command("import_figures", path)

        self.figure.refresh_from_db()
        self.assertEqual(self.figure.role, Figure.ROLE.TRIANGULATION.value)
        self.assertEqual(self.figure.start_date, date(2020, 6, 1))

    def test_the_clear_token_clears_a_stock_figures_end_date(self):
        # A stock figure's end date is its reporting date, and clearing it is a real edit.
        self.assertIn(self.figure.category, Figure.stock_list())
        path = write_sheet(["id", "end_date"], [{"id": self.figure.id, "end_date": "<clear>"}])
        call_command("import_figures", path)

        self.figure.refresh_from_db()
        self.assertIsNone(self.figure.end_date)

    def test_clearing_a_flow_figures_end_date_is_refused(self):
        # FigureSerializer._validate_category compares a flow figure's end_date against today with
        # no null guard, so a cleared end_date makes the figure raise TypeError on every later save.
        figure = self._figure(category=Figure.FIGURE_CATEGORY_TYPES.NEW_DISPLACEMENT)
        self.assertIn(figure.category, Figure.flow_list())

        path = write_sheet(["id", "end_date"], [{"id": figure.id, "end_date": "<clear>"}])
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)

        self.assertIn("must keep an end date", out.getvalue())
        figure.refresh_from_db()
        self.assertIsNotNone(figure.end_date)

    def test_a_date_beyond_the_apps_future_bound_is_refused(self):
        # Writing a date the app would reject leaves a row that fails validation on every later
        # edit, which is the trap the narrow serializer exists to avoid.
        path = write_sheet(
            ["id", "start_date", "end_date"],
            [{"id": self.figure.id, "start_date": date(2999, 1, 1), "end_date": date(2999, 12, 31)}],
        )
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)

        self.assertIn("years in the future", out.getvalue())
        self.figure.refresh_from_db()
        self.assertEqual(self.figure.start_date, date(2020, 6, 1))

    def test_a_future_end_date_on_a_flow_figure_is_refused(self):
        figure = self._figure(category=Figure.FIGURE_CATEGORY_TYPES.NEW_DISPLACEMENT)
        path = write_sheet(["id", "end_date"], [{"id": figure.id, "end_date": date(2030, 1, 1)}])
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)

        self.assertIn("past date", out.getvalue())

    def test_a_malformed_key_is_a_row_error_not_a_traceback(self):
        # A key column takes raw sheet text; a typo must join the per-row error list rather than
        # replace the whole report with a traceback carrying no row number.
        for column, bad in (("uuid", "not-a-uuid"), ("id", "12a")):
            with self.subTest(column=column):
                path = write_sheet([column, "role"], [{column: bad, "role": "TRIANGULATION"}])
                out = StringIO()
                with self.assertRaises(CommandError):
                    call_command("import_figures", path, stdout=out)
                self.assertIn(f"Row 2: {column}", out.getvalue())
                self.assertIn("is not a valid", out.getvalue())

    def test_the_readme_states_the_two_key_rule(self):
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        call_command("import_figures", make_template=tmp.name, stdout=StringIO())
        workbook = load_workbook(tmp.name)

        text = "\n".join(str(cell.value) for row in workbook["README"].iter_rows() for cell in row if cell.value is not None)
        self.assertIn("exactly one of", text)
        # Neither key is marked required on its own, since supplying both is rejected.
        shape = [
            (row[0].value, row[2].value) for row in workbook["README"].iter_rows() if row and row[0].value in ("id", "uuid")
        ]
        self.assertEqual(sorted(shape), [("id", "no"), ("uuid", "no")])

    def test_an_end_date_accuracy_cell_is_written_rather_than_cleared(self):
        # FigureSerializer nulls end_date_accuracy for a stock category (IDPS is one). Without that
        # mixin the cell lands, which is the point of the narrow serializer.
        self.figure.end_date_accuracy = None
        self.figure.save()
        path = write_sheet(["id", "end_date_accuracy"], [{"id": self.figure.id, "end_date_accuracy": "MONTH"}])
        out = StringIO()
        call_command("import_figures", path, stdout=out)

        self.figure.refresh_from_db()
        self.assertEqual(self.figure.end_date_accuracy, DATE_ACCURACY.MONTH.value)
        self.assertNotIn("CELL_IGNORED", out.getvalue())

    def test_enum_columns_resolve_by_member_name(self):
        path = write_sheet(
            ["id", "role", "start_date_accuracy"],
            [{"id": self.figure.id, "role": "TRIANGULATION", "start_date_accuracy": "YEAR"}],
        )
        call_command("import_figures", path)

        self.figure.refresh_from_db()
        self.assertEqual(self.figure.role, Figure.ROLE.TRIANGULATION.value)
        self.assertEqual(self.figure.start_date_accuracy, DATE_ACCURACY.YEAR.value)

    def test_an_unknown_enum_value_fails_the_row(self):
        path = write_sheet(["id", "role"], [{"id": self.figure.id, "role": "NOT_A_ROLE"}])
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)

        self.assertIn("invalid value 'NOT_A_ROLE'", out.getvalue())
        self.figure.refresh_from_db()
        self.assertEqual(self.figure.role, Figure.ROLE.RECOMMENDED.value)

    def test_an_unparseable_date_fails_the_row(self):
        path = write_sheet(["id", "start_date"], [{"id": self.figure.id, "start_date": "not-a-date"}])
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)

        self.assertIn("start_date", out.getvalue())
        self.figure.refresh_from_db()
        self.assertEqual(self.figure.start_date, date(2020, 6, 1))

    # ----- the one cross-field rule kept -----

    def test_a_start_date_after_the_supplied_end_date_fails_the_row(self):
        path = write_sheet(
            ["id", "start_date", "end_date"],
            [{"id": self.figure.id, "start_date": date(2020, 7, 1), "end_date": date(2020, 6, 1)}],
        )
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)

        self.assertIn("is after end_date", out.getvalue())
        self.figure.refresh_from_db()
        self.assertEqual(self.figure.start_date, date(2020, 6, 1))

    def test_a_start_date_after_the_stored_end_date_fails_the_row(self):
        # Only start_date is supplied, so the check has to consult the stored end_date.
        path = write_sheet(["id", "start_date"], [{"id": self.figure.id, "start_date": date(2021, 1, 1)}])
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)

        self.assertIn("is after end_date 2020-06-30", out.getvalue())
        self.figure.refresh_from_db()
        self.assertEqual(self.figure.start_date, date(2020, 6, 1))

    def test_a_figure_with_no_end_date_accepts_any_start_date(self):
        self.figure.end_date = None
        self.figure.save()
        path = write_sheet(["id", "start_date"], [{"id": self.figure.id, "start_date": date(2021, 1, 1)}])
        call_command("import_figures", path)

        self.figure.refresh_from_db()
        self.assertEqual(self.figure.start_date, date(2021, 1, 1))

    def test_a_figure_whose_stored_dates_are_inverted_can_still_have_them_fixed(self):
        # The case the narrow serializer exists for: bad history must not block its own repair.
        Figure.objects.filter(pk=self.figure.pk).update(start_date=date(2020, 9, 1), end_date=date(2020, 6, 30))
        path = write_sheet(
            ["id", "start_date", "end_date"],
            [{"id": self.figure.id, "start_date": date(2020, 6, 1), "end_date": date(2020, 6, 30)}],
        )
        call_command("import_figures", path)

        self.figure.refresh_from_db()
        self.assertEqual(self.figure.start_date, date(2020, 6, 1))

    # ----- the date rules are scoped to rows that supply a date -----

    # A role correction is the whole reason this importer exists, and the figures needing one are
    # the oldest — the ones whose stored dates predate today's rules. Checking those dates on a row
    # that does not touch them would refuse the correction, and since the import is all-or-nothing,
    # one such figure would stop every other row in the sheet.

    def test_a_role_only_row_is_accepted_on_a_flow_figure_with_no_stored_end_date(self):
        figure = self._figure(category=Figure.FIGURE_CATEGORY_TYPES.NEW_DISPLACEMENT)
        Figure.objects.filter(pk=figure.pk).update(end_date=None)

        path = write_sheet(["id", "role"], [{"id": figure.id, "role": "TRIANGULATION"}])
        call_command("import_figures", path)

        figure.refresh_from_db()
        self.assertEqual(figure.role, Figure.ROLE.TRIANGULATION.value)
        self.assertIsNone(figure.end_date)

    def test_a_role_only_row_is_accepted_on_a_flow_figure_whose_stored_end_date_is_future(self):
        figure = self._figure(category=Figure.FIGURE_CATEGORY_TYPES.NEW_DISPLACEMENT)
        Figure.objects.filter(pk=figure.pk).update(end_date=date(2030, 1, 1))

        path = write_sheet(["id", "role"], [{"id": figure.id, "role": "TRIANGULATION"}])
        call_command("import_figures", path)

        figure.refresh_from_db()
        self.assertEqual(figure.role, Figure.ROLE.TRIANGULATION.value)
        self.assertEqual(figure.end_date, date(2030, 1, 1))

    def test_a_role_only_row_is_accepted_on_a_figure_whose_stored_dates_are_inverted(self):
        Figure.objects.filter(pk=self.figure.pk).update(start_date=date(2020, 9, 1), end_date=date(2020, 6, 30))

        path = write_sheet(["id", "role"], [{"id": self.figure.id, "role": "TRIANGULATION"}])
        call_command("import_figures", path)

        self.figure.refresh_from_db()
        self.assertEqual(self.figure.role, Figure.ROLE.TRIANGULATION.value)
        self.assertEqual(self.figure.start_date, date(2020, 9, 1))

    def test_an_accuracy_only_row_is_accepted_on_a_figure_whose_stored_dates_are_inverted(self):
        # An accuracy is not a date and no rule reads one, so an accuracy cannot make any of the
        # date rules newly true. Folding the accuracies into the group would refuse this row over
        # dates it never touches.
        Figure.objects.filter(pk=self.figure.pk).update(start_date=date(2020, 9, 1), end_date=date(2020, 6, 30))

        path = write_sheet(["id", "end_date_accuracy"], [{"id": self.figure.id, "end_date_accuracy": "MONTH"}])
        call_command("import_figures", path)

        self.figure.refresh_from_db()
        self.assertEqual(self.figure.end_date_accuracy, DATE_ACCURACY.MONTH.value)
        self.assertEqual(self.figure.end_date, date(2020, 6, 30))

    # ----- naming a row by id or by uuid -----

    def test_a_row_identified_by_uuid_patches_that_figure(self):
        other = self._figure()
        path = write_sheet(["uuid", "role"], [{"uuid": str(self.figure.uuid), "role": "TRIANGULATION"}])
        call_command("import_figures", path)

        self.figure.refresh_from_db()
        other.refresh_from_db()
        self.assertEqual(self.figure.role, Figure.ROLE.TRIANGULATION.value)
        self.assertEqual(other.role, Figure.ROLE.RECOMMENDED.value)

    def test_an_unknown_uuid_fails_the_row(self):
        path = write_sheet(
            ["uuid", "role"],
            [{"uuid": "4a1c9f2e-0000-4000-8000-000000000000", "role": "TRIANGULATION"}],
        )
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)

        self.assertIn("no Figure found with uuid 4a1c9f2e-0000-4000-8000-000000000000", out.getvalue())

    def test_supplying_both_keys_fails_the_row(self):
        path = write_sheet(
            ["id", "uuid", "role"],
            [{"id": self.figure.id, "uuid": str(self.figure.uuid), "role": "TRIANGULATION"}],
        )
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)

        self.assertIn("exactly one of id · uuid is required; 2 given", out.getvalue())
        self.figure.refresh_from_db()
        self.assertEqual(self.figure.role, Figure.ROLE.RECOMMENDED.value)

    def test_supplying_neither_key_fails_the_row(self):
        path = write_sheet(["id", "uuid", "role"], [{"id": None, "uuid": None, "role": "TRIANGULATION"}])
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)
        output = out.getvalue()

        self.assertIn("exactly one of id · uuid is required; none given", output)
        self.assertIn("only updates existing Figure rows", output)

    def test_a_uuid_shared_by_two_figures_fails_the_row(self):
        # Figure.uuid lost its unique constraint in 2021 and helix stores uuids it did not
        # generate, so a shared uuid must not resolve to an arbitrary figure.
        shared = self.figure.uuid
        twin = self._figure()
        Figure.objects.filter(pk=twin.pk).update(uuid=shared)

        path = write_sheet(["uuid", "role"], [{"uuid": str(shared), "role": "TRIANGULATION"}])
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)
        output = out.getvalue()

        self.assertIn(f"uuid {shared} matches more than one Figure", output)
        self.assertIn(str(self.figure.pk), output)
        self.assertIn(str(twin.pk), output)

    def test_one_figure_named_by_id_on_one_row_and_uuid_on_another_is_a_duplicate(self):
        path = write_sheet(
            ["id", "uuid", "role"],
            [
                {"id": self.figure.id, "uuid": None, "role": "TRIANGULATION"},
                {"id": None, "uuid": str(self.figure.uuid), "role": "RECOMMENDED"},
            ],
        )
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)

        self.assertIn(f"Figure {self.figure.id} also appears on row 2", out.getvalue())

    def test_a_uuid_keyed_edit_leaves_the_uuid_alone(self):
        original = self.figure.uuid
        path = write_sheet(["uuid", "role"], [{"uuid": str(original), "role": "TRIANGULATION"}])
        out = StringIO()
        call_command("import_figures", path, stdout=out)

        self.figure.refresh_from_db()
        self.assertEqual(self.figure.uuid, original)
        self.assertNotIn("uuid=", out.getvalue())
        self.assertNotIn("id=", out.getvalue())

    def test_blank_id_rejected_no_create(self):
        path = write_sheet(["id", "role"], [{"id": None, "role": "TRIANGULATION"}])
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=StringIO())

        self.assertEqual(Figure.objects.count(), 1)

    def test_an_unknown_id_fails_the_row(self):
        path = write_sheet(["id", "role"], [{"id": 9999999, "role": "TRIANGULATION"}])
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)

        self.assertIn("no Figure found with id 9999999", out.getvalue())

    # ----- template -----

    def test_make_template_writes_the_expected_columns(self):
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        call_command("import_figures", make_template=tmp.name, stdout=StringIO())

        workbook = load_workbook(tmp.name)
        self.assertIn("Data", workbook.sheetnames)
        headers = [cell.value for cell in workbook["Data"][1]]
        self.assertEqual(headers, ImportFiguresCommand().import_columns())
        self.assertEqual(headers[:2], ["id", "uuid"])
        # Neither key is required on its own: a row supplies exactly one.
        self.assertEqual(ImportFiguresCommand().required_create_columns(), set())

    # ----- an edited template -----

    def test_the_readme_sheet_is_not_required(self):
        path = write_sheet(["id", "role"], [{"id": self.figure.id, "role": "TRIANGULATION"}])
        self.assertEqual(load_workbook(path).sheetnames, ["Data"])
        call_command("import_figures", path)

        self.figure.refresh_from_db()
        self.assertEqual(self.figure.role, Figure.ROLE.TRIANGULATION.value)

    def test_columns_may_be_dropped_and_reordered(self):
        path = write_sheet(["role", "id"], [{"id": self.figure.id, "role": "TRIANGULATION"}])
        call_command("import_figures", path)

        self.figure.refresh_from_db()
        self.assertEqual(self.figure.role, Figure.ROLE.TRIANGULATION.value)
        self.assertEqual(self.figure.start_date, date(2020, 6, 1))

    def test_dropping_every_key_column_is_refused(self):
        path = write_sheet(["role"], [{"role": "TRIANGULATION"}])
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)

        self.assertIn("exactly one of id · uuid is required; none given", out.getvalue())

    def test_a_sheet_of_only_the_key_column_changes_nothing(self):
        path = write_sheet(["id"], [{"id": self.figure.id}])
        out = StringIO()
        call_command("import_figures", path, stdout=out)

        self.assertIn("1 of the updated rows had no effective change.", out.getvalue())

    def test_a_header_only_sheet_imports_nothing(self):
        path = write_sheet(["id", "role"], [])
        out = StringIO()
        call_command("import_figures", path, stdout=out)

        self.assertIn("Created 0, updated 0.", out.getvalue())

    def test_a_duplicated_header_is_refused(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet.append(["id", "role", "role"])
        worksheet.append([self.figure.id, "TRIANGULATION", None])
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        workbook.save(tmp.name)

        with self.assertRaises(CommandError) as caught:
            call_command("import_figures", tmp.name, stdout=StringIO())

        self.assertIn("Duplicate column(s): role", str(caught.exception))
        self.figure.refresh_from_db()
        self.assertEqual(self.figure.role, Figure.ROLE.RECOMMENDED.value)

    def test_a_cleared_header_cell_is_refused_rather_than_shifting_the_columns(self):
        path = write_sheet(
            ["id", "start_date_accuracy", "role"],
            [{"id": self.figure.id, "start_date_accuracy": "WEEK", "role": "TRIANGULATION"}],
        )
        workbook = load_workbook(path)
        workbook["Data"]["B1"] = None  # header cleared, column still there
        workbook.save(path)

        with self.assertRaises(CommandError) as caught:
            call_command("import_figures", path, stdout=StringIO())

        self.assertIn("Column(s) B have no header", str(caught.exception))
        self.figure.refresh_from_db()
        self.assertEqual(self.figure.role, Figure.ROLE.RECOMMENDED.value)

    def test_a_trailing_blank_header_over_an_empty_column_is_tolerated(self):
        path = write_sheet(["id", "role"], [{"id": self.figure.id, "role": "TRIANGULATION"}])
        workbook = load_workbook(path)
        workbook["Data"]["D1"] = "   "
        workbook.save(path)

        call_command("import_figures", path)
        self.figure.refresh_from_db()
        self.assertEqual(self.figure.role, Figure.ROLE.TRIANGULATION.value)

    def test_a_value_in_an_unheaded_column_is_refused(self):
        path = write_sheet(["id", "role"], [{"id": self.figure.id, "role": "TRIANGULATION"}])
        workbook = load_workbook(path)
        workbook["Data"]["D2"] = "orphaned value"
        workbook.save(path)

        with self.assertRaises(CommandError) as caught:
            call_command("import_figures", path, stdout=StringIO())

        self.assertIn("the header row does not name", str(caught.exception))

    def test_a_blank_row_between_rows_is_skipped(self):
        other = self._figure()
        path = write_sheet(["id", "role"], [{"id": self.figure.id, "role": "TRIANGULATION"}])
        workbook = load_workbook(path)
        worksheet = workbook["Data"]
        worksheet.append([None, None])
        worksheet.append([other.id, "TRIANGULATION"])
        workbook.save(path)

        out = StringIO()
        call_command("import_figures", path, stdout=out)

        self.assertIn("Created 0, updated 2.", out.getvalue())

    # ----- row counting and numbering -----

    def test_three_data_rows_update_exactly_three_figures(self):
        # The header is consumed by read_rows while apply_rows numbers rows from 2, so the count
        # and the reported row numbers are two separate chances to be off by one.
        figures = [self.figure, self._figure(), self._figure()]
        path = write_sheet(
            ["id", "start_date"],
            [{"id": figure.id, "start_date": date(2020, 6, 2 + index)} for index, figure in enumerate(figures)],
        )
        out = StringIO()
        call_command("import_figures", path, stdout=out)
        output = out.getvalue()

        self.assertIn("Created 0, updated 3.", output)
        for index, figure in enumerate(figures):
            figure.refresh_from_db()
            self.assertEqual(figure.start_date, date(2020, 6, 2 + index))

        changed = [line for line in output.splitlines() if line.startswith("ROW_UPDATED")]
        self.assertEqual(len(changed), 3)
        self.assertEqual(sorted(line.split("\trow=")[1].split("\t")[0] for line in changed), ["2", "3", "4"])

    def test_a_row_number_in_an_error_points_at_the_offending_sheet_row(self):
        path = write_sheet(
            ["id", "role"],
            [
                {"id": self.figure.id, "role": "TRIANGULATION"},
                {"id": self._figure().id, "role": "TRIANGULATION"},
                {"id": 9999999, "role": "TRIANGULATION"},
            ],
        )
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)

        self.assertIn("Row 4: id: no Figure found with id 9999999", out.getvalue())

    # ----- the changelog -----

    def test_the_changelog_names_only_the_fields_that_moved(self):
        path = write_sheet(
            ["id", "role", "start_date"],
            [{"id": self.figure.id, "role": "TRIANGULATION", "start_date": date(2020, 6, 1)}],
        )
        out = StringIO()
        call_command("import_figures", path, stdout=out)
        output = out.getvalue()

        self.assertIn(f"ROW_UPDATED\tfigure={self.figure.id}\trow=2", output)
        # An enum reports its member name, not the translated label.
        self.assertIn("role=RECOMMENDED->TRIANGULATION", output)
        # Supplied but identical, so it did not move and is not reported.
        self.assertNotIn("start_date=", output)

    def test_a_row_that_changes_nothing_reports_no_changelog_line(self):
        path = write_sheet(["id", "role"], [{"id": self.figure.id, "role": "RECOMMENDED"}])
        out = StringIO()
        call_command("import_figures", path, stdout=out)
        output = out.getvalue()

        self.assertNotIn("ROW_UPDATED", output)
        self.assertIn("1 of the updated rows had no effective change.", output)

    def test_an_edit_leaves_the_review_status_alone(self):
        # The narrow serializer does not run update_figure_status, so correcting a date no longer
        # un-approves an approved figure.
        Figure.objects.filter(pk=self.figure.pk).update(review_status=Figure.FIGURE_REVIEW_STATUS.APPROVED)
        path = write_sheet(["id", "start_date"], [{"id": self.figure.id, "start_date": date(2020, 6, 5)}])
        call_command("import_figures", path)

        self.figure.refresh_from_db()
        self.assertEqual(self.figure.start_date, date(2020, 6, 5))
        self.assertEqual(self.figure.review_status, Figure.FIGURE_REVIEW_STATUS.APPROVED.value)

    def test_an_edit_leaves_the_derived_total_alone(self):
        # total_figures is derived from reported and household_size, neither of which is editable
        # here, so an edit must not disturb it.
        path = write_sheet(["id", "role"], [{"id": self.figure.id, "role": "TRIANGULATION"}])
        out = StringIO()
        call_command("import_figures", path, stdout=out)

        self.figure.refresh_from_db()
        self.assertEqual(self.figure.total_figures, 100)
        self.assertNotIn("total_figures", out.getvalue())

    # ----- all-or-nothing -----

    def test_dry_run_rolls_back(self):
        path = write_sheet(["id", "role"], [{"id": self.figure.id, "role": "TRIANGULATION"}])
        out = StringIO()
        call_command("import_figures", path, dry_run=True, stdout=out)

        self.figure.refresh_from_db()
        self.assertEqual(self.figure.role, Figure.ROLE.RECOMMENDED.value)
        self.assertIn("DRY RUN", out.getvalue())

    def test_one_invalid_row_leaves_every_other_row_unapplied(self):
        other = self._figure()
        path = write_sheet(
            ["id", "role"],
            [
                {"id": self.figure.id, "role": "TRIANGULATION"},
                {"id": other.id, "role": "TRIANGULATION"},
                {"id": 9999999, "role": "TRIANGULATION"},
            ],
        )
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)

        self.assertIn("nothing committed", out.getvalue())
        self.figure.refresh_from_db()
        other.refresh_from_db()
        self.assertEqual(self.figure.role, Figure.ROLE.RECOMMENDED.value)
        self.assertEqual(other.role, Figure.ROLE.RECOMMENDED.value)

    def test_a_failure_while_saving_rolls_back_the_rows_already_saved(self):
        # Validation passed for every row, so this is the case per-item isolation would paper over.
        other = self._figure()
        path = write_sheet(
            ["id", "role"],
            [{"id": self.figure.id, "role": "TRIANGULATION"}, {"id": other.id, "role": "TRIANGULATION"}],
        )

        real_save = FigureRoleAndDatesSerializer.save
        calls = {"n": 0}

        def failing_save(self, **kwargs):
            calls["n"] += 1
            if calls["n"] == 2:
                raise RuntimeError("save blew up on the second row")
            return real_save(self, **kwargs)

        with mock.patch.object(FigureRoleAndDatesSerializer, "save", failing_save):
            with self.assertRaises(RuntimeError):
                call_command("import_figures", path, stdout=StringIO())

        self.assertEqual(calls["n"], 2)
        self.figure.refresh_from_db()
        other.refresh_from_db()
        self.assertEqual(self.figure.role, Figure.ROLE.RECOMMENDED.value)
        self.assertEqual(other.role, Figure.ROLE.RECOMMENDED.value)

    # ----- rows are validated and saved one at a time -----

    # A serializer deep-copies every declared field on construction, so holding one per row made a
    # large sheet cost gigabytes. Rows are saved as they are read instead, and the transaction — not
    # a deferred second pass — is what keeps the run all-or-nothing.

    def test_a_failing_run_prints_no_changelog_for_the_rows_it_saved(self):
        # Earlier rows are written before the bad row is reached, and then rolled back. Printing
        # their changelog lines would claim changes the database never kept.
        other = self._figure()
        path = write_sheet(
            ["id", "role"],
            [
                {"id": self.figure.id, "role": "TRIANGULATION"},
                {"id": other.id, "role": "TRIANGULATION"},
                {"id": 9999999, "role": "TRIANGULATION"},
            ],
        )
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)

        output = out.getvalue()
        self.assertNotIn("ROW_UPDATED", output)
        self.assertIn("nothing committed", output)

    def test_every_invalid_row_is_reported_not_just_the_first(self):
        # Validation keeps going after a row fails, so the operator gets the whole list in one run.
        path = write_sheet(
            ["id", "role"],
            [
                {"id": 9999998, "role": "TRIANGULATION"},
                {"id": self.figure.id, "role": "NOT_A_ROLE"},
                {"id": 9999999, "role": "TRIANGULATION"},
            ],
        )
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)

        output = out.getvalue()
        self.assertIn("3 error(s)", output)
        self.assertIn("Row 2: id: no Figure found with id 9999998", output)
        self.assertIn("Row 3: role:", output)
        self.assertIn("Row 4: id: no Figure found with id 9999999", output)

    def test_no_row_is_saved_once_a_row_has_failed(self):
        # The run will roll back, so writing later rows only to discard them is wasted work.
        other = self._figure()
        path = write_sheet(
            ["id", "role"],
            [
                {"id": 9999999, "role": "TRIANGULATION"},
                {"id": self.figure.id, "role": "TRIANGULATION"},
                {"id": other.id, "role": "TRIANGULATION"},
            ],
        )
        calls = {"n": 0}
        real_save = FigureRoleAndDatesSerializer.save

        def counting_save(self, **kwargs):
            calls["n"] += 1
            return real_save(self, **kwargs)

        with mock.patch.object(FigureRoleAndDatesSerializer, "save", counting_save):
            with self.assertRaises(CommandError):
                call_command("import_figures", path, stdout=StringIO())

        self.assertEqual(calls["n"], 0)

    def test_the_reader_yields_rows_rather_than_building_a_list(self):
        # The whole point of the change: a sheet is never materialised in full.
        path = write_sheet(["id", "role"], [{"id": self.figure.id, "role": "TRIANGULATION"}])
        rows = read_rows(path, data_sheet="Data", allowed_columns=ImportFiguresCommand().import_columns())

        self.assertFalse(isinstance(rows, list))
        self.assertIs(iter(rows), iter(rows))  # an iterator, consumed once
        self.assertEqual([row["role"] for row in rows], ["TRIANGULATION"])

    def test_a_bad_header_fails_before_any_row_is_read(self):
        # Header checks stay eager despite the lazy row iteration, so a malformed sheet is rejected
        # up front rather than part-way through the run.
        path = write_sheet(["id", "role", "not_a_column"], [{"id": self.figure.id, "role": "TRIANGULATION"}])
        with self.assertRaises(CommandError):
            read_rows(path, data_sheet="Data", allowed_columns=ImportFiguresCommand().import_columns())

    # ----- match keys are resolved a chunk at a time -----

    def test_a_chunk_resolves_all_its_rows_in_one_query(self):
        # The point of chunked resolution: cost is per chunk, not per row. Resolving row by row was
        # most of the run on a large sheet.
        figures = [self._figure() for _ in range(5)]
        chunk = [{"id": figure.id, "role": "TRIANGULATION"} for figure in figures]
        command = ImportFiguresCommand()

        with self.assertNumQueries(1):
            resolved = command.resolve_chunk(chunk)

        self.assertEqual(sorted(resolved["id"]), sorted(figure.id for figure in figures))

    def test_a_chunk_costs_one_query_per_match_column_in_use(self):
        # A column no row fills is not queried at all.
        by_id, by_uuid = self._figure(), self._figure()
        chunk = [
            {"id": by_id.id, "role": "TRIANGULATION"},
            {"uuid": str(by_uuid.uuid), "role": "TRIANGULATION"},
        ]
        with self.assertNumQueries(2):
            resolved = ImportFiguresCommand().resolve_chunk(chunk)

        self.assertIn(by_id.id, resolved["id"])
        self.assertIn(by_uuid.uuid, resolved["uuid"])

    def test_chunked_resolution_still_finds_an_ambiguous_key(self):
        # Every match is kept, not just the first, so a non-unique uuid is still caught. Figure.uuid
        # carries whatever uuid an external import supplied and is not unique.
        shared = self.figure.uuid
        twin = self._figure()
        Figure.objects.filter(pk=twin.pk).update(uuid=shared)

        resolved = ImportFiguresCommand().resolve_chunk([{"uuid": str(shared)}])
        self.assertEqual(len(resolved["uuid"][shared]), 2)

    def test_a_malformed_key_does_not_cost_its_chunk_the_query(self):
        # One unparseable cell must not stop the rest of the chunk resolving; it becomes that row's
        # own error when the row is prepared.
        figure = self._figure()
        chunk = [{"id": "not-a-number"}, {"id": figure.id}]

        with self.assertNumQueries(1):
            resolved = ImportFiguresCommand().resolve_chunk(chunk)

        self.assertIn(figure.id, resolved["id"])

    def test_rows_keep_their_sheet_numbers_across_chunk_boundaries(self):
        # Row numbers come from the chunk offset now, so a small chunk size must not renumber them.
        figures = [self._figure() for _ in range(3)]
        rows = [{"id": figure.id, "role": "TRIANGULATION"} for figure in figures]
        rows.append({"id": 9999999, "role": "TRIANGULATION"})  # row 5
        path = write_sheet(["id", "role"], rows)

        out = StringIO()
        with mock.patch.object(ImportFiguresCommand, "RESOLVE_CHUNK_SIZE", 2):
            with self.assertRaises(CommandError):
                call_command("import_figures", path, stdout=out)

        self.assertIn("Row 5: id: no Figure found with id 9999999", out.getvalue())

    def test_a_reused_serializer_does_not_carry_the_previous_rows_errors(self):
        other = self._figure()
        path = write_sheet(
            ["id", "role"],
            [
                {"id": self.figure.id, "role": "NOT_A_ROLE"},  # row 2 fails
                {"id": other.id, "role": "TRIANGULATION"},  # row 3 is fine
            ],
        )
        out = StringIO()
        with self.assertRaises(CommandError):
            call_command("import_figures", path, stdout=out)

        output = out.getvalue()
        self.assertIn("1 error(s)", output)
        self.assertIn("Row 2: role:", output)
        self.assertNotIn("Row 3", output)

    def test_a_reused_serializer_does_not_carry_the_previous_rows_validated_data(self):
        # Row 2 supplies role, row 3 supplies only start_date. Leaked validated data would write
        # row 2's role onto row 3's figure.
        other = self._figure()
        path = write_sheet(
            ["id", "role", "start_date"],
            [
                {"id": self.figure.id, "role": "TRIANGULATION"},
                {"id": other.id, "start_date": date(2020, 6, 5)},
            ],
        )
        call_command("import_figures", path, stdout=StringIO())

        self.figure.refresh_from_db()
        other.refresh_from_db()
        self.assertEqual(self.figure.role, Figure.ROLE.TRIANGULATION.value)
        self.assertEqual(other.role, Figure.ROLE.RECOMMENDED.value)  # untouched by row 2
        self.assertEqual(other.start_date, date(2020, 6, 5))
