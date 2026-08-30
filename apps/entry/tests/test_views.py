import json
import tempfile

from django.core.files.base import ContentFile
from openpyxl import load_workbook
from rest_framework import status

from apps.contrib.tasks import generate_external_endpoint_dump_file
from apps.entry.models import ExternalApiDump, Figure
from apps.entry.serializers import FigureReadOnlySerializer
from apps.entry.views import get_idu_data, get_idu_data_excel, get_idu_data_geojson
from utils.factories import (
    ClientFactory,
    CountryFactory,
    EntryFactory,
    EventFactory,
    FigureFactory,
    FigureLocationFactory,
    OrganizationFactory,
)
from utils.tests import HelixAPITestCase, HelixTestCase


class TestGetIduData(HelixTestCase):
    """The IDU export must only publish source links when the requesting client
    has 'share_source' enabled"""

    def setUp(self):
        super().setUp()
        self.country = CountryFactory.create()
        self.event = EventFactory.create()
        self.source = OrganizationFactory.create(name="ACLED")
        self.source_url = "https://example.com/source-report"
        self.entry = EntryFactory.create(url=self.source_url, is_confidential=False)
        self.figure = FigureFactory.create(
            entry=self.entry,
            country=self.country,
            event=self.event,
            category=Figure.FIGURE_CATEGORY_TYPES.NEW_DISPLACEMENT.value,
            role=Figure.ROLE.RECOMMENDED.value,
            include_idu=True,
            excerpt_idu="Some displacement happened.",
        )
        self.figure.sources.add(self.source)

    def _get_rows(self, include_sources):
        return list(get_idu_data({"include_sources": include_sources, "id": self.figure.id}))

    def test_source_links_excluded_when_sharing_disabled(self):
        rows = self._get_rows(include_sources=False)
        self.assertEqual(len(rows), 1)
        row = rows[0]

        # The source URL must be blanked out and no anchor/url must be included in
        # the popup text
        self.assertEqual(row["entry_url_or_document_url"], "")
        self.assertNotIn("<a href", row["standard_popup_text"])
        self.assertNotIn(self.source_url, row["standard_popup_text"])
        self.assertNotIn(self.source_url, row["custom_link_text"])

        # Source names are not blanked out only the links
        self.assertIn(self.source.name, row["sources_name"])

    def test_source_links_included_when_sharing_enabled(self):
        rows = self._get_rows(include_sources=True)
        self.assertEqual(len(rows), 1)
        row = rows[0]

        self.assertEqual(row["entry_url_or_document_url"], self.source_url)
        self.assertIn("<a href", row["standard_popup_text"])
        self.assertIn(self.source_url, row["standard_popup_text"])
        self.assertIn(self.source_url, row["custom_link_text"])

        self.assertIn(self.source.name, row["sources_name"])


class TestIduDumpGeneration(HelixTestCase):
    """The streamed dumps must stay valid, parseable, and value-identical to the
    per-record serializer output"""

    def setUp(self):
        super().setUp()
        self.country = CountryFactory.create(idmc_short_name="Nepal")
        self.event = EventFactory.create()
        self.entry = EntryFactory.create(url="https://example.com/source-report", is_confidential=False)
        self.location = FigureLocationFactory.create()
        self.figure = FigureFactory.create(
            entry=self.entry,
            country=self.country,
            event=self.event,
            category=Figure.FIGURE_CATEGORY_TYPES.NEW_DISPLACEMENT.value,
            role=Figure.ROLE.RECOMMENDED.value,
            include_idu=True,
            excerpt_idu="Some displacement happened.",
            geo_locations=[self.location],
        )

    def test_json_dump_matches_per_record_serializer_output(self):
        generate_external_endpoint_dump_file(
            ExternalApiDump.ExternalApiType.IDUS_ALL,
            FigureReadOnlySerializer,
            get_idu_data,
            "idus_all.json",
            ExternalApiDump.Format.JSON,
        )

        dump = ExternalApiDump.objects.get(
            api_type=ExternalApiDump.ExternalApiType.IDUS_ALL,
            format=ExternalApiDump.Format.JSON,
            include_sources=False,
        )
        self.assertEqual(dump.status, ExternalApiDump.Status.COMPLETED)
        with dump.dump_file.open("rb") as fp:
            payload = json.load(fp)

        expected = FigureReadOnlySerializer(get_idu_data({"include_sources": False}), many=True).data
        self.assertEqual(payload, json.loads(json.dumps(expected)))
        self.assertEqual(payload[0]["id"], self.figure.id)

    def test_geojson_dump_is_valid_with_expected_feature(self):
        doc = json.loads(b"".join(get_idu_data_geojson({"include_sources": False})))

        self.assertEqual(doc["type"], "FeatureCollection")
        self.assertEqual(len(doc["features"]), 1)
        feature = doc["features"][0]
        self.assertEqual(feature["geometry"]["type"], "MultiPoint")
        self.assertTrue(feature["geometry"]["coordinates"])
        self.assertEqual(feature["properties"]["id"], self.figure.id)

    def test_excel_dump_rows_match_serializer_output(self):
        wb = get_idu_data_excel({"include_sources": False})
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            wb.close()
            loaded = load_workbook(tmp.name, read_only=True)

        rows = list(loaded["IDUS_Data"].iter_rows(max_row=2, values_only=True))
        self.assertEqual(rows[0][0], "Id")
        expected = FigureReadOnlySerializer(next(iter(get_idu_data({"include_sources": False})))).data
        self.assertEqual(rows[1][0], self.figure.id)
        self.assertEqual(rows[1][1], expected["country"])


class TestIduExportSourceRouting(HelixAPITestCase):
    """The IDU endpoint must serve the dump that matches the client's
    'share_source' flag"""

    def setUp(self):
        super().setUp()
        self.url = "/external-api/idus/last-180-days/"
        self.client_with_sources = ClientFactory.create(code="share-on", is_active=True, share_source=True)
        self.client_without_sources = ClientFactory.create(code="share-off", is_active=True, share_source=False)
        self.dump_with_sources = self._create_dump(include_sources=True, filename="idus_with_sources.json")
        self.dump_without_sources = self._create_dump(include_sources=False, filename="idus_without_sources.json")

    def _create_dump(self, include_sources, filename):
        dump = ExternalApiDump.objects.create(
            api_type=ExternalApiDump.ExternalApiType.IDUS,
            format=ExternalApiDump.Format.JSON,
            include_sources=include_sources,
            status=ExternalApiDump.Status.COMPLETED,
        )
        dump.dump_file.save(filename, ContentFile(b"[]"), save=True)
        return dump

    def test_client_without_share_source_is_served_dump_without_sources(self):
        response = self.client.get(f"{self.url}?client_id={self.client_without_sources.code}")
        self.assertEqual(response.status_code, status.HTTP_302_FOUND)
        self.assertIn("idus_without_sources", response.url)
        self.assertNotIn("idus_with_sources", response.url)

    def test_client_with_share_source_is_served_dump_with_sources(self):
        response = self.client.get(f"{self.url}?client_id={self.client_with_sources.code}")
        self.assertEqual(response.status_code, status.HTTP_302_FOUND)
        self.assertIn("idus_with_sources", response.url)
        self.assertNotIn("idus_without_sources", response.url)


class TestIduGeojsonGeometry(HelixTestCase):
    """The IDU geojson routes are unauthenticated and read by GIS tooling outside
    Helix, which parses the geometry rather than the properties. A document that
    merely loads as JSON is not enough: the positions have to be numbers in
    [longitude, latitude] order, one per location, so these tests pin the
    geometry against the spec instead of against the serializer."""

    def setUp(self):
        super().setUp()
        self.country = CountryFactory.create(idmc_short_name="Nepal")
        self.event = EventFactory.create()
        self.entry = EntryFactory.create(url="https://example.com/source-report", is_confidential=False)

    def _create_figure(self, locations):
        return FigureFactory.create(
            entry=self.entry,
            country=self.country,
            event=self.event,
            category=Figure.FIGURE_CATEGORY_TYPES.NEW_DISPLACEMENT.value,
            role=Figure.ROLE.RECOMMENDED.value,
            include_idu=True,
            excerpt_idu="Some displacement happened.",
            geo_locations=locations,
        )

    def _assert_is_number(self, value):
        # `bool` is a subclass of `int` and would otherwise pass as a number
        self.assertNotIsInstance(value, bool)
        self.assertIsInstance(value, (int, float))

    def _assert_valid_geojson(self, doc):
        self.assertEqual(doc["type"], "FeatureCollection")
        for feature in doc["features"]:
            self.assertEqual(feature["type"], "Feature")
            geometry = feature["geometry"]
            self.assertEqual(geometry["type"], "MultiPoint")
            self.assertIsInstance(geometry["coordinates"], list)
            for position in geometry["coordinates"]:
                self.assertIsInstance(position, list)
                self.assertEqual(len(position), 2)
                longitude, latitude = position
                self._assert_is_number(longitude)
                self._assert_is_number(latitude)
                self.assertGreaterEqual(longitude, -180)
                self.assertLessEqual(longitude, 180)
                self.assertGreaterEqual(latitude, -90)
                self.assertLessEqual(latitude, 90)

    def _get_document(self):
        doc = json.loads(b"".join(get_idu_data_geojson({"include_sources": False})))
        self._assert_valid_geojson(doc)
        return doc

    def test_single_location_emits_one_numeric_longitude_latitude_position(self):
        self._create_figure([FigureLocationFactory.create(display_name="loc-a", lat=12.5, lon=145.75)])

        features = self._get_document()["features"]

        self.assertEqual(len(features), 1)
        self.assertEqual(features[0]["geometry"]["coordinates"], [[145.75, 12.5]])

    def test_multiple_locations_emit_one_position_each_without_interleaving(self):
        self._create_figure(
            [
                FigureLocationFactory.create(display_name="loc-a", lat=12.5, lon=145.75),
                FigureLocationFactory.create(display_name="loc-b", lat=-13.25, lon=46.5),
                FigureLocationFactory.create(display_name="loc-c", lat=0.5, lon=-72.25),
            ]
        )

        features = self._get_document()["features"]

        self.assertEqual(len(features), 1)
        coordinates = features[0]["geometry"]["coordinates"]
        self.assertEqual(len(coordinates), 3)
        # Aggregation order is not part of the contract, the set of positions is
        self.assertEqual(
            sorted(tuple(position) for position in coordinates),
            sorted([(145.75, 12.5), (46.5, -13.25), (-72.25, 0.5)]),
        )

    def test_figure_without_location_is_omitted_and_leaves_its_neighbours_intact(self):
        located = self._create_figure([FigureLocationFactory.create(display_name="loc-a", lat=12.5, lon=145.75)])
        unlocated = self._create_figure([])

        features = self._get_document()["features"]

        published_ids = [feature["properties"]["id"] for feature in features]
        self.assertEqual(published_ids, [located.id])
        self.assertNotIn(unlocated.id, published_ids)

    def test_figure_without_any_location_does_not_raise(self):
        self._create_figure([])

        self.assertEqual(self._get_document()["features"], [])
