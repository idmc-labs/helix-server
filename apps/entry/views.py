import typing
from pathlib import Path
from types import MappingProxyType
from typing import Optional

from django.contrib.postgres.aggregates.general import ArrayAgg, StringAgg
from django.contrib.postgres.fields import ArrayField
from django.db.models import Avg, Case, CharField, F, Func, Q, Value, When
from django.db.models.functions import Cast, Coalesce, Concat, ExtractYear, Lower
from django.db.models.sql.constants import LOUTER
from django.shortcuts import redirect
from django.utils import timezone
from django_cte import With
from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import extend_schema
from openpyxl import Workbook
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.viewsets import ViewSet

from apps.common.utils import (
    EXTERNAL_ARRAY_SEPARATOR,
    EXTERNAL_TUPLE_SEPARATOR,
    extract_event_code_data,
    extract_location_data,
)
from apps.crisis.models import Crisis
from apps.entry.models import ExternalApiDump, Figure
from apps.entry.serializers import FigureReadOnlySerializer
from apps.gidd.views import client_id
from helix.storages import TemporaryStorageEnableAuthString, get_external_storage
from utils.common import track_gidd
from utils.db import Array
from utils.serializers import make_flat_to_representation
from utils.streaming import stream_json_object_with_array

external_storage = get_external_storage()


def get_idu_data(filters=None):
    include_sources = False
    if filters:
        include_sources = filters.pop("include_sources", False)

    # The flat form joined FOUR to-many relations (geo_locations,
    # event__event_code, sources, entry__publishers) into one GROUP BY, so
    # every aggregate processed the cross-product of all four fan-outs.
    # Aggregating each relation in its own figure-grouped CTE keeps every scan
    # linear and drops the GROUP BY from the outer query entirely.
    idu_figures = Figure.objects.filter(
        category__in=[
            Figure.FIGURE_CATEGORY_TYPES.RETURNEES.value,
            Figure.FIGURE_CATEGORY_TYPES.NEW_DISPLACEMENT.value,
            Figure.FIGURE_CATEGORY_TYPES.CROSS_BORDER_RETURN.value,
            Figure.FIGURE_CATEGORY_TYPES.FAILED_RETURN_RETURNEE_DISPLACEMENT.value,
        ],
        excerpt_idu__isnull=False,
        include_idu=True,
    )
    geo_cte = With(
        idu_figures.values("id")
        .annotate(
            centroid_lat=Avg("geo_locations__lat"),
            centroid_lon=Avg("geo_locations__lon"),
            locations=ArrayAgg(
                Array(
                    F("geo_locations__display_name"),
                    Concat(
                        F("geo_locations__lat"),
                        Value(EXTERNAL_TUPLE_SEPARATOR),
                        F("geo_locations__lon"),
                        output_field=CharField(),
                    ),
                    Cast("geo_locations__accuracy", CharField()),
                    Cast("geo_locations__identifier", CharField()),
                    output_field=ArrayField(CharField()),
                ),
                distinct=True,
                filter=~Q(Q(geo_locations__display_name__isnull=True) | Q(geo_locations__display_name="")),
            ),
        )
        .values("id", "centroid_lat", "centroid_lon", "locations"),
        name="idu_geo_agg",
    )
    event_code_cte = With(
        idu_figures.values("id")
        .annotate(
            event_codes=ArrayAgg(
                Array(
                    F("event__event_code__event_code"),
                    Cast(F("event__event_code__event_code_type"), CharField()),
                    output_field=ArrayField(CharField()),
                ),
                distinct=True,
                filter=Q(event__event_code__country__id=F("country__id")),
            ),
        )
        .values("id", "event_codes"),
        name="idu_event_code_agg",
    )
    sources_cte = With(
        idu_figures.values("id")
        .annotate(
            sources_name=StringAgg("sources__name", EXTERNAL_ARRAY_SEPARATOR, distinct=True, output_field=CharField()),
        )
        .values("id", "sources_name"),
        name="idu_sources_agg",
    )
    publishers_cte = With(
        idu_figures.values("id")
        .annotate(
            publishers_name=StringAgg("entry__publishers__name", " ", distinct=True, output_field=CharField()),
        )
        .values("id", "publishers_name"),
        name="idu_publishers_agg",
    )

    base_query = (
        Figure.objects.annotate(
            displacement_date=Coalesce("end_date", "start_date"),
        )
        .filter(
            category__in=[
                Figure.FIGURE_CATEGORY_TYPES.RETURNEES.value,
                Figure.FIGURE_CATEGORY_TYPES.NEW_DISPLACEMENT.value,
                Figure.FIGURE_CATEGORY_TYPES.CROSS_BORDER_RETURN.value,
                Figure.FIGURE_CATEGORY_TYPES.FAILED_RETURN_RETURNEE_DISPLACEMENT.value,
            ],
            excerpt_idu__isnull=False,
            include_idu=True,
            entry__is_confidential=False,
        )
        .filter(
            ~(Q(entry__document_url__isnull=True) | Q(entry__document_url=""))
            | ~(Q(entry__url__isnull=True) | Q(entry__url=""))
        )
    )
    for relation_cte in (geo_cte, event_code_cte, sources_cte, publishers_cte):
        base_query = relation_cte.join(base_query, id=relation_cte.col.id, _join_type=LOUTER).with_cte(relation_cte)

    base_query = (
        base_query.annotate(
            country_name=F("country__idmc_short_name"),
            iso3=F("country__iso3"),
            figure_role=F("role"),
            centroid_lat=geo_cte.col.centroid_lat,
            centroid_lon=geo_cte.col.centroid_lon,
            locations=geo_cte.col.locations,
            event_codes=event_code_cte.col.event_codes,
            centroid=Case(
                When(
                    centroid_lat__isnull=False,
                    then=Concat(
                        Value("["),
                        F("centroid_lat"),
                        Value(EXTERNAL_TUPLE_SEPARATOR),
                        F("centroid_lon"),
                        Value("]"),
                        output_field=CharField(),
                    ),
                ),
                default=Value(""),
            ),
            displacement_start_date=F("start_date"),
            displacement_end_date=F("end_date"),
            year=Coalesce(ExtractYear("start_date", "year"), ExtractYear("end_date", "year")),
            event_name=F("event__name"),
            event_start_date=F("event__start_date"),
            event_end_date=F("event__end_date"),
            disaster_category_name=F("disaster_category__name"),
            disaster_sub_category_name=F("disaster_sub_category__name"),
            type_name=Case(
                When(figure_cause=Crisis.CRISIS_TYPE.CONFLICT, then=F("violence__name")),
                default=F("disaster_sub_type__type__name"),
                output_field=CharField(),
            ),
            subtype_name=Case(
                When(figure_cause=Crisis.CRISIS_TYPE.CONFLICT, then=F("violence_sub_type__name")),
                default=F("disaster_sub_type__name"),
                output_field=CharField(),
            ),
            figure_term_label=Case(
                When(term=0, then=Lower(Value(Figure.FIGURE_TERMS.EVACUATED.label))),
                When(term=1, then=Lower(Value(Figure.FIGURE_TERMS.DISPLACED.label))),
                When(term=2, then=Lower(Value(Figure.FIGURE_TERMS.FORCED_TO_FLEE.label))),
                When(term=3, then=Lower(Value(Figure.FIGURE_TERMS.RELOCATED.label))),
                When(term=4, then=Lower(Value(Figure.FIGURE_TERMS.SHELTERED.label))),
                When(term=5, then=Lower(Value(Figure.FIGURE_TERMS.IN_RELIEF_CAMP.label))),
                When(term=6, then=Lower(Value(Figure.FIGURE_TERMS.DESTROYED_HOUSING.label))),
                When(term=8, then=Lower(Value(Figure.FIGURE_TERMS.PARTIALLY_DESTROYED_HOUSING.label))),
                When(term=9, then=Lower(Value(Figure.FIGURE_TERMS.UNINHABITABLE_HOUSING.label))),
                When(term=10, then=Lower(Value(Figure.FIGURE_TERMS.HOMELESS.label))),
                When(term=11, then=Lower(Value(Figure.FIGURE_TERMS.RETURNS.label))),
                When(term=12, then=Lower(Value(Figure.FIGURE_TERMS.MULTIPLE_OR_OTHER.label))),
                output_field=CharField(),
            ),
            quantifier_label=Case(
                When(quantifier=0, then=Lower(Value(Figure.QUANTIFIER.MORE_THAN_OR_EQUAL.label))),
                When(quantifier=1, then=Lower(Value(Figure.QUANTIFIER.LESS_THAN_OR_EQUAL.label))),
                When(quantifier=2, then=Value("total")),
                When(quantifier=3, then=Lower(Value(Figure.QUANTIFIER.APPROXIMATELY.label))),
                output_field=CharField(),
            ),
            total_figures_text=Func(
                F("total_figures"), Value("999G999G999G990D"), function="to_char", output_field=CharField()
            ),
            sources_name=sources_cte.col.sources_name,
            displacement_occurred_transformed=Case(
                When(displacement_occurred=0, then=Value("Displacement reporting preventive evacuations")),
                When(
                    displacement_occurred__in=[1, 2, 3], then=Value("Displacement without preventive evacuations reported")
                ),
                output_field=CharField(),
            ),
            custom_figure_text=Case(
                When(
                    total_figures=1,
                    category=Figure.FIGURE_CATEGORY_TYPES.RETURN.value,
                    then=Concat(
                        F("country__idmc_short_name"),
                        Value(": "),
                        F("total_figures_text"),
                        Value(" return "),
                        Concat(Value("("), F("figure_term_label"), Value("),")),
                        Value(" "),
                        Func(F("start_date"), Value("DD Month"), function="to_char", output_field=CharField()),
                        Value(" - "),
                        Func(F("end_date"), Value("DD Month"), function="to_char", output_field=CharField()),
                        output_field=CharField(),
                    ),
                ),
                When(
                    (~Q(total_figures=1) & Q(category=Figure.FIGURE_CATEGORY_TYPES.RETURN.value)),
                    then=Concat(
                        F("country__idmc_short_name"),
                        Value(": "),
                        F("total_figures_text"),
                        Value(" returns "),
                        Concat(Value("("), F("figure_term_label"), Value("),")),
                        Value(" "),
                        Func(F("start_date"), Value("DD Month"), function="to_char", output_field=CharField()),
                        Value(" - "),
                        Func(F("end_date"), Value("DD Month"), function="to_char", output_field=CharField()),
                        output_field=CharField(),
                    ),
                ),
                When(
                    (Q(total_figures=1) & ~Q(term=Figure.FIGURE_TERMS.DISPLACED.value)),
                    then=Concat(
                        F("country__idmc_short_name"),
                        Value(": "),
                        F("total_figures_text"),
                        Value(" displacement "),
                        Concat(Value("("), F("figure_term_label"), Value("),")),
                        Value(" "),
                        Func(F("start_date"), Value("DD Month"), function="to_char", output_field=CharField()),
                        Value(" - "),
                        Func(F("end_date"), Value("DD Month"), function="to_char", output_field=CharField()),
                        output_field=CharField(),
                    ),
                ),
                When(
                    (
                        ~Q(total_figures=1)
                        & Q(
                            Q(term=Figure.FIGURE_TERMS.DISPLACED.value) | Q(term=Figure.FIGURE_TERMS.MULTIPLE_OR_OTHER.value)
                        )
                    ),
                    then=Concat(
                        F("country__idmc_short_name"),
                        Value(": "),
                        F("total_figures_text"),
                        Value(" displacements, "),
                        # THIS may be problematic
                        # Concat(Value('('), F('figure_term_label'), Value('),')),
                        # Value(' '),
                        Func(F("start_date"), Value("DD Month"), function="to_char", output_field=CharField()),
                        Value(" - "),
                        Func(F("end_date"), Value("DD Month"), function="to_char", output_field=CharField()),
                        output_field=CharField(),
                    ),
                ),
                When(
                    (~Q(total_figures=1) & ~Q(term=Figure.FIGURE_TERMS.DISPLACED.value)),
                    then=Concat(
                        F("country__idmc_short_name"),
                        Value(": "),
                        F("total_figures_text"),
                        Value(" displacements "),
                        Concat(Value("("), F("figure_term_label"), Value("),")),
                        Value(" "),
                        Func(F("start_date"), Value("DD Month"), function="to_char", output_field=CharField()),
                        Value(" - "),
                        Func(F("end_date"), Value("DD Month"), function="to_char", output_field=CharField()),
                        output_field=CharField(),
                    ),
                ),
            ),
            standard_info_text=Concat(
                Value("<b> "),
                F("custom_figure_text"),
                Value(" </b>"),
            ),
        )
        # `id` tiebreaker: (start_date, end_date) is not a total order, so the
        # dump row order would otherwise be plan-dependent.
        .order_by("-start_date", "-end_date", "id")
    )

    if not include_sources:
        base_query = base_query.annotate(
            entry_url_or_document_url=Value("", output_field=CharField()),
            custom_link_text=Concat(
                publishers_cte.col.publishers_name,
                Value(" - "),
                Func(F("entry__publish_date"), Value("DD Month YYYY"), function="to_char", output_field=CharField()),
                output_field=CharField(),
            ),
            standard_popup_text=Concat(
                Value("<b> "),
                F("custom_figure_text"),
                Value(" </b> <br> "),
                F("excerpt_idu"),
                Value(" <br> "),
                F("custom_link_text"),
                output_field=CharField(),
            ),
        )
    else:
        base_query = base_query.annotate(
            entry_url_or_document_url=Case(
                When(entry__document__isnull=False, then=F("entry__document_url")),
                When(entry__document__isnull=True, then=F("entry__url")),
                output_field=CharField(),
            ),
            custom_link_text=Concat(
                Value('<a href="'),
                Case(
                    When(entry__url__isnull=False, then=F("entry__url")),
                    When(entry__document_url__isnull=False, then=F("entry__document_url")),
                ),
                Value('"'),
                Value('target="_blank">'),
                publishers_cte.col.publishers_name,
                Value(" - "),
                Func(F("entry__publish_date"), Value("DD Month YYYY"), function="to_char", output_field=CharField()),
                Value("</a>"),
                output_field=CharField(),
            ),
            standard_popup_text=Concat(
                Value("<b> "),
                F("custom_figure_text"),
                Value(" </b> <br> "),
                F("excerpt_idu"),
                Value(" <br> "),
                F("custom_link_text"),
                output_field=CharField(),
            ),
        )

    # Apply filters if provided
    if filters:
        base_query = base_query.filter(**filters)

    # `.iterator()` streams rows from a server-side cursor instead of filling
    # the queryset result cache. NOTE: incompatible with `prefetch_related` on
    # Django 3.2 — adding one here would silently re-buffer.
    for figure_data in base_query.values().iterator(chunk_size=2000):
        # Raw CTE columns skip ArrayAgg/StringAgg's client-side NULL->[]/""
        # conversion — restore it here. From Django 5.0 that conversion lives in
        # the aggregate's `convert_value`, which a CTE-nested aggregate never
        # reaches: a `default=` on these aggregates would be silently dropped
        # rather than merely redundant, so this compensation is load-bearing.
        locations_data = figure_data.pop("locations", None) or []
        location_parse = extract_location_data(locations_data)

        event_codes_data = figure_data.pop("event_codes", None) or []
        event_code_parse = extract_event_code_data(event_codes_data)

        yield {
            **figure_data,
            "sources_name": figure_data["sources_name"] or "",
            "locations_name": location_parse["display_name"],
            "locations_coordinates": location_parse["lat_lon"],
            "locations_accuracy": location_parse["accuracy"],
            "locations_type": location_parse["type_of_points"],
            "event_codes": event_code_parse["code"],
            "event_code_types": event_code_parse["code_type"],
        }


def get_idu_export_field_descriptions() -> typing.List[str]:
    serializer = FigureReadOnlySerializer()
    return [f"{field.label}: {field.help_text}" for field in serializer.fields.values() if field.help_text]


def get_idu_data_excel(filters=None):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("IDUS_Data")
    ws.append(
        [
            "Id",
            "Country",
            "Iso3",
            "Latitude",
            "Longitude",
            "Centroid",
            "Role",
            "DisplacementType",
            "Qualifier",
            "Figure",
            "DisplacementDate",
            "DisplacementStartDate",
            "DisplacementEndDate",
            "Year",
            "EventId",
            "EventName",
            "EventCodes",
            "EventCodeTypes",
            "EventStartDate",
            "EventEndDate",
            "Category",
            "Subcategory",
            "Type",
            "Subtype",
            "StandardPopupText",
            "StandardInfoText",
            "OldId",
            "Sources",
            "SourceUrl",
            "LocationsName",
            "LocationsCoordinates",
            "LocationsAccuracy",
            "LocationsType",
            "DisplacementOccurred",
            "CreatedAt",
        ]
    )

    if filters:
        idu_data = get_idu_data(filters)
    else:
        idu_data = get_idu_data()

    serialize_record = make_flat_to_representation(FigureReadOnlySerializer())
    for obj in idu_data:
        item = serialize_record(obj)
        ws.append(
            [
                item["id"],
                item["country"],
                item["iso3"],
                item["latitude"],
                item["longitude"],
                item["centroid"],
                item["role"],
                item["displacement_type"],
                item["qualifier"],
                item["figure"],
                item["displacement_date"],
                item["displacement_start_date"],
                item["displacement_end_date"],
                item["year"],
                item["event_id"],
                item["event_name"],
                item["event_codes"],
                item["event_code_types"],
                item["event_start_date"],
                item["event_end_date"],
                item["category"],
                item["subcategory"],
                item["type"],
                item["subtype"],
                item["standard_popup_text"],
                item["standard_info_text"],
                item["old_id"],
                item["sources"],
                item["source_url"],
                item["locations_name"],
                item["locations_coordinates"],
                item["locations_accuracy"],
                item["locations_type"],
                item["displacement_occurred"],
                item["created_at"],
            ]
        )

    ws2 = wb.create_sheet("README")
    for description in get_idu_export_field_descriptions():
        ws2.append([description])

    return wb


def get_idu_data_geojson(filters=None):
    def format_coordinates(coordinates: typing.Optional[str]) -> typing.List[typing.List[float]]:
        """Turn the exported `locations_coordinates` string into GeoJSON
        positions. The export stores each location as "lat, lon" and joins
        locations with the array separator, while a GeoJSON position is
        [longitude, latitude] — so the pair is split and swapped. A pair that
        does not parse as two numbers is dropped rather than raised on, because
        this runs inside a streaming response that has already sent its
        headers."""
        positions = []
        for pair in (coordinates or "").split(EXTERNAL_ARRAY_SEPARATOR):
            lat, _, lon = pair.partition(EXTERNAL_TUPLE_SEPARATOR)
            try:
                positions.append([float(lon), float(lat)])
            except ValueError:
                continue
        return positions

    def remove_null_from_dict(data: dict) -> dict:
        return {key: value for key, value in data.items() if value is not None}

    if filters:
        idu_data = get_idu_data(filters)
    else:
        idu_data = get_idu_data()

    readme_text = "\n\n".join(get_idu_export_field_descriptions())
    scalar_fields = {
        "type": "FeatureCollection",
        "readme": readme_text,
        "lastUpdated": timezone.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    serialize_record = make_flat_to_representation(FigureReadOnlySerializer())

    def feature_iterator():
        for obj in idu_data:
            item = serialize_record(obj)

            coordinates = format_coordinates(item["locations_coordinates"])
            # A row with nothing mappable is omitted rather than published with an
            # empty or null geometry, matching the GIDD geojson export.
            if not coordinates:
                continue

            geometry = {
                "type": "MultiPoint",
                "coordinates": coordinates,
            }

            yield {
                "type": "Feature",
                "geometry": geometry,
                "properties": remove_null_from_dict(
                    {
                        "id": item["id"],
                        "country": item["country"],
                        "iso3": item["iso3"],
                        "latitude": item["latitude"],
                        "longitude": item["longitude"],
                        "centroid": item["centroid"],
                        "role": item["role"],
                        "displacement_type": item["displacement_type"],
                        "qualifier": item["qualifier"],
                        "figure": item["figure"],
                        "displacement_date": item["displacement_date"],
                        "displacement_start_date": item["displacement_start_date"],
                        "displacement_end_date": item["displacement_end_date"],
                        "year": item["year"],
                        "event_id": item["event_id"],
                        "event_name": item["event_name"],
                        "event_codes": item["event_codes"],
                        "event_code_types": item["event_code_types"],
                        "event_start_date": item["event_start_date"],
                        "event_end_date": item["event_end_date"],
                        "category": item["category"],
                        "subcategory": item["subcategory"],
                        "type": item["type"],
                        "subtype": item["subtype"],
                        "standard_popup_text": item["standard_popup_text"],
                        "standard_info_text": item["standard_info_text"],
                        "old_id": item["old_id"],
                        "sources": item["sources"],
                        "source_url": item["source_url"],
                        "locations_name": item["locations_name"],
                        "locations_coordinates": item["locations_coordinates"],
                        "locations_accuracy": item["locations_accuracy"],
                        "locations_type": item["locations_type"],
                        "displacement_occurred": item["displacement_occurred"],
                        "created_at": item["created_at"],
                    }
                ),
            }

    return stream_json_object_with_array(
        scalar_fields=scalar_fields,
        array_key="features",
        items=feature_iterator(),
    )


class FigureViewSet(viewsets.ReadOnlyModelViewSet):
    # TODO Add url for this viewset
    serializer_class = FigureReadOnlySerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        return get_idu_data()


class ExternalEndpointBaseCachedViewMixin:
    ENDPOINT_TYPE = None

    CONTENT_TYPES = MappingProxyType(
        {
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "json": "application/json",
            "geojson": "application/geo+json",
        }
    )

    def get_content_type(self, filename: str) -> Optional[str]:
        extension = Path(filename).suffix.lower().lstrip(".")
        return self.CONTENT_TYPES.get(extension)

    def build_download_params(self, filename: str) -> dict:
        params = {
            "ResponseContentDisposition": f'attachment; filename="{filename}"',
        }

        content_type = self.get_content_type(filename)
        if content_type:
            params["ResponseContentType"] = content_type

        return params

    def download_file(self, request, obj):
        filename = Path(obj.dump_file.name).name
        params = self.build_download_params(filename)
        with TemporaryStorageEnableAuthString(external_storage):
            url = external_storage.url(
                obj.dump_file.name,
                parameters=params,
            )

        return redirect(url)

    # The IDU dumps are generated both with and without sources, and a client's `share_source`
    # flag selects which it may read. An endpoint generated only one way must not key the lookup on
    # that flag, or every client with `share_source` set gets a 404 for a dump that exists.
    DUMP_VARIES_BY_CLIENT = True

    @client_id
    def get(self, request, data_format):
        # Check if request is comming from valid client
        client_id = request.GET.get("client_id", None)
        # Track client. Use a format-specific api_type so excel/geojson hits are
        # counted separately from json (which falls back to the base type).
        tracking_type = ExternalApiDump.TRACKING_API_TYPE.get(
            (self.ENDPOINT_TYPE, data_format),
            self.ENDPOINT_TYPE,
        )
        client = track_gidd(
            client_id,
            tracking_type,
        )
        # NOTE: Sending empty array so client don't break.
        _empty_response = []
        try:
            api_dump = ExternalApiDump.objects.get(
                api_type=self.ENDPOINT_TYPE,
                include_sources=client.share_source if self.DUMP_VARIES_BY_CLIENT else False,
                format=data_format,
            )
        except ExternalApiDump.DoesNotExist:
            return Response(_empty_response, status=status.HTTP_404_NOT_FOUND)

        if api_dump.status == ExternalApiDump.Status.COMPLETED:
            return self.download_file(request, api_dump)

        if api_dump.status == ExternalApiDump.Status.FAILED:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Finally, for pending. If we have a dump send it
        if api_dump.dump_file.name is not None:
            return self.download_file(request, api_dump)

        # Else send 202 response
        return Response(_empty_response, status=status.HTTP_202_ACCEPTED)


@extend_schema(
    description=Path("docs/idus/main-description.md").read_text(),
    responses=FigureReadOnlySerializer,
    tags=["IDU"],
)
class BaseIdusCachedView(ExternalEndpointBaseCachedViewMixin, ViewSet):
    @client_id
    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        return self._export(request, ExternalApiDump.Format.EXCEL)

    @client_id
    @action(detail=False, methods=["get"], url_path="export-json")
    def export_json(self, request):
        return self._export(request, ExternalApiDump.Format.JSON)

    @client_id
    @action(detail=False, methods=["get"], url_path="export-geojson")
    def export_geojson(self, request):
        return self._export(request, ExternalApiDump.Format.GEOJSON)

    def _export(self, request, fmt):
        return super().get(request, data_format=fmt)


class IdusFlatCachedView(BaseIdusCachedView):
    ENDPOINT_TYPE = ExternalApiDump.ExternalApiType.IDUS


class IdusAllFlatCachedView(BaseIdusCachedView):
    ENDPOINT_TYPE = ExternalApiDump.ExternalApiType.IDUS_ALL


class IdusAllDisasterCachedView(BaseIdusCachedView):
    ENDPOINT_TYPE = ExternalApiDump.ExternalApiType.IDUS_ALL_DISASTER


@extend_schema(
    description=(
        "The reference lists the IDU feeds are keyed against: disaster and violence typologies, "
        "geographical groups, and countries with their bounding boxes. Answers with a redirect to "
        "the generated dump, as the other IDU endpoints do."
    ),
    responses={(302, "application/json"): OpenApiTypes.STR},
    tags=["IDU"],
)
class IduReferencesCachedView(ExternalEndpointBaseCachedViewMixin, ViewSet):
    ENDPOINT_TYPE = ExternalApiDump.ExternalApiType.IDU_REFERENCES
    # One dump serves every client: the references carry no source information to withhold.
    DUMP_VARIES_BY_CLIENT = False

    @client_id
    @action(detail=False, methods=["get"], url_path="export-json")
    def export_json(self, request):
        return super().get(request, data_format=ExternalApiDump.Format.JSON)
