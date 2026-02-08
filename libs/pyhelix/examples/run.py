# /// script
# requires-python = "~=3.8.0"
# dependencies = [
#   "pydantic-settings",
#   "pyhelix",
#   "openpyxl",
#   "colorlog",
# ]
#
# [tool.uv.sources]
# pyhelix = { path = "../", editable = true }
#
# ///

from __future__ import annotations

import json
import logging
import logging.config
import pathlib
import typing
import uuid

import typing_extensions
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError
from pydantic_settings import BaseSettings, SettingsConfigDict

from pyhelix.api.api import HelixClient, HelixEndpoint
from pyhelix.constants import CRISIS_TYPE
from pyhelix.hulk import HulkDataHandler
from pyhelix.models import (
    HulkAttachmentImport,
    HulkEntryImport,
    HulkEntryImportTypeEnum,
    HulkEventImport,
    HulkFigureImport,
    HulkFigureImportLocation,
    HulkSourcePreviewImport,
)
from pyhelix.parsers import validate_and_parse_enum

logger = logging.getLogger(__name__)


BASE_DIR = pathlib.Path(__file__).parent
IMPORT_DATASET_DIR = BASE_DIR / "dataset"
OUTPUT_DATASET_DIR = BASE_DIR / "generated"
OUTPUT_DATASET_DIR.mkdir(exist_ok=True)

ENTRY_RAW_FILE = IMPORT_DATASET_DIR / "entry.json"
EVENTS_RAW_FILE = IMPORT_DATASET_DIR / "events.xlsx"
FIGURES_RAW_FILE = IMPORT_DATASET_DIR / "figures.xlsx"


LOGGING_CONFIG = {
    "version": 1,
    "disable_existing_loggers": False,
    "formatters": {
        "colored": {
            "()": "colorlog.ColoredFormatter",
            "format": "%(log_color)s[%(levelname)s]%(reset)s %(red)s%(name)-8s%(reset)s %(message)s",
            "log_colors": {
                "DEBUG": "cyan",
                "INFO": "green",
                "WARNING": "yellow",
                "ERROR": "red",
                "CRITICAL": "bold_red",
            },
        },
    },
    "handlers": {
        "console": {
            "class": "colorlog.StreamHandler",
            "formatter": "colored",
            "level": "INFO",
        }
    },
    "root": {
        "handlers": ["console"],
        "level": "INFO",
    },
}


logging.config.dictConfig(LOGGING_CONFIG)


class Settings(BaseSettings):
    model_config = SettingsConfigDict(env_file=BASE_DIR / ".env", env_file_encoding="utf-8")

    HELIX_BASE_DOMAIN: str = HelixEndpoint.BaseDomain.PRODUCTION
    HELIX_EMAIL: str
    HELIX_PASSWORD: str


def read_from_sheet(sheet: Worksheet, header_row=1, max_col=None) -> typing.Generator[dict[typing.Any, typing.Any]]:
    iter_rows = sheet.iter_rows(min_row=header_row, max_col=max_col, values_only=True)
    headers = next(iter_rows)
    for row in iter_rows:
        yield {header: row[i] for i, header in enumerate(headers)}


class ImportContext(typing.NamedTuple):
    helix_client: HelixClient
    entries_map: dict[str, uuid.UUID] = {}
    events_map: dict[str, typing.Tuple[uuid.UUID, CRISIS_TYPE]] = {}


class MainTriggerResult(typing.TypedDict):
    violence_sub_type_id: int | None
    disaster_sub_type_id: int | None
    other_sub_type_id: int | None


def parse_main_trigger(context: ImportContext, cause: CRISIS_TYPE, main_trigger) -> MainTriggerResult:
    violence_sub_type_id = None
    disaster_sub_type_id = None
    other_sub_type_id = None

    if cause == CRISIS_TYPE.CONFLICT:
        violence_sub_type_id = context.helix_client.violence_sub_type_manager.search(main_trigger)
    elif cause == CRISIS_TYPE.DISASTER:
        disaster_sub_type_id = context.helix_client.disaster_sub_type_manager.search(main_trigger)
    elif cause == CRISIS_TYPE.OTHER:
        other_sub_type_id = context.helix_client.other_sub_type_manager.search(main_trigger)
    else:
        typing_extensions.assert_never(cause)

    return {
        "violence_sub_type_id": violence_sub_type_id,
        "disaster_sub_type_id": disaster_sub_type_id,
        "other_sub_type_id": other_sub_type_id,
    }


def import_entry(hulk_handler: HulkDataHandler, context: ImportContext):
    with ENTRY_RAW_FILE.open("r") as fp:
        entry_data = json.load(fp)

    for hulk_import_type in [
        HulkEntryImportTypeEnum.URL,
        HulkEntryImportTypeEnum.DOCUMENT,
    ]:
        publishers_id = []
        for publisher in entry_data["publishers"].split(","):
            publisher_id = context.helix_client.organization_manager.search(publisher)
            if publisher_id:
                publishers_id.append(publisher_id)
                continue
            # TODO: Do we treat this as error
            logger.warning("Organization not found for <%s>", publisher)

        source_preview_uuid = None
        attachment_uuid = None
        if hulk_import_type == HulkEntryImportTypeEnum.URL:
            # TODO: Use separate uuid for source_preview to avoid conflict?
            source_preview_uuid = entry_data["uuid"]
            try:
                source_preview_import = HulkSourcePreviewImport(
                    uuid=source_preview_uuid,
                    file_url=entry_data["url"],
                )
                hulk_handler.handle_import_object(source_preview_import)
            except ValidationError as e:
                hulk_handler.handle_import_error(HulkSourcePreviewImport, e)
                continue
        else:
            # TODO: Use separate uuid for source_preview to avoid conflict?
            attachment_uuid = entry_data["uuid"]
            try:
                attachment_import = HulkAttachmentImport(
                    uuid=attachment_uuid,
                    file_url=entry_data["url"],  # XXX: This should be a s3 url
                )
                hulk_handler.handle_import_object(attachment_import)
            except ValidationError as e:
                hulk_handler.handle_import_error(HulkAttachmentImport, e)
                return

        try:
            entry_import = HulkEntryImport(
                uuid=entry_data["uuid"],
                hulk_import_type=hulk_import_type,
                source_preview_uuid=source_preview_uuid,
                attachment_uuid=attachment_uuid,
                entry_title=entry_data["entry_title"],
                is_confidential=entry_data["confidential_source"] == 1,
                publish_date=entry_data["publication_date"],
                publishers_id=publishers_id,
                extra_context={
                    "publishers": entry_data["publishers"],
                },
            )
            hulk_handler.handle_import_object(entry_import)
        except ValidationError as e:
            hulk_handler.handle_import_error(HulkEntryImport, e)
            return

        context.entries_map.update(
            {
                entry_import.entry_title: entry_import.uuid,
            }
        )


def import_events(hulk_handler: HulkDataHandler, context: ImportContext):
    with EVENTS_RAW_FILE.open("rb") as payload:
        workbook = load_workbook(payload, data_only=True, read_only=True)

        sheet = workbook.active
        assert sheet is not None, f"There is no sheet in {EVENTS_RAW_FILE}"

        for event_data in read_from_sheet(sheet):
            event_cause = validate_and_parse_enum(CRISIS_TYPE, event_data["event_cause"])

            if event_cause is None:
                logger.error("Event cause should be valid: %s", event_cause)
                continue

            main_trigger_metadata = parse_main_trigger(context, event_cause, event_data["main_trigger"])

            countries_id = []
            for country in event_data["countries"].split(","):
                country_id = context.helix_client.country_manager.search(country)
                if country_id:
                    countries_id.append(country_id)
                    continue
                # TODO: Treat this as error?
                logger.warning("Country not found for <%s>", country)

            try:
                event_import = HulkEventImport(
                    uuid=event_data["uuid"],
                    event_name=event_data["event_name"],
                    event_cause=event_cause,
                    violence_sub_type_id=main_trigger_metadata["violence_sub_type_id"],
                    disaster_sub_type_id=main_trigger_metadata["disaster_sub_type_id"],
                    other_sub_type_id=main_trigger_metadata["other_sub_type_id"],
                    start_date=event_data["start_date"],
                    start_date_accuracy=event_data["start_date_accuracy"],
                    end_date=event_data["end_date"],
                    end_date_accuracy=event_data["end_date_accuracy"],
                    event_narrative=event_data["event_narrative"],
                    countries_id=countries_id,
                    event_codes=[],  # TODO
                    extra_context={
                        "countries": event_data["countries"],
                        "main_trigger": event_data["main_trigger"],
                    },
                )
                hulk_handler.handle_import_object(event_import)
            except ValidationError as e:
                hulk_handler.handle_import_error(HulkEventImport, e)
                continue

            context.events_map.update(
                {
                    event_import.event_name: (
                        event_import.uuid,
                        event_cause,
                    ),
                }
            )


def import_figures(hulk_handler: HulkDataHandler, context: ImportContext):
    with FIGURES_RAW_FILE.open("rb") as payload:
        workbook = load_workbook(payload, data_only=True, read_only=True)

        sheet = workbook.active
        assert sheet is not None, f"There is no sheet in {FIGURES_RAW_FILE}"

        entry_uuid = list(context.entries_map.values())[0]

        assert entry_uuid is not None, "entry not found"

        for figure_data in read_from_sheet(sheet):
            figure_uuid = figure_data["uuid"]

            event_raw = figure_data["event"]
            event_metadata = context.events_map.get(event_raw)
            if event_metadata is None:
                hulk_handler.handle_import_error_raw(
                    HulkFigureImport,
                    {
                        "uuid": figure_uuid,
                        "error": f"event_uuid not found for: {event_raw}",
                    },
                )
                continue

            figure_cause_raw = figure_data["figure_cause"]
            figure_cause = validate_and_parse_enum(CRISIS_TYPE, figure_cause_raw)
            if figure_cause is None:
                hulk_handler.handle_import_error_raw(
                    HulkFigureImport,
                    {
                        "uuid": figure_uuid,
                        "error": f"figure_cause not found for: {figure_cause_raw}",
                    },
                )
                continue

            main_trigger_metadata = parse_main_trigger(context, figure_cause, figure_data["main_trigger_figure"])

            country_raw = figure_data["country"]
            country_id = context.helix_client.country_manager.search(country_raw)
            if country_id is None:
                hulk_handler.handle_import_error_raw(
                    HulkFigureImport,
                    {
                        "uuid": figure_uuid,
                        "error": f"country not found for: {country_raw}",
                    },
                )
                continue

            organizations_raw = json.loads(figure_data["sources"])
            organizations_id = []
            for organization_raw in organizations_raw:
                organization_id = context.helix_client.organization_manager.search(
                    organization_raw.replace("- Peru", "")  # TODO: Fix this in dataset
                )
                if organization_id:
                    organizations_id.append(organization_id)
                    continue
                # TODO: Treat this as error?
                logger.warning("Organization not found for <%s>", organization_raw)

            figure_tags_raw = json.loads(figure_data["tags"])
            figure_tags_id = []
            for figure_tag_raw in figure_tags_raw:
                figure_tag_id = context.helix_client.figure_tag_manager.search(figure_tag_raw)
                if figure_tag_id:
                    figure_tags_id.append(figure_tag_id)
                    continue
                # TODO: Treat this as error?
                logger.warning("FigureTag not found for <%s>", figure_tag_raw)

            try:
                figure_locations = [
                    HulkFigureImportLocation(
                        uuid=loc_uuid,
                        bounding_box=None,
                        display_name=loc_display_name,
                        country_name=country_raw,
                        country_code="TEST",  # TODO
                        identifier=loc_identifier,
                        accuracy=loc_accuracy,
                        geocoder=loc_geocoder,
                        latitude=loc_latitude,
                        longitude=loc_longitude,
                    )
                    for (
                        loc_uuid,
                        loc_display_name,
                        loc_identifier,
                        loc_accuracy,
                        _,
                        loc_latitude,
                        loc_longitude,
                        loc_geocoder,
                    ) in json.loads(figure_data["location"])
                ]

                figure_import = HulkFigureImport(
                    uuid=figure_uuid,
                    entry_uuid=entry_uuid,
                    event_uuid=event_metadata[0],
                    figure_cause=figure_cause,
                    violence_sub_type_id=main_trigger_metadata["violence_sub_type_id"],
                    context_of_violences_id=[],  # TODO
                    osv_sub_type_id=None,  # TODO
                    disaster_sub_type_id=main_trigger_metadata["disaster_sub_type_id"],
                    other_sub_type_id=main_trigger_metadata["other_sub_type_id"],
                    category=figure_data["category"],
                    term=figure_data["term"],
                    quantifier=figure_data["quantifier"],
                    unit=figure_data["unit"],
                    figure_role=figure_data["figure_role"],
                    country_id=country_id,
                    start_date=figure_data["start_date"],
                    start_date_accuracy=figure_data["start_date_accuracy"],
                    end_date=figure_data["end_date"],
                    end_date_accuracy=figure_data["end_date_accuracy"],
                    stock_date=figure_data["stock_date"],
                    stock_date_accuracy=figure_data["stock_date_accuracy"],
                    stock_reporting_date=figure_data["stock_reporting_date"],
                    reported_figure=figure_data["reported_figure"],
                    is_housing_destruction=int(figure_data["housing_destruction_toggle"] or 0) == 1,
                    household_size=figure_data["household_size"],
                    displacement_occurred=figure_data["displacement_occured"],
                    is_disaggregated=figure_data["disaggregation_available"],
                    analysis_text=figure_data["analysis_text"],
                    source_excerpt_text=figure_data["source_excerpt_text"],
                    include_idu=figure_data["include_IDU"],
                    idu_text=figure_data["idu_text"],
                    tags_id=figure_tags_id,
                    sources_id=organizations_id,
                    locations=figure_locations,
                    extra_context={
                        "main_trigger": figure_data["main_trigger_figure"],
                    },
                )
                hulk_handler.handle_import_object(figure_import)
            except ValidationError as e:
                hulk_handler.handle_import_error(HulkFigureImport, e)
                continue


def main():
    settings = Settings()

    helix_client = HelixClient(
        endpoint=HelixEndpoint(
            base_domain=settings.HELIX_BASE_DOMAIN,
            email=settings.HELIX_EMAIL,
            password=settings.HELIX_PASSWORD,
        ),
    )

    with HulkDataHandler(export_dir=OUTPUT_DATASET_DIR, helix_client=helix_client) as hulk_handler:
        context = ImportContext(helix_client=helix_client)

        data_generators: list[typing.Callable[[HulkDataHandler, ImportContext],]] = [
            # Order matters
            import_entry,
            import_events,
            import_figures,
        ]
        for data_generator in data_generators:
            data_generator(hulk_handler, context)

        # NOTE: Just for verbose information about the outputs
        print(json.dumps(hulk_handler.debug_metadata(), indent=2))


main()
